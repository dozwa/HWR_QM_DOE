"""Kommandozeilen-Interface fuer den Statapult-Simulator.

Subcommands:
  shoot   - Einzelschuss oder Wiederholungen
  batch   - Batch-Modus aus CSV-Datei
  msa     - MSA-Daten generieren (fuer Gage R&R)
  control - Regelkarten-Daten generieren
  info    - Faktor-Uebersicht anzeigen
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import List, Optional

from .config import CatapultConfig
from .factors import ALL_FACTORS, STANDARD_FACTORS, factors_info, get_defaults
from .io_utils import (
    format_multiple_results_text,
    format_result_json,
    format_result_text,
    read_csv,
    write_csv,
)
from .simulator import Statapult


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="statapult",
        description="Virtueller Statapult-Simulator fuer Six Sigma DOE-Uebungen.",
    )
    subparsers = parser.add_subparsers(dest="command", help="Verfuegbare Befehle")

    # --- shoot ---
    shoot_p = subparsers.add_parser("shoot", help="Einzelschuss durchfuehren")
    _add_factor_args(shoot_p)
    shoot_p.add_argument("--seed", type=int, default=None, help="Zufalls-Seed")
    shoot_p.add_argument("--repeat", type=int, default=1, help="Anzahl Wiederholungen")
    shoot_p.add_argument(
        "--noise-level", type=float, default=1.0,
        help="Rausch-Multiplikator (0=deterministisch)",
    )
    shoot_p.add_argument(
        "--format", choices=["text", "csv", "json"], default="text",
        dest="output_format", help="Ausgabeformat",
    )
    shoot_p.add_argument("--verbose", "-v", action="store_true", help="Ausfuehrliche Ausgabe")
    shoot_p.add_argument("--config", type=str, default=None, help="YAML-Konfigurationsdatei")
    shoot_p.add_argument("--drag", action="store_true", help="Luftwiderstand aktivieren")

    # --- batch ---
    batch_p = subparsers.add_parser("batch", help="Batch-Modus aus CSV-Datei")
    batch_p.add_argument("-i", "--input", required=True, help="Eingabe-CSV")
    batch_p.add_argument("-o", "--output", default=None, help="Ausgabe-CSV (sonst stdout)")
    batch_p.add_argument("--seed", type=int, default=None, help="Zufalls-Seed")
    batch_p.add_argument("--noise-level", type=float, default=1.0, help="Rausch-Multiplikator")
    batch_p.add_argument("--config", type=str, default=None, help="YAML-Konfigurationsdatei")
    batch_p.add_argument("--drag", action="store_true", help="Luftwiderstand aktivieren")

    # --- msa ---
    msa_p = subparsers.add_parser("msa", help="MSA-Daten generieren")
    _add_factor_args(msa_p)
    msa_p.add_argument("--operators", type=int, default=3, help="Anzahl Operatoren")
    msa_p.add_argument(
        "--measurements-per-operator", type=int, default=10,
        help="Messungen pro Operator",
    )
    msa_p.add_argument("-o", "--output", default=None, help="Ausgabe-CSV")
    msa_p.add_argument("--seed", type=int, default=None, help="Zufalls-Seed")
    msa_p.add_argument("--config", type=str, default=None, help="YAML-Konfigurationsdatei")

    # --- control ---
    ctrl_p = subparsers.add_parser("control", help="Regelkarten-Daten generieren")
    _add_factor_args(ctrl_p)
    ctrl_p.add_argument("--shots", type=int, default=25, help="Anzahl Schuesse")
    ctrl_p.add_argument(
        "--drift", type=float, default=0.0,
        help="Drift-Rate (cm pro Schuss)",
    )
    ctrl_p.add_argument("-o", "--output", default=None, help="Ausgabe-CSV")
    ctrl_p.add_argument("--seed", type=int, default=None, help="Zufalls-Seed")
    ctrl_p.add_argument("--config", type=str, default=None, help="YAML-Konfigurationsdatei")

    # --- info ---
    subparsers.add_parser("info", help="Faktor-Uebersicht anzeigen")

    # --- fill ---
    fill_p = subparsers.add_parser(
        "fill",
        help="Excel-Vorlagen des Notebooks mit simulierten Daten befuellen",
    )
    fill_p.add_argument(
        "file", type=str,
        help="Pfad zur Excel-Vorlage (MSA, DoE oder Konfirmation)",
    )
    fill_p.add_argument("-o", "--output", default=None, help="Ausgabe-Pfad (sonst ueberschreibt Eingabe)")
    fill_p.add_argument("--seed", type=int, default=None, help="Zufalls-Seed")
    fill_p.add_argument("--noise-level", type=float, default=1.0, help="Rausch-Multiplikator")
    _add_factor_args(fill_p)

    return parser


def _add_factor_args(parser: argparse.ArgumentParser) -> None:
    """Fuegt Faktor-Argumente zu einem Subparser hinzu."""
    for key, f in ALL_FACTORS.items():
        cli_name = f"--{key.replace('_', '-')}"
        parser.add_argument(
            cli_name, type=float, default=None,
            metavar=f.einheit.upper(),
            help=f"{f.name} (Default: {f.default}, Range: {f.low}-{f.high})",
        )


def _get_settings(args: argparse.Namespace) -> dict:
    """Extrahiert Faktor-Settings aus geparsten Argumenten."""
    settings = {}
    for key in ALL_FACTORS:
        attr = key  # argparse konvertiert bindestriche zu unterstrichen
        value = getattr(args, attr, None)
        if value is not None:
            settings[key] = value
    return settings


def _load_config(args: argparse.Namespace) -> CatapultConfig:
    """Laedt Konfiguration aus args.config oder Default."""
    config_path = getattr(args, "config", None)
    if config_path:
        config = CatapultConfig.from_yaml(config_path)
    else:
        config = CatapultConfig.default()

    if getattr(args, "drag", False):
        config.enable_drag = True

    return config


def cmd_shoot(args: argparse.Namespace) -> None:
    config = _load_config(args)
    katapult = Statapult(config=config, seed=args.seed)
    settings = _get_settings(args)

    if args.repeat == 1:
        result = katapult.shoot(settings, noise_level=args.noise_level)
        if args.output_format == "json":
            print(format_result_json(result))
        elif args.output_format == "csv":
            print(write_csv([{"Wurfweite_cm": round(result.wurfweite_cm, 1)}]))
        else:
            print(format_result_text(result, verbose=args.verbose))
    else:
        results = katapult.shoot_multiple(
            settings, n=args.repeat, noise_level=args.noise_level
        )
        if args.output_format == "json":
            data = [r.to_dict() for r in results]
            print(json.dumps(data, indent=2, ensure_ascii=False))
        elif args.output_format == "csv":
            rows = [
                {"Wurf": i + 1, "Wurfweite_cm": round(r.wurfweite_cm, 1)}
                for i, r in enumerate(results)
            ]
            print(write_csv(rows))
        else:
            print(format_multiple_results_text(results))


def cmd_batch(args: argparse.Namespace) -> None:
    config = _load_config(args)
    katapult = Statapult(config=config, seed=args.seed)

    rows = read_csv(args.input)
    results = []
    for i, row_settings in enumerate(rows, 1):
        # Konvertiere alle Werte zu float wo moeglich
        settings = {}
        for k, v in row_settings.items():
            if isinstance(v, (int, float)):
                settings[k] = float(v)
        result = katapult.shoot(settings, noise_level=args.noise_level)
        row_out = dict(row_settings)
        row_out["Ergebnis"] = round(result.wurfweite_cm, 1)
        results.append(row_out)
        print(f"  Versuch {i:>3d}: {result.wurfweite_cm:>8.1f} cm", file=sys.stderr)

    if args.output:
        write_csv(results, args.output)
        print(f"\nErgebnisse gespeichert: {args.output}", file=sys.stderr)
    else:
        print(write_csv(results))


def cmd_msa(args: argparse.Namespace) -> None:
    config = _load_config(args)
    katapult = Statapult(config=config, seed=args.seed)
    settings = _get_settings(args)

    rows = []
    for op in range(1, args.operators + 1):
        operator_id = f"Operator_{op}"
        for m in range(1, args.measurements_per_operator + 1):
            result = katapult.shoot(
                settings, noise_level=1.0, operator_id=operator_id
            )
            rows.append({
                "Operator": operator_id,
                "Messung": m,
                "Wurfweite_cm": round(result.wurfweite_cm, 1),
            })

    if args.output:
        write_csv(rows, args.output)
        print(f"MSA-Daten gespeichert: {args.output} ({len(rows)} Messungen)")
    else:
        print(write_csv(rows))


def cmd_control(args: argparse.Namespace) -> None:
    config = _load_config(args)
    config.noise.drift_rate = args.drift
    katapult = Statapult(config=config, seed=args.seed)
    settings = _get_settings(args)

    rows = []
    for i in range(1, args.shots + 1):
        result = katapult.shoot(settings, noise_level=1.0)
        rows.append({
            "Schuss": i,
            "Wurfweite_cm": round(result.wurfweite_cm, 1),
        })

    if args.output:
        write_csv(rows, args.output)
        print(f"Regelkarten-Daten gespeichert: {args.output} ({len(rows)} Schuesse)")
    else:
        print(write_csv(rows))


def cmd_info(_args: argparse.Namespace) -> None:
    print(factors_info())


def cmd_fill(args: argparse.Namespace) -> None:
    from .excel_filler import fill_doe, fill_konfirmation, fill_msa

    filepath = args.file
    if not Path(filepath).exists():
        print(f"Fehler: Datei nicht gefunden: {filepath}", file=sys.stderr)
        sys.exit(1)

    settings = _get_settings(args) or None
    output = args.output

    # Auto-detect Vorlage anhand Dateiname oder Inhalt
    name_lower = Path(filepath).name.lower()
    if "msa" in name_lower:
        result = fill_msa(
            filepath, output_path=output, settings=settings,
            seed=args.seed, noise_level=args.noise_level,
        )
        print(f"MSA-Vorlage befuellt: {result}")
    elif "doe" in name_lower or "versuch" in name_lower:
        result = fill_doe(
            filepath, output_path=output,
            seed=args.seed, noise_level=args.noise_level,
        )
        print(f"DoE-Vorlage befuellt: {result}")
    elif "konfirm" in name_lower:
        result = fill_konfirmation(
            filepath, output_path=output, settings=settings,
            seed=args.seed, noise_level=args.noise_level,
        )
        print(f"Konfirmations-Vorlage befuellt: {result}")
    else:
        # Versuche Inhalt zu erkennen
        from openpyxl import load_workbook
        wb = load_workbook(filepath, read_only=True)
        sheets = [s.lower() for s in wb.sheetnames]
        wb.close()

        if "type-1" in sheets or "reproduzierbarkeit" in sheets:
            result = fill_msa(
                filepath, output_path=output, settings=settings,
                seed=args.seed, noise_level=args.noise_level,
            )
            print(f"MSA-Vorlage befuellt: {result}")
        elif "versuchsergebnisse" in sheets:
            result = fill_doe(
                filepath, output_path=output,
                seed=args.seed, noise_level=args.noise_level,
            )
            print(f"DoE-Vorlage befuellt: {result}")
        elif "konfirmation" in sheets:
            result = fill_konfirmation(
                filepath, output_path=output, settings=settings,
                seed=args.seed, noise_level=args.noise_level,
            )
            print(f"Konfirmations-Vorlage befuellt: {result}")
        else:
            print(f"Fehler: Konnte Vorlagen-Typ nicht erkennen. Sheets: {wb.sheetnames}", file=sys.stderr)
            sys.exit(1)


def main(argv: Optional[List[str]] = None) -> None:
    parser = build_parser()
    args = parser.parse_args(argv)

    if args.command is None:
        parser.print_help()
        sys.exit(0)

    commands = {
        "shoot": cmd_shoot,
        "batch": cmd_batch,
        "msa": cmd_msa,
        "control": cmd_control,
        "info": cmd_info,
        "fill": cmd_fill,
    }
    commands[args.command](args)
