"""Befuellt die Excel-Vorlagen des DMAIC-Notebooks mit simulierten Daten.

Unterstuetzte Vorlagen:
- MSA_Messung_Template.xlsx (Sheets: Type-1, Reproduzierbarkeit)
- DoE_Versuchsergebnisse.xlsx (Sheet: Versuchsergebnisse)
- Konfirmation.xlsx (Sheet: Konfirmation)
"""

from __future__ import annotations

import os
from copy import copy
from pathlib import Path
from typing import Dict, List, Optional

import numpy as np

from .factors import ALL_FACTORS, validate_settings
from .simulator import Statapult


def fill_msa(
    input_path: str,
    output_path: Optional[str] = None,
    settings: Optional[Dict[str, float]] = None,
    referenzwert: Optional[float] = None,
    seed: Optional[int] = None,
    noise_level: float = 1.0,
) -> str:
    """Befuellt eine MSA_Messung_Template.xlsx mit simulierten Messdaten.

    Parameters
    ----------
    input_path : str
        Pfad zur leeren MSA-Vorlage (von helper.generate_msa_template()).
    output_path : str, optional
        Pfad fuer die befuellte Datei. Default: ueberschreibt input_path.
    settings : dict, optional
        Katapult-Einstellungen. Default: Standard-Einstellungen.
    referenzwert : float, optional
        Referenzwert fuer Type-1. Default: wahrer Wert bei gegebenen Settings.
    seed : int, optional
        Zufalls-Seed.
    noise_level : float
        Rausch-Multiplikator.

    Returns
    -------
    str
        Pfad zur befuellten Datei.
    """
    from openpyxl import load_workbook

    if output_path is None:
        output_path = input_path

    wb = load_workbook(input_path)
    katapult = Statapult(seed=seed)
    settings = validate_settings(settings or {})

    # Wahren Wert berechnen fuer Referenz
    true_distance = katapult.shoot(settings, noise_level=0).wurfweite_cm
    if referenzwert is None:
        referenzwert = round(true_distance, 1)

    # --- Sheet: Type-1 ---
    ws_t1 = wb["Type-1"]
    header_row, num_personen = _find_msa_data_section(ws_t1, "Messung-Nr.")
    if header_row:
        data_row = header_row + 1
        row = data_row
        while ws_t1.cell(row=row, column=1).value is not None:
            # Referenzwert eintragen
            ws_t1.cell(row=row, column=2, value=referenzwert)
            # Fuer jede Person eine Messung simulieren
            for p in range(num_personen):
                operator_id = f"Person {p + 1}"
                # Type-1: alle messen denselben Punkt -> nur Messrauschen
                measurement = referenzwert + katapult.config.noise.measurement_noise(katapult.rng)
                if operator_id not in katapult.config.noise._operator_biases:
                    katapult.config.noise.get_operator_bias(operator_id, katapult.rng)
                bias = katapult.config.noise._operator_biases[operator_id]
                measurement += bias
                ws_t1.cell(row=row, column=3 + p, value=round(measurement, 1))
            row += 1

    # --- Sheet: Reproduzierbarkeit ---
    ws_repr = wb["Reproduzierbarkeit"]
    header_row, num_personen = _find_msa_data_section(ws_repr, "Wurf-ID")
    if header_row:
        data_row = header_row + 1
        row = data_row
        while ws_repr.cell(row=row, column=1).value is not None:
            # Pro Wurf: einmal schiessen, dann jede Person misst
            shot = katapult.shoot(settings, noise_level=noise_level)
            true_val = shot.wurfweite_cm
            for p in range(num_personen):
                operator_id = f"Person {p + 1}"
                # Jede Person misst mit eigenem Bias + Messrauschen
                bias = katapult.config.noise.get_operator_bias(operator_id, katapult.rng)
                noise = katapult.config.noise.measurement_noise(katapult.rng)
                measurement = true_val + bias + noise
                ws_repr.cell(row=row, column=2 + p, value=round(measurement, 1))
            row += 1

    wb.save(output_path)
    return os.path.abspath(output_path)


def fill_doe(
    input_path: str,
    output_path: Optional[str] = None,
    seed: Optional[int] = None,
    noise_level: float = 1.0,
) -> str:
    """Befuellt eine DoE_Versuchsergebnisse.xlsx mit simulierten Wurfweiten.

    Liest den Versuchsplan, fuehrt fuer jede Zeile einen simulierten Schuss
    durch und schreibt das Ergebnis in die letzte Spalte.

    Parameters
    ----------
    input_path : str
        Pfad zur DoE-Vorlage (von helper.erstelle_doe_excel()).
    output_path : str, optional
        Pfad fuer die befuellte Datei. Default: ueberschreibt input_path.
    seed : int, optional
        Zufalls-Seed.
    noise_level : float
        Rausch-Multiplikator.

    Returns
    -------
    str
        Pfad zur befuellten Datei.
    """
    from openpyxl import load_workbook

    if output_path is None:
        output_path = input_path

    wb = load_workbook(input_path)
    ws = wb["Versuchsergebnisse"]
    katapult = Statapult(seed=seed)

    # Header lesen (Zeile 1)
    headers = []
    col = 1
    while ws.cell(row=1, column=col).value is not None:
        headers.append(ws.cell(row=1, column=col).value)
        col += 1

    # Ergebnis-Spalte finden (letzte Spalte, enthaelt "Ergebnis" oder "Weite")
    result_col = len(headers)
    for i, h in enumerate(headers):
        if h and ("ergebnis" in str(h).lower() or "weite" in str(h).lower()):
            result_col = i + 1

    # Faktor-Spalten identifizieren: Spalten mit Einheit in Klammern (Original-Werte)
    factor_cols = {}
    for i, h in enumerate(headers):
        if h is None:
            continue
        h_str = str(h)
        # Suche nach "Name (Einheit)" Pattern (nicht "kodiert")
        if "(" in h_str and ")" in h_str and "kodiert" not in h_str.lower():
            # Faktor-Name extrahieren
            name_part = h_str.split("(")[0].strip()
            # Mapping auf Simulator-Keys
            for key, factor in ALL_FACTORS.items():
                if factor.name.lower() in name_part.lower() or key in name_part.lower().replace("-", "_").replace(" ", "_"):
                    factor_cols[key] = i + 1  # 1-indexed
                    break

    # Datenzeilen verarbeiten
    row = 2
    while ws.cell(row=row, column=1).value is not None:
        settings = {}
        for key, col_idx in factor_cols.items():
            val = ws.cell(row=row, column=col_idx).value
            if val is not None:
                settings[key] = float(val)

        result = katapult.shoot(settings, noise_level=noise_level)
        ws.cell(row=row, column=result_col, value=round(result.wurfweite_cm, 1))
        row += 1

    wb.save(output_path)
    return os.path.abspath(output_path)


def fill_konfirmation(
    input_path: str,
    output_path: Optional[str] = None,
    settings: Optional[Dict[str, float]] = None,
    seed: Optional[int] = None,
    noise_level: float = 1.0,
) -> str:
    """Befuellt eine Konfirmation.xlsx mit simulierten Wurfweiten.

    Liest die empfohlenen Einstellungen aus der Vorlage oder verwendet
    die uebergebenen Settings.

    Parameters
    ----------
    input_path : str
        Pfad zur Konfirmations-Vorlage (von helper.erstelle_konfirmation_excel()).
    output_path : str, optional
        Pfad fuer die befuellte Datei. Default: ueberschreibt input_path.
    settings : dict, optional
        Katapult-Einstellungen. Wenn None, werden die Einstellungen aus
        der Vorlage gelesen.
    seed : int, optional
        Zufalls-Seed.
    noise_level : float
        Rausch-Multiplikator.

    Returns
    -------
    str
        Pfad zur befuellten Datei.
    """
    from openpyxl import load_workbook

    if output_path is None:
        output_path = input_path

    wb = load_workbook(input_path)
    ws = wb["Konfirmation"]

    # Settings aus Vorlage lesen falls nicht uebergeben
    if settings is None:
        settings = _read_konfirmation_settings(ws)

    katapult = Statapult(seed=seed)

    # Datentabelle finden (Header mit "Wurf-ID")
    header_row = None
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val and "wurf" in str(val).lower() and "id" in str(val).lower():
            header_row = row
            break

    if header_row is None:
        raise ValueError("Konnte Datentabelle in Konfirmation.xlsx nicht finden")

    # Datenzeilen befuellen
    row = header_row + 1
    while ws.cell(row=row, column=1).value is not None:
        result = katapult.shoot(settings, noise_level=noise_level)
        ws.cell(row=row, column=2, value=round(result.wurfweite_cm, 1))
        row += 1

    wb.save(output_path)
    return os.path.abspath(output_path)


# ======================================================================
# Hilfsfunktionen
# ======================================================================

def _find_msa_data_section(ws, first_header: str) -> tuple[Optional[int], int]:
    """Findet die Header-Zeile und Anzahl Personen in einem MSA-Sheet."""
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val and first_header.lower() in str(val).lower():
            # Personen zaehlen: alle weiteren Spalten mit Werten
            num_personen = 0
            col = 2
            while ws.cell(row=row, column=col).value is not None:
                num_personen += 1
                col += 1
            # Abzug fuer "Referenzwert" Spalte bei Type-1
            if first_header == "Messung-Nr.":
                num_personen -= 1  # Referenzwert-Spalte abziehen
            return row, max(1, num_personen)
    return None, 0


def _read_konfirmation_settings(ws) -> Dict[str, float]:
    """Liest die empfohlenen Einstellungen aus einer Konfirmation.xlsx."""
    settings = {}
    for row in range(1, ws.max_row + 1):
        val_a = ws.cell(row=row, column=1).value
        val_b = ws.cell(row=row, column=2).value
        if val_a and val_b and ":" in str(val_a):
            name = str(val_a).strip().rstrip(":")
            # Wert aus "123.4 cm" oder "123.4 Grad" extrahieren
            val_str = str(val_b).strip().split()[0]
            try:
                value = float(val_str)
            except ValueError:
                continue
            # Auf Simulator-Key mappen
            for key, factor in ALL_FACTORS.items():
                if factor.name.lower() in name.lower() or key in name.lower().replace("-", "_").replace(" ", "_"):
                    settings[key] = value
                    break
    return settings
