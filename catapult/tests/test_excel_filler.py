"""End-to-End Tests fuer die Excel-Fill-Operationen.

Testet den kompletten Workflow:
1. Template mit helper.py erzeugen
2. Mit statapult fill befuellen
3. Pruefen ob die befuellte Datei korrekt ist
4. Pruefen ob helper.py die befuellte Datei korrekt einliest
"""

import os
import sys
import tempfile

import numpy as np
import pandas as pd
import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))
os.environ.setdefault("MPLBACKEND", "Agg")

from statapult.excel_filler import fill_doe, fill_konfirmation, fill_msa


@pytest.fixture
def tmp_dir():
    """Erzeugt ein temporaeres Verzeichnis fuer Test-Dateien."""
    with tempfile.TemporaryDirectory() as d:
        yield d


@pytest.fixture
def msa_template(tmp_dir):
    """Erzeugt eine MSA-Vorlage mit helper.py."""
    from helper import generate_msa_template

    path = os.path.join(tmp_dir, "MSA_Template.xlsx")
    generate_msa_template(
        gruppenname="TestGruppe", num_personen=3,
        num_messungen=10, num_wuerfe=10, messmodus="1D",
        output_path=path,
    )
    return path


@pytest.fixture
def doe_template(tmp_dir):
    """Erzeugt eine DoE-Vorlage mit helper.py."""
    from helper import erstelle_doe_excel, generiere_versuchsplan

    faktoren = [
        {"name": "Abzugswinkel", "einheit": "Grad", "low": 130, "high": 170},
        {"name": "Gummiband-Position", "einheit": "cm", "low": 8, "high": 18},
        {"name": "Becherposition", "einheit": "cm", "low": 8, "high": 22},
    ]
    plan = generiere_versuchsplan(faktoren, wiederholungen=1, centerpoints=3, seed=42)
    path = os.path.join(tmp_dir, "DoE_Versuchsergebnisse.xlsx")
    erstelle_doe_excel(plan, faktoren, output_path=path)
    return path, faktoren


@pytest.fixture
def konfirmation_template(tmp_dir):
    """Erzeugt eine Konfirmations-Vorlage mit helper.py."""
    from helper import erstelle_konfirmation_excel

    einstellungen = {
        "Abzugswinkel": {"original": 160.0, "einheit": "Grad"},
        "Stoppwinkel": {"original": 100.0, "einheit": "Grad"},
        "Gummiband-Position": {"original": 15.0, "einheit": "cm"},
        "Becherposition": {"original": 18.0, "einheit": "cm"},
        "Pin-Hoehe": {"original": 13.0, "einheit": "cm"},
    }
    path = os.path.join(tmp_dir, "Konfirmation.xlsx")
    erstelle_konfirmation_excel(einstellungen, zielweite=350.0, output_path=path)
    return path


# ======================================================================
# MSA Template Tests
# ======================================================================

class TestFillMSA:
    def test_fills_without_error(self, msa_template, tmp_dir):
        """fill_msa laeuft fehlerfrei durch."""
        out = os.path.join(tmp_dir, "MSA_filled.xlsx")
        result = fill_msa(msa_template, output_path=out, seed=42)
        assert os.path.exists(result)

    def test_type1_has_values(self, msa_template, tmp_dir):
        """Type-1 Sheet hat numerische Werte in den Personen-Spalten."""
        from openpyxl import load_workbook

        out = os.path.join(tmp_dir, "MSA_filled.xlsx")
        fill_msa(msa_template, output_path=out, seed=42)

        wb = load_workbook(out, data_only=True)
        ws = wb["Type-1"]

        # Finde Datenbereich (nach Header mit "Messung-Nr.")
        data_start = None
        for row in range(1, 20):
            if ws.cell(row=row, column=1).value and "Messung" in str(ws.cell(row=row, column=1).value):
                data_start = row + 1
                break
        assert data_start is not None, "Header-Zeile nicht gefunden"

        # Prüfe: Referenzwert (Spalte 2) und Personen (Spalten 3-5) befuellt
        ref = ws.cell(row=data_start, column=2).value
        assert isinstance(ref, (int, float)), f"Referenzwert fehlt: {ref}"

        p1 = ws.cell(row=data_start, column=3).value
        assert isinstance(p1, (int, float)), f"Person 1 Messung fehlt: {p1}"

    def test_reproduzierbarkeit_has_values(self, msa_template, tmp_dir):
        """Reproduzierbarkeit Sheet hat Werte fuer alle Personen."""
        from openpyxl import load_workbook

        out = os.path.join(tmp_dir, "MSA_filled.xlsx")
        fill_msa(msa_template, output_path=out, seed=42)

        wb = load_workbook(out, data_only=True)
        ws = wb["Reproduzierbarkeit"]

        data_start = None
        for row in range(1, 20):
            val = ws.cell(row=row, column=1).value
            if val and str(val).strip() == "Wurf-ID":
                data_start = row + 1
                break
        assert data_start is not None, "Header 'Wurf-ID' nicht gefunden"

        # Alle 3 Personen-Spalten befuellt
        for col in [2, 3, 4]:
            val = ws.cell(row=data_start, column=col).value
            assert isinstance(val, (int, float)), f"Spalte {col} leer"

    def test_operator_bias_visible(self, msa_template, tmp_dir):
        """Verschiedene Operatoren zeigen unterschiedliche Mittelwerte."""
        from openpyxl import load_workbook

        out = os.path.join(tmp_dir, "MSA_filled.xlsx")
        fill_msa(msa_template, output_path=out, seed=42)

        wb = load_workbook(out, data_only=True)
        ws = wb["Reproduzierbarkeit"]

        data_start = None
        for row in range(1, 20):
            val = ws.cell(row=row, column=1).value
            if val and str(val).strip() == "Wurf-ID":
                data_start = row + 1
                break

        # Werte pro Person sammeln
        persons = {2: [], 3: [], 4: []}
        row = data_start
        while ws.cell(row=row, column=1).value is not None:
            for col in persons:
                val = ws.cell(row=row, column=col).value
                if isinstance(val, (int, float)):
                    persons[col].append(val)
            row += 1

        means = [np.mean(v) for v in persons.values() if v]
        assert len(means) >= 2
        # Mittelwerte sollten sich leicht unterscheiden (Operator-Bias)
        assert max(means) - min(means) > 0.1, "Kein Operator-Bias sichtbar"

    def test_msa_readable_by_helper(self, msa_template, tmp_dir):
        """Die befuellte MSA-Datei kann von helper.py eingelesen werden."""
        from openpyxl import load_workbook

        from helper import analysiere_type1

        out = os.path.join(tmp_dir, "MSA_filled.xlsx")
        fill_msa(msa_template, output_path=out, seed=42)

        wb = load_workbook(out, data_only=True)
        ws = wb["Type-1"]

        # Daten extrahieren wie das Notebook es tut
        data_start = None
        for row in range(1, 20):
            if ws.cell(row=row, column=1).value and "Messung" in str(ws.cell(row=row, column=1).value):
                data_start = row + 1
                break

        ref_wert = ws.cell(row=data_start, column=2).value
        messdaten = {}
        for p in range(3):
            col_name = f"Person {p+1}"
            vals = []
            row = data_start
            while ws.cell(row=row, column=1).value is not None:
                v = ws.cell(row=row, column=3 + p).value
                if isinstance(v, (int, float)):
                    vals.append(v)
                row += 1
            messdaten[col_name] = vals

        df = pd.DataFrame(messdaten)
        result = analysiere_type1(df, ref_wert)
        assert len(result) == 3
        for person, data in result.items():
            assert "bias" in data
            assert "repeatability" in data
            assert data["n"] == 10


# ======================================================================
# DoE Template Tests
# ======================================================================

class TestFillDoE:
    def test_fills_without_error(self, doe_template, tmp_dir):
        """fill_doe laeuft fehlerfrei durch."""
        path, faktoren = doe_template
        out = os.path.join(tmp_dir, "DoE_filled.xlsx")
        result = fill_doe(path, output_path=out, seed=42)
        assert os.path.exists(result)

    def test_ergebnis_spalte_befuellt(self, doe_template, tmp_dir):
        """Die Ergebnis-Spalte enthaelt numerische Werte > 0."""
        from openpyxl import load_workbook

        path, faktoren = doe_template
        out = os.path.join(tmp_dir, "DoE_filled.xlsx")
        fill_doe(path, output_path=out, seed=42)

        wb = load_workbook(out, data_only=True)
        ws = wb["Versuchsergebnisse"]

        # Letzte Spalte = Ergebnis
        headers = []
        col = 1
        while ws.cell(row=1, column=col).value:
            headers.append(ws.cell(row=1, column=col).value)
            col += 1
        result_col = len(headers)

        # Alle Datenzeilen prüfen
        row = 2
        count = 0
        while ws.cell(row=row, column=1).value is not None:
            val = ws.cell(row=row, column=result_col).value
            assert isinstance(val, (int, float)), f"Zeile {row}: kein Ergebnis"
            assert val > 0, f"Zeile {row}: Ergebnis <= 0"
            count += 1
            row += 1
        assert count > 0

    def test_ergebnisse_variieren(self, doe_template, tmp_dir):
        """Verschiedene Faktoreinstellungen geben verschiedene Ergebnisse."""
        from openpyxl import load_workbook

        path, faktoren = doe_template
        out = os.path.join(tmp_dir, "DoE_filled.xlsx")
        fill_doe(path, output_path=out, seed=42)

        wb = load_workbook(out, data_only=True)
        ws = wb["Versuchsergebnisse"]

        headers = []
        col = 1
        while ws.cell(row=1, column=col).value:
            headers.append(ws.cell(row=1, column=col).value)
            col += 1
        result_col = len(headers)

        ergebnisse = []
        row = 2
        while ws.cell(row=row, column=1).value is not None:
            val = ws.cell(row=row, column=result_col).value
            if isinstance(val, (int, float)):
                ergebnisse.append(val)
            row += 1

        assert len(set(round(e, 0) for e in ergebnisse)) > 3, "Zu wenig Variation"

    def test_doe_readable_by_helper(self, doe_template, tmp_dir):
        """Die befuellte DoE-Datei kann von helper.fitte_modell gelesen werden."""
        from helper import fitte_modell

        path, faktoren = doe_template
        out = os.path.join(tmp_dir, "DoE_filled.xlsx")
        fill_doe(path, output_path=out, seed=42)

        df = pd.read_excel(out)
        modell = fitte_modell(df, faktoren)
        assert modell.rsquared_adj > 0.9, f"R2_adj={modell.rsquared_adj:.4f} zu niedrig"

    def test_reproducible_with_seed(self, doe_template, tmp_dir):
        """Gleicher Seed = gleiche Ergebnisse."""
        from openpyxl import load_workbook

        path, faktoren = doe_template
        out1 = os.path.join(tmp_dir, "DoE_1.xlsx")
        out2 = os.path.join(tmp_dir, "DoE_2.xlsx")
        fill_doe(path, output_path=out1, seed=42)
        fill_doe(path, output_path=out2, seed=42)

        wb1 = load_workbook(out1, data_only=True)
        wb2 = load_workbook(out2, data_only=True)
        ws1 = wb1["Versuchsergebnisse"]
        ws2 = wb2["Versuchsergebnisse"]

        headers = []
        col = 1
        while ws1.cell(row=1, column=col).value:
            headers.append(ws1.cell(row=1, column=col).value)
            col += 1
        rc = len(headers)

        row = 2
        while ws1.cell(row=row, column=1).value is not None:
            assert ws1.cell(row=row, column=rc).value == ws2.cell(row=row, column=rc).value
            row += 1


# ======================================================================
# Konfirmation Template Tests
# ======================================================================

class TestFillKonfirmation:
    def test_fills_without_error(self, konfirmation_template, tmp_dir):
        """fill_konfirmation laeuft fehlerfrei durch."""
        out = os.path.join(tmp_dir, "Konf_filled.xlsx")
        result = fill_konfirmation(
            konfirmation_template, output_path=out, seed=42
        )
        assert os.path.exists(result)

    def test_weite_spalte_befuellt(self, konfirmation_template, tmp_dir):
        """Die Weite-Spalte enthaelt 20 Messwerte."""
        from openpyxl import load_workbook

        out = os.path.join(tmp_dir, "Konf_filled.xlsx")
        fill_konfirmation(konfirmation_template, output_path=out, seed=42)

        wb = load_workbook(out, data_only=True)
        ws = wb["Konfirmation"]

        # Datentabelle finden
        data_start = None
        for row in range(1, 25):
            val = ws.cell(row=row, column=1).value
            if val and "Wurf" in str(val) and "ID" in str(val):
                data_start = row + 1
                break
        assert data_start is not None

        werte = []
        row = data_start
        while ws.cell(row=row, column=1).value is not None:
            val = ws.cell(row=row, column=2).value
            if isinstance(val, (int, float)):
                werte.append(val)
            row += 1

        assert len(werte) == 20, f"Erwartet 20, bekommen {len(werte)}"
        assert all(w > 0 for w in werte)

    def test_werte_nahe_zielweite(self, konfirmation_template, tmp_dir):
        """Die Konfirmationswerte sollten nahe der Zielweite (350 cm) liegen."""
        from openpyxl import load_workbook

        out = os.path.join(tmp_dir, "Konf_filled.xlsx")
        fill_konfirmation(konfirmation_template, output_path=out, seed=42)

        wb = load_workbook(out, data_only=True)
        ws = wb["Konfirmation"]

        data_start = None
        for row in range(1, 25):
            val = ws.cell(row=row, column=1).value
            if val and "Wurf" in str(val) and "ID" in str(val):
                data_start = row + 1
                break

        werte = []
        row = data_start
        while ws.cell(row=row, column=1).value is not None:
            val = ws.cell(row=row, column=2).value
            if isinstance(val, (int, float)):
                werte.append(val)
            row += 1

        mean = np.mean(werte)
        # Settings entsprechen ca. 350-400 cm (abhaengig von Defaults)
        assert 200 < mean < 600, f"Mittelwert {mean:.1f} unplausibel"

    def test_reads_settings_from_template(self, konfirmation_template, tmp_dir):
        """fill_konfirmation liest die Einstellungen aus der Vorlage."""
        from openpyxl import load_workbook

        out = os.path.join(tmp_dir, "Konf_filled.xlsx")
        # Ohne explizite Settings -> liest aus Vorlage
        fill_konfirmation(konfirmation_template, output_path=out, seed=42)

        wb = load_workbook(out, data_only=True)
        ws = wb["Konfirmation"]

        data_start = None
        for row in range(1, 25):
            val = ws.cell(row=row, column=1).value
            if val and "Wurf" in str(val) and "ID" in str(val):
                data_start = row + 1
                break

        val = ws.cell(row=data_start, column=2).value
        assert isinstance(val, (int, float)), "Kein Wert eingetragen"


# ======================================================================
# CLI fill Subcommand Tests
# ======================================================================

class TestCLIFill:
    def test_fill_doe_cli(self, doe_template, tmp_dir):
        """CLI: statapult fill DoE-Datei funktioniert."""
        import subprocess

        path, _ = doe_template
        out = os.path.join(tmp_dir, "cli_doe.xlsx")
        result = subprocess.run(
            [sys.executable, "-m", "statapult", "fill", path,
             "-o", out, "--seed", "42"],
            capture_output=True, text=True,
        )
        assert result.returncode == 0, f"Stderr: {result.stderr}"
        assert os.path.exists(out)

    def test_fill_msa_cli(self, msa_template, tmp_dir):
        """CLI: statapult fill MSA-Datei funktioniert."""
        import subprocess

        out = os.path.join(tmp_dir, "cli_msa.xlsx")
        result = subprocess.run(
            [sys.executable, "-m", "statapult", "fill", msa_template,
             "-o", out, "--seed", "42"],
            capture_output=True, text=True,
        )
        assert result.returncode == 0, f"Stderr: {result.stderr}"
        assert os.path.exists(out)

    def test_fill_konfirmation_cli(self, konfirmation_template, tmp_dir):
        """CLI: statapult fill Konfirmation-Datei funktioniert."""
        import subprocess

        out = os.path.join(tmp_dir, "cli_konf.xlsx")
        result = subprocess.run(
            [sys.executable, "-m", "statapult", "fill", konfirmation_template,
             "-o", out, "--seed", "42"],
            capture_output=True, text=True,
        )
        assert result.returncode == 0, f"Stderr: {result.stderr}"
        assert os.path.exists(out)

    def test_fill_nonexistent_file(self, tmp_dir):
        """CLI: Fehlermeldung bei nicht-existierender Datei."""
        import subprocess

        result = subprocess.run(
            [sys.executable, "-m", "statapult", "fill", "/nope.xlsx"],
            capture_output=True, text=True,
        )
        assert result.returncode != 0
