#!/usr/bin/env python3
"""
DMAIC Katapult-Versuch – Hilfsfunktionen (helper.py)

Enthält die gesamte statistische Logik für das Jupyter Notebook.
Studierenden wird dieser Code nicht direkt gezeigt – sie interagieren
über einfache Funktionsaufrufe im Notebook.

Modul-Gliederung:
  1. Konfiguration (Farben, Dataclass, Theme)
  2. Allgemeine Hilfsfunktionen (Ampel, Zielscheibe, Prio-2-Blöcke)
  3. DEFINE – Testwürfe, CV, Projektcharter
  4. MEASURE – MSA (Type-1, Gage R&R ANOVA), Baseline
  5. ANALYZE – DoE-Generierung, Regression, hierarchisches Pruning
  6. IMPROVE – Optimierung, Konturplot, Konfirmation
  7. CONTROL – I-MR-Kontrollkarte, Normalverteilung, Cpk
  8. EXPORT – ZIP-Erstellung

Abhängigkeiten (pip install):
  numpy, pandas, scipy, statsmodels, matplotlib, openpyxl
"""

from __future__ import annotations

import hashlib
import io
import json
import os
import tempfile
import warnings
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple

import matplotlib
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import numpy as np
import pandas as pd
from scipy import optimize, stats
from IPython.display import HTML, display

# statsmodels – lazy importiert in den Funktionen die es brauchen,
# damit der Import-Fehler klar lokalisierbar ist.

# ═══════════════════════════════════════════════════════════════════
# 1. KONFIGURATION
# ═══════════════════════════════════════════════════════════════════

# QM-Farbpalette (konsistent mit QM-01)
BLUE = "#2563EB"
RED = "#DC2626"
GREEN = "#16A34A"
ORANGE = "#EA580C"
GRAY = "#6B7280"
LIGHT_BLUE = "#DBEAFE"
LIGHT_RED = "#FEE2E2"
LIGHT_GREEN = "#DCFCE7"
LIGHT_YELLOW = "#FEF9C3"
LIGHT_PURPLE = "#F3E8FF"

COLORS = dict(
    blue=BLUE, red=RED, green=GREEN, orange=ORANGE, gray=GRAY,
    light_blue=LIGHT_BLUE, light_red=LIGHT_RED, light_green=LIGHT_GREEN,
    light_yellow=LIGHT_YELLOW, light_purple=LIGHT_PURPLE,
)


def setup_theme():
    """Setzt das globale Matplotlib-Theme für alle Plots."""
    plt.rcParams.update({
        "figure.dpi": 150,
        "savefig.dpi": 150,
        "font.family": "sans-serif",
        "font.size": 11,
        "axes.titlesize": 14,
        "axes.labelsize": 12,
        "figure.facecolor": "white",
        "axes.facecolor": "white",
        "axes.grid": False,
    })


@dataclass
class Projekt:
    """Sammelt alle Daten und Ergebnisse über die DMAIC-Phasen hinweg."""
    gruppenname: str = ""
    gruppennummer: int = 1
    seed: int = 42
    zielweite: float = 0.0
    toleranz: float = 15.0
    messmodus: str = "1D"  # "1D" oder "2D"

    # DEFINE
    testwuerfe: np.ndarray = field(default_factory=lambda: np.array([]))
    charter: Dict[str, str] = field(default_factory=dict)
    # Faktoren werden bereits in DEFINE definiert (Master-Liste).
    faktoren: List[Dict] = field(default_factory=list)
    # Katapult-Vermessung (Min/Max-Konfiguration mit je 3 Würfen)
    vermessung_min_wuerfe: np.ndarray = field(default_factory=lambda: np.array([]))
    vermessung_max_wuerfe: np.ndarray = field(default_factory=lambda: np.array([]))
    vermessung_min_einstellung: Dict[str, float] = field(default_factory=dict)
    vermessung_max_einstellung: Dict[str, float] = field(default_factory=dict)
    vermessung_beschreibung: str = ""
    # Manuelle Annäherung an die Zielweite (OFAT): Protokoll der Iterationen +
    # finalisierte initiale Einstellung für Testwürfe und Baseline.
    annaeherung_log: List[Dict] = field(default_factory=list)
    initiale_einstellung: Dict[str, float] = field(default_factory=dict)

    # MEASURE
    msa_type1: Optional[Dict] = None
    msa_grr: Optional[Dict] = None
    msa_rohdaten: Optional[pd.DataFrame] = None  # Long-Format MSA-Daten für Visualisierungen nach Restore
    baseline_wuerfe: np.ndarray = field(default_factory=lambda: np.array([]))
    baseline_stats: Optional[Dict] = None

    # ANALYZE – ggf. reduzierte Teilmenge + Centerpoint-Entscheidung je Faktor
    faktoren_doe: List[Dict] = field(default_factory=list)
    versuchsplan: Optional[pd.DataFrame] = None
    doe_ergebnisse: Optional[pd.DataFrame] = None
    modell: Any = None
    modell_gepruned: Any = None
    pruning_log: List[str] = field(default_factory=list)

    # IMPROVE
    optimale_einstellung: Optional[Dict] = None
    konfirmation_wuerfe: np.ndarray = field(default_factory=lambda: np.array([]))
    konfirmation_ergebnis: Optional[Dict] = None

    # CONTROL
    imr_ergebnis: Optional[Dict] = None
    cpk_ergebnis: Optional[Dict] = None

    # PERSISTENZ-CONFIG
    versuchsplan_config: Optional[Dict] = None   # {"wiederholungen", "blocking", "centerpoints"}
    optimierung_config: Optional[Dict] = None    # {"strategie", "lambda_gewicht"}

    # Alle erzeugten Figuren (für ZIP-Export)
    figuren: Dict[str, plt.Figure] = field(default_factory=dict)
    csv_daten: Dict[str, pd.DataFrame] = field(default_factory=dict)


def init_projekt(
    gruppenname: str,
    gruppennummer: int,
    zielweite: float = 300.0,
    toleranz: float = 15.0,
) -> Projekt:
    """Initialisiert ein Projekt mit deterministischem Seed und nutzerdefinierter Zielweite.

    Der Seed wird aus Gruppenname + Gruppennummer abgeleitet und für Verzeichnisnamen
    sowie Randomisierungen in MSA/ANALYZE verwendet. Die Zielweite wird **nicht** mehr
    zufällig gezogen, sondern vom Nutzer vorgegeben (Default 300 cm).
    """
    seed_str = f"{gruppenname.strip().lower()}_{gruppennummer}"
    seed = int(hashlib.md5(seed_str.encode()).hexdigest()[:8], 16) % (2**31)

    p = Projekt(
        gruppenname=gruppenname,
        gruppennummer=gruppennummer,
        seed=seed,
        zielweite=float(zielweite),
        toleranz=toleranz,
    )
    return p


def _effektive_faktoren(projekt: Projekt) -> List[Dict]:
    """Gibt `faktoren_doe` zurück, falls in ANALYZE eine Teilmenge gewählt wurde,
    sonst die in DEFINE definierte Master-Liste `faktoren`."""
    return projekt.faktoren_doe if projekt.faktoren_doe else projekt.faktoren


# ═══════════════════════════════════════════════════════════════════
# 2. ALLGEMEINE HILFSFUNKTIONEN
# ═══════════════════════════════════════════════════════════════════

def ampel(wert: float, schwellen: List[Tuple[float, str, str]],
          einheit: str = "", titel: str = "") -> str:
    """
    Erzeugt eine Ampel-Bewertung als HTML.

    schwellen: Liste von (obere_grenze, symbol, text).
    Die letzte Schwelle sollte float('inf') haben.

    Beispiel:
        ampel(8.5, [(10, '✅', 'Akzeptabel'), (30, '⚠️', 'Bedingt'), (inf, '❌', 'Nicht akzeptabel')])
    """
    for grenze, symbol, text in schwellen:
        if wert <= grenze:
            farbe = "#16A34A" if "✅" in symbol else "#EA580C" if "⚠️" in symbol else "#DC2626"
            html = f"""
            <div style="padding:12px; border-left:4px solid {farbe};
                        background:{farbe}11; border-radius:4px; margin:8px 0;">
                <span style="font-size:1.3em;">{symbol}</span>
                <strong>{titel}</strong> {wert:.2f}{einheit} — {text}
            </div>"""
            return html
    return ""


def zeige_ampel(wert, schwellen, einheit="", titel=""):
    """Zeigt eine Ampel-Bewertung direkt im Notebook an."""
    display(HTML(ampel(wert, schwellen, einheit, titel)))


def prio2_block(titel: str, inhalt: str) -> str:
    """Erzeugt einen einklappbaren Prio-2-Block (details/summary)."""
    return f"""
    <details style="margin:10px 0; padding:8px; background:#F9FAFB;
                     border:1px solid #E5E7EB; border-radius:6px;">
        <summary style="cursor:pointer; font-weight:bold; color:{BLUE};">
            🔍 Für Neugierige: {titel}
        </summary>
        <div style="margin-top:8px; padding:8px; font-size:0.95em;">
            {inhalt}
        </div>
    </details>"""


def zeige_prio2(titel: str, inhalt: str):
    """Zeigt einen Prio-2-Block direkt an."""
    display(HTML(prio2_block(titel, inhalt)))


def hinweis_bericht(text: str):
    """Zeigt einen 'Für den Bericht'-Hinweis an."""
    html = f"""
    <div style="padding:10px; border-left:4px solid {BLUE};
                background:{LIGHT_BLUE}; border-radius:4px; margin:10px 0;">
        📋 <strong>Für den Bericht:</strong> {text}
    </div>"""
    display(HTML(html))


def _save_fig(projekt: Projekt, fig: plt.Figure, name: str):
    """Speichert eine Figur im Projekt für späteren Export."""
    projekt.figuren[name] = fig


# ═══════════════════════════════════════════════════════════════════
# 2b. ZIELSCHEIBEN-VISUALISIERUNG
# ═══════════════════════════════════════════════════════════════════

def _draw_target_background(ax, zielweite=0, toleranz=None, modus="1D"):
    """Zeichnet den Zielscheiben-Hintergrund."""
    if modus == "2D":
        # Volle 2D-Zielscheibe
        for r, c in [(1.0, "#F3F4F6"), (0.7, "#E5E7EB"), (0.4, "#D1D5DB"), (0.15, "#9CA3AF")]:
            circle = plt.Circle((0, 0), r, facecolor=c, edgecolor=GRAY, linewidth=0.8)
            ax.add_patch(circle)
        ax.axhline(0, color=GRAY, linewidth=0.5, linestyle="--", alpha=0.4)
        ax.axvline(0, color=GRAY, linewidth=0.5, linestyle="--", alpha=0.4)
        if toleranz:
            tol_circle = plt.Circle((0, 0), toleranz / 100,
                                     fill=False, edgecolor=GREEN, linewidth=2, linestyle="--")
            ax.add_patch(tol_circle)
        ax.set_xlim(-1.3, 1.3)
        ax.set_ylim(-1.3, 1.3)
        ax.set_aspect("equal")
    else:
        # 1D-Streifendiagramm
        ax.axhline(0, color=GRAY, linewidth=1, alpha=0.3)
        ax.axvline(zielweite, color=GREEN, linewidth=2, linestyle="--", label="Zielweite")
        if toleranz:
            ax.axvspan(zielweite - toleranz, zielweite + toleranz,
                       alpha=0.1, color=GREEN, label=f"Toleranz ±{toleranz} cm")


def plot_zielscheibe(daten_weite, ziel, toleranz, modus="1D",
                     daten_quer=None, titel="Zielscheibe",
                     farbe=BLUE, ax=None, show_stats=True):
    """
    Plottet Wurfweiten als Zielscheiben-Darstellung.

    Parameter:
        daten_weite: Array der Wurfweiten (cm)
        ziel: Zielweite (cm)
        toleranz: Toleranz (cm)
        modus: "1D" oder "2D"
        daten_quer: Array der Querversätze (nur bei 2D)
        titel: Plot-Titel
        farbe: Punkt-Farbe
        ax: Vorhandenes Axes-Objekt (optional)
    """
    own_fig = ax is None
    if own_fig:
        fig, ax = plt.subplots(figsize=(8, 5) if modus == "1D" else (6, 6))

    if modus == "2D" and daten_quer is not None:
        # Normalisiere auf Zielscheiben-Skala
        max_range = max(toleranz * 2, np.ptp(daten_weite), np.ptp(daten_quer)) * 1.2
        x_norm = (daten_weite - ziel) / max_range
        y_norm = daten_quer / max_range
        _draw_target_background(ax, modus="2D", toleranz=toleranz)
        ax.scatter(x_norm, y_norm, c=farbe, s=60, zorder=5,
                   edgecolors="white", linewidths=0.8)
        ax.set_xlabel("Weite (relativ zum Ziel)")
        ax.set_ylabel("Querversatz (relativ)")
    else:
        # 1D-Modus
        _draw_target_background(ax, zielweite=ziel, toleranz=toleranz, modus="1D")
        y_jitter = np.random.default_rng(42).uniform(-0.3, 0.3, len(daten_weite))
        ax.scatter(daten_weite, y_jitter, c=farbe, s=60, zorder=5,
                   edgecolors="white", linewidths=0.8)
        ax.set_xlabel("Wurfweite (cm)")
        ax.set_yticks([])
        ax.set_ylim(-1, 1)

    if show_stats:
        mu = np.mean(daten_weite)
        sigma = np.std(daten_weite, ddof=1) if len(daten_weite) > 1 else 0
        stats_text = f"μ = {mu:.1f} cm | σ = {sigma:.1f} cm | n = {len(daten_weite)}"
        ax.set_title(f"{titel}\n{stats_text}", fontsize=12, fontweight="bold")
    else:
        ax.set_title(titel, fontsize=12, fontweight="bold")

    ax.legend(loc="upper right", fontsize=9)
    if own_fig:
        fig.tight_layout()
        return fig
    return ax


def plot_4_zielscheiben_referenz():
    """Plottet die 4 Referenz-Zielscheiben (Accuracy × Precision)."""
    np.random.seed(42)
    fig, axes = plt.subplots(2, 2, figsize=(8, 8))

    configs = [
        ("Hohe Accuracy\nHohe Precision", np.random.normal(0, 0.08, (15, 2)), GREEN),
        ("Niedrige Accuracy\nHohe Precision", np.random.normal(0, 0.08, (15, 2)) + [0.5, 0.3], ORANGE),
        ("Hohe Accuracy\nNiedrige Precision", np.random.normal(0, 0.35, (15, 2)), BLUE),
        ("Niedrige Accuracy\nNiedrige Precision", np.random.normal(0, 0.35, (15, 2)) + [0.4, -0.3], RED),
    ]

    for ax, (title, pts, col) in zip(axes.flat, configs):
        _draw_target_background(ax, modus="2D")
        ax.scatter(pts[:, 0], pts[:, 1], c=col, s=35, zorder=5,
                   edgecolors="white", linewidths=0.5)
        ax.set_title(title, fontsize=11, fontweight="bold")
        ax.set_xticks([])
        ax.set_yticks([])

    fig.suptitle("Accuracy vs. Precision – Referenz", fontsize=14, fontweight="bold", y=0.98)
    fig.tight_layout(rect=[0, 0, 1, 0.95])
    return fig


# ═══════════════════════════════════════════════════════════════════
# 3. DEFINE
# ═══════════════════════════════════════════════════════════════════

def berechne_testwurf_statistik(wuerfe: np.ndarray) -> Dict:
    """Berechnet Statistiken der 5 Testwürfe."""
    mu = np.mean(wuerfe)
    sigma = np.std(wuerfe, ddof=1) if len(wuerfe) > 1 else 0
    cv = (sigma / mu * 100) if mu != 0 else 0
    return {"mean": mu, "std": sigma, "cv": cv, "n": len(wuerfe)}


def zeige_testwurf_ergebnis(projekt: Projekt):
    """Zeigt die Testwurf-Auswertung inkl. Zielscheibe und CV-Warnung."""
    stats = berechne_testwurf_statistik(projekt.testwuerfe)

    # Zielscheibe (1D-Modus: nur Weite)
    fig = plot_zielscheibe(
        projekt.testwuerfe, projekt.zielweite, projekt.toleranz,
        modus="1D", titel="Testwürfe – Ist-Zustand"
    )
    _save_fig(projekt, fig, "define_testwuerfe")
    plt.show()

    # CV-Bewertung
    cv_schwellen = [
        (15, "✅", "Guter Start – weiter mit MSA"),
        (30, "⚠️", "Grenzwertig – Einstellungen und Arretierung prüfen"),
        (float("inf"), "❌", "Katapult hat ein Problem! Arretierung prüfen, ggf. nachbessern."),
    ]
    zeige_ampel(stats["cv"], cv_schwellen, einheit="%", titel="Variationskoeffizient (CV):")


# ───────────────────────────────────────────────────────────────────
# 3a. KATAPULT-VERMESSUNG (Min/Max-Charakterisierung)
# ───────────────────────────────────────────────────────────────────

def speichere_vermessung(
    projekt: Projekt,
    min_wuerfe,
    max_wuerfe,
    min_einstellung: Dict[str, float],
    max_einstellung: Dict[str, float],
    beschreibung: str = "",
) -> None:
    """Speichert die Katapult-Vermessung (Min/Max) im Projekt und persistiert.

    Leere / ungültige Würfe (≤ 0) werden aussortiert. Einstellungen werden
    als Mapping Faktorname → Wert übernommen. Bei weniger als einem gültigen
    Wurf je Seite wird gewarnt, aber nichts überschrieben.
    """
    min_list = list(min_wuerfe) if min_wuerfe is not None else []
    max_list = list(max_wuerfe) if max_wuerfe is not None else []
    _min = np.array([w for w in min_list if w and w > 0], dtype=float)
    _max = np.array([w for w in max_list if w and w > 0], dtype=float)

    if len(_min) > 0:
        projekt.vermessung_min_wuerfe = _min
    if len(_max) > 0:
        projekt.vermessung_max_wuerfe = _max

    if min_einstellung:
        projekt.vermessung_min_einstellung = dict(min_einstellung)
    if max_einstellung:
        projekt.vermessung_max_einstellung = dict(max_einstellung)

    if beschreibung:
        projekt.vermessung_beschreibung = beschreibung

    # Nur warnen, wenn Werte übergeben wurden, die komplett herausgefiltert wurden.
    if min_list and len(_min) == 0:
        print("⚠️ Keine gültigen Min-Würfe (Werte > 0 erwartet).")
    if max_list and len(_max) == 0:
        print("⚠️ Keine gültigen Max-Würfe (Werte > 0 erwartet).")

    speichere_fortschritt(projekt)


def zeige_vermessung(projekt: Projekt) -> None:
    """Visualisiert die gemessene Katapult-Spanne inkl. Zielweiten-Marker."""
    min_w = projekt.vermessung_min_wuerfe
    max_w = projekt.vermessung_max_wuerfe

    if len(min_w) == 0 or len(max_w) == 0:
        print("ℹ️ Vermessung noch nicht vollständig – bitte Min- und Max-Würfe eintragen.")
        return

    min_mu = float(np.mean(min_w))
    max_mu = float(np.mean(max_w))
    ziel = float(projekt.zielweite)

    fig, ax = plt.subplots(figsize=(9, 2.6))
    ax.hlines(0, min_mu, max_mu, color=BLUE, linewidth=6, alpha=0.35,
              label=f"Gemessene Spanne ({min_mu:.0f}–{max_mu:.0f} cm)")
    ax.scatter(min_w, [0] * len(min_w), color=BLUE, s=60, zorder=5,
               edgecolors="white", linewidths=0.8, label="Min-Würfe")
    ax.scatter(max_w, [0] * len(max_w), color=ORANGE, s=60, zorder=5,
               edgecolors="white", linewidths=0.8, label="Max-Würfe")
    ax.axvline(ziel, color=GREEN, linewidth=2.5, linestyle="--",
               label=f"Zielweite {ziel:.0f} cm")
    if projekt.toleranz:
        ax.axvspan(ziel - projekt.toleranz, ziel + projekt.toleranz,
                   alpha=0.12, color=GREEN)
    ax.set_yticks([])
    ax.set_xlabel("Wurfweite (cm)")
    ax.set_title("Katapult-Charakterisierung: Min/Max-Vermessung",
                 fontsize=12, fontweight="bold")
    ax.legend(loc="upper right", fontsize=9)
    margin = max(30.0, 0.08 * (max_mu - min_mu))
    ax.set_xlim(min(min_mu, ziel) - margin, max(max_mu, ziel) + margin)
    fig.tight_layout()
    _save_fig(projekt, fig, "define_vermessung")
    plt.show()

    # Ampel: Zielweite innerhalb der Spanne?
    if min_mu <= ziel <= max_mu:
        farbe, symbol, text = GREEN, "✅", "Zielweite liegt innerhalb der gemessenen Spanne."
    else:
        farbe, symbol, text = ORANGE, "⚠️", (
            "Zielweite liegt außerhalb der gemessenen Spanne – ggf. anpassen."
        )
    display(HTML(f"""
    <div style="padding:10px; border-left:4px solid {farbe}; background:{farbe}11;
                border-radius:4px; margin:8px 0;">
        <span style="font-size:1.2em;">{symbol}</span> <strong>Plausibilität:</strong> {text}
    </div>"""))

    # Einstellungs-Tabelle
    faktoren = projekt.faktoren
    if faktoren and (projekt.vermessung_min_einstellung or projekt.vermessung_max_einstellung):
        rows = ""
        for f in faktoren:
            name = f["name"]
            lo = projekt.vermessung_min_einstellung.get(name, "—")
            hi = projekt.vermessung_max_einstellung.get(name, "—")
            einheit = f.get("einheit", "")
            rows += (f"<tr><td style='padding:6px;font-weight:bold;'>{name}</td>"
                     f"<td style='padding:6px;'>{lo} {einheit}</td>"
                     f"<td style='padding:6px;'>{hi} {einheit}</td></tr>")
        display(HTML(f"""
        <table style="border-collapse:collapse; width:100%; border:1px solid #E5E7EB; margin-top:8px;">
            <tr style="background:{LIGHT_BLUE};">
                <th style="padding:8px;text-align:left;">Faktor</th>
                <th style="padding:8px;text-align:left;">Min-Konfiguration</th>
                <th style="padding:8px;text-align:left;">Max-Konfiguration</th>
            </tr>
            {rows}
        </table>"""))

    if projekt.vermessung_beschreibung:
        display(HTML(
            f"<div style='margin-top:8px; padding:8px; background:#F9FAFB; "
            f"border:1px solid #E5E7EB; border-radius:4px;'>"
            f"<strong>Beschreibung:</strong> {projekt.vermessung_beschreibung}</div>"
        ))


def setze_zielweite(projekt: Projekt, zielweite: float) -> None:
    """Setzt `projekt.zielweite`, speichert den Fortschritt und warnt bei Inkonsistenz.

    Warnt, wenn die gewählte Zielweite außerhalb der durch die Vermessung belegten
    Spanne [mean(min_wuerfe), mean(max_wuerfe)] liegt. Die Zielweite wird dennoch
    gesetzt, damit bewusst anspruchsvolle Ziele möglich bleiben.
    """
    try:
        z = float(zielweite)
    except (TypeError, ValueError):
        print(f"⚠️ Zielweite '{zielweite}' ist keine Zahl – wird ignoriert.")
        return

    if z <= 0:
        print(f"⚠️ Zielweite muss > 0 sein (erhalten: {z}).")
        return

    projekt.zielweite = z

    if len(projekt.vermessung_min_wuerfe) > 0 and len(projekt.vermessung_max_wuerfe) > 0:
        lo = float(np.mean(projekt.vermessung_min_wuerfe))
        hi = float(np.mean(projekt.vermessung_max_wuerfe))
        if not (lo <= z <= hi):
            print(f"⚠️ Zielweite {z:.0f} cm liegt außerhalb der gemessenen Spanne "
                  f"[{lo:.0f}, {hi:.0f}] cm – Erreichen ist unsicher.")
        else:
            print(f"✅ Zielweite {z:.0f} cm liegt innerhalb der gemessenen Spanne "
                  f"[{lo:.0f}, {hi:.0f}] cm.")
    else:
        print(f"ℹ️ Zielweite auf {z:.0f} cm gesetzt "
              f"(noch keine Min/Max-Vermessung zur Plausibilitätsprüfung vorhanden).")

    speichere_fortschritt(projekt)


def zeige_faktoren_legende(projekt: Projekt) -> None:
    """Druckt eine klar nummerierte Legende der in DEFINE gesetzten Faktoren.

    Wird am Anfang aller Zellen aufgerufen, in denen der Student Werte oder
    Würfe pro Faktor einträgt, damit die generischen Formularfelder (`wert_1`,
    `wert_2` …) eindeutig zugeordnet werden können.
    """
    if not projekt.faktoren:
        print("⚠️ Noch keine Faktoren definiert. Bitte zuerst Schritt 1 (Faktoren definieren) ausführen.")
        return
    lines = ["📋 Eure Faktoren aus Schritt 1 (Reihenfolge entspricht den Eingabefeldern):"]
    for i, f in enumerate(projekt.faktoren, start=1):
        lines.append(f"   {i}. {f['name']} [{f['low']} – {f['high']} {f['einheit']}]")
    print("\n".join(lines))


def protokolliere_annaeherung(projekt: Projekt,
                              einstellung: Dict[str, float],
                              wuerfe) -> None:
    """Speichert eine Iteration der manuellen Annäherung (OFAT).

    Führt eine "nur ein Faktor geändert"-Prüfung gegen den letzten Eintrag und
    warnt bei Verletzung. Die Einstellung wird immer gespeichert — die
    Studierenden entscheiden selbst, ob sie verworfen wird.
    """
    w = np.array([v for v in wuerfe if v and v > 0], dtype=float)
    if len(w) == 0:
        print("⚠️ Keine gültigen Würfe eingegeben (Werte > 0 erwartet).")
        return

    eintrag = {
        "iteration": len(projekt.annaeherung_log) + 1,
        "einstellung": dict(einstellung),
        "wuerfe": w.tolist(),
        "mean": float(w.mean()),
        "abweichung_vom_ziel": float(w.mean() - projekt.zielweite),
    }
    projekt.annaeherung_log.append(eintrag)

    # OFAT-Plausibilität: wie viele Faktoren haben sich vs. letzte Iteration geändert?
    if len(projekt.annaeherung_log) >= 2:
        vorher = projekt.annaeherung_log[-2]["einstellung"]
        geaendert = [name for name, val in einstellung.items()
                     if name in vorher and float(val) != float(vorher[name])]
        if len(geaendert) > 1:
            print(f"⚠️ OFAT-Hinweis: {len(geaendert)} Faktoren wurden gegenüber der letzten "
                  f"Iteration verändert ({', '.join(geaendert)}). Für eine saubere "
                  f"Annäherung solltet ihr pro Iteration nur **einen** Faktor verändern.")

    diff = eintrag["abweichung_vom_ziel"]
    print(f"✅ Iteration {eintrag['iteration']} protokolliert: μ={eintrag['mean']:.1f} cm "
          f"(Abweichung zur Zielweite: {diff:+.1f} cm).")
    speichere_fortschritt(projekt)


def setze_initiale_einstellung(projekt: Projekt,
                               einstellung: Optional[Dict[str, float]] = None) -> None:
    """Übernimmt eine Einstellung als "initiale Einstellung" für Testwürfe und Baseline.

    Wird ohne Argument aufgerufen, übernimmt sie die Einstellung der **letzten**
    Annäherungs-Iteration. Sonst wird das übergebene Dict verwendet.
    """
    if einstellung is None:
        if not projekt.annaeherung_log:
            print("⚠️ Keine Annäherungs-Iteration vorhanden — bitte zuerst eine Iteration "
                  "protokollieren oder Einstellung direkt übergeben.")
            return
        einstellung = projekt.annaeherung_log[-1]["einstellung"]

    if not einstellung:
        print("⚠️ Leere Einstellung — keine Änderung gespeichert.")
        return

    projekt.initiale_einstellung = dict(einstellung)
    print("✅ Initiale Einstellung übernommen:")
    for name, val in projekt.initiale_einstellung.items():
        einheit = next((f["einheit"] for f in projekt.faktoren if f["name"] == name), "")
        print(f"   • {name}: {val} {einheit}")
    speichere_fortschritt(projekt)


def pruefe_excel_faktoren_konsistenz(projekt: Projekt,
                                     excel_faktoren: List[Dict]) -> List[str]:
    """Vergleicht die aus der DoE-Excel geparsten Faktoren mit projekt.faktoren_doe.

    Zurück: Liste menschenlesbarer Diskrepanz-Meldungen (leer → alles konsistent).
    Bei Abweichungen wird zusätzlich ein Hinweis angezeigt. Die Entscheidung,
    welche Variante weitergenutzt wird, trifft der Aufrufer — im Notebook ist
    das die Excel-Seite (weil dort die tatsächlichen Messungen herkommen).
    """
    meldungen: List[str] = []
    referenz = _effektive_faktoren(projekt)
    if not referenz:
        return meldungen  # nichts zu vergleichen

    ref_namen = [f["name"] for f in referenz]
    excel_namen = [f["name"] for f in excel_faktoren]

    # Zusätzliche / fehlende Faktoren
    nur_excel = [n for n in excel_namen if n not in ref_namen]
    nur_define = [n for n in ref_namen if n not in excel_namen]
    for n in nur_excel:
        meldungen.append(f"Faktor '{n}' in Excel vorhanden, aber nicht in DEFINE/ANALYZE definiert.")
    for n in nur_define:
        meldungen.append(f"Faktor '{n}' in DEFINE/ANALYZE definiert, aber nicht in Excel.")

    # Stufen-Abweichungen auf gemeinsamen Faktoren
    ref_map = {f["name"]: f for f in referenz}
    for fx in excel_faktoren:
        if fx["name"] not in ref_map:
            continue
        fr = ref_map[fx["name"]]
        if abs(float(fx["low"]) - float(fr["low"])) > 1e-6:
            meldungen.append(
                f"Faktor '{fx['name']}': Low={fx['low']} in Excel vs. {fr['low']} in DEFINE."
            )
        if abs(float(fx["high"]) - float(fr["high"])) > 1e-6:
            meldungen.append(
                f"Faktor '{fx['name']}': High={fx['high']} in Excel vs. {fr['high']} in DEFINE."
            )
        if fx.get("einheit") and fr.get("einheit") and fx["einheit"] != fr["einheit"]:
            meldungen.append(
                f"Faktor '{fx['name']}': Einheit '{fx['einheit']}' in Excel vs. "
                f"'{fr['einheit']}' in DEFINE."
            )

    if meldungen:
        html = "<br>".join(f"• {m}" for m in meldungen)
        display(HTML(f"""
        <div style="padding:10px; border-left:4px solid {ORANGE}; background:{LIGHT_YELLOW};
                     border-radius:4px; margin:8px 0;">
            ⚠️ <strong>Abweichung zwischen Excel und Faktordefinition:</strong><br>
            {html}<br>
            <em>Die Excel-Werte sind maßgeblich, da die Messungen unter diesen Bedingungen
            stattfanden.</em>
        </div>"""))
    return meldungen


def formatiere_charter(projekt: Projekt) -> str:
    """Formatiert die Projektcharter als HTML-Tabelle."""
    rows = ""
    for key, val in projekt.charter.items():
        rows += f"<tr><td style='font-weight:bold; padding:6px;'>{key}</td>"
        rows += f"<td style='padding:6px;'>{val}</td></tr>"
    return f"""
    <table style="border-collapse:collapse; width:100%; border:1px solid #E5E7EB;">
        <tr style="background:{LIGHT_BLUE};">
            <th colspan="2" style="padding:10px; text-align:left;">
                📋 Projektcharter – {projekt.gruppenname}
            </th>
        </tr>
        {rows}
    </table>"""


# ═══════════════════════════════════════════════════════════════════
# 3b. MSA-TEMPLATE-GENERIERUNG
# ═══════════════════════════════════════════════════════════════════

# Excel-Styling-Konstanten (openpyxl, lazy importiert)
_MSA_STYLES_LOADED = False
_HEADER_FONT = _HEADER_FILL = _META_KEY_FONT = _META_VAL_FONT = None
_NORMAL_FONT = _THIN_BORDER = _CENTER_ALIGN = _LEFT_ALIGN = None


def _ensure_msa_styles():
    """Lädt openpyxl-Styling-Konstanten beim ersten Aufruf."""
    global _MSA_STYLES_LOADED, _HEADER_FONT, _HEADER_FILL, _META_KEY_FONT
    global _META_VAL_FONT, _NORMAL_FONT, _THIN_BORDER, _CENTER_ALIGN, _LEFT_ALIGN
    if _MSA_STYLES_LOADED:
        return
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    _HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    _HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    _META_KEY_FONT = Font(name="Calibri", bold=True, size=11)
    _META_VAL_FONT = Font(name="Calibri", size=11, italic=True, color="666666")
    _NORMAL_FONT = Font(name="Calibri", size=11)
    _THIN_BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    _CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
    _LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
    _MSA_STYLES_LOADED = True


def _msa_apply_header(cell):
    _ensure_msa_styles()
    cell.font = _HEADER_FONT
    cell.fill = _HEADER_FILL
    cell.alignment = _CENTER_ALIGN
    cell.border = _THIN_BORDER


def _msa_apply_data(cell, alignment=None):
    _ensure_msa_styles()
    cell.font = _NORMAL_FONT
    cell.border = _THIN_BORDER
    cell.alignment = alignment or _CENTER_ALIGN


def _msa_write_metadata(ws, gruppenname, messmodus, start_row=1):
    """Schreibt den Metadaten-Kopfblock und gibt die nächste Zeile zurück."""
    _ensure_msa_styles()
    from openpyxl.styles import Border, Font, Side
    meta = [
        ("Gruppe:", gruppenname if gruppenname else "<Gruppenname eintragen>"),
        ("Datum:", date.today().strftime("%d.%m.%Y")),
        ("Messmittel:", "<z.\u202fB. Maßband, Laser-Entfernungsmesser>"),
        ("Messmodus:", messmodus),
    ]
    row = start_row
    for key, value in meta:
        key_cell = ws.cell(row=row, column=1, value=key)
        key_cell.font = _META_KEY_FONT
        key_cell.alignment = _LEFT_ALIGN
        val_cell = ws.cell(row=row, column=2, value=value)
        val_cell.font = _META_VAL_FONT
        val_cell.alignment = _LEFT_ALIGN
        val_cell.border = Border(bottom=Side(style="thin", color="AAAAAA"))
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        row += 1
    row += 1  # Leerzeile
    return row


def _msa_build_type1(wb, gruppenname, num_personen, num_messungen, messmodus):
    """Erstellt das Type-1-Blatt."""
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    ws = wb.active
    ws.title = "Type-1"
    data_start = _msa_write_metadata(ws, gruppenname, messmodus)

    hint = ws.cell(row=data_start, column=1, value=(
        "Anleitung: Alle Personen messen unabhängig denselben "
        "Referenzpunkt (z.\u202fB. Klebestreifen auf dem Boden). "
        "Referenzwert = bekannte Distanz in cm."))
    hint.font = Font(name="Calibri", size=10, italic=True, color="888888")
    ws.merge_cells(start_row=data_start, start_column=1,
                   end_row=data_start, end_column=2 + num_personen)
    data_start += 2

    headers = ["Messung-Nr.", "Referenzwert (cm)"]
    headers += [f"Person {p} (cm)" for p in range(1, num_personen + 1)]
    for col_idx, h in enumerate(headers, start=1):
        _msa_apply_header(ws.cell(row=data_start, column=col_idx, value=h))

    for i in range(1, num_messungen + 1):
        row = data_start + i
        _msa_apply_data(ws.cell(row=row, column=1, value=i))
        ref = ws.cell(row=row, column=2, value="<Ref>")
        _msa_apply_data(ref)
        ref.font = Font(name="Calibri", size=11, italic=True, color="999999")
        for p in range(num_personen):
            _msa_apply_data(ws.cell(row=row, column=3 + p, value=None))

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 20
    for p in range(num_personen):
        ws.column_dimensions[get_column_letter(3 + p)].width = 18


def _msa_build_reproduzierbarkeit(wb, gruppenname, num_personen, num_wuerfe, messmodus):
    """Erstellt das Reproduzierbarkeits-Blatt."""
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    ws = wb.create_sheet(title="Reproduzierbarkeit")
    data_start = _msa_write_metadata(ws, gruppenname, messmodus)

    hint = ws.cell(row=data_start, column=1, value=(
        "Anleitung: Nach jedem Wurf den Auftrittspunkt markieren "
        "(Kreppband). Jede Person misst unabhängig die Distanz – "
        "ohne Absprache. Erst danach Markierung entfernen."))
    hint.font = Font(name="Calibri", size=10, italic=True, color="888888")
    ws.merge_cells(start_row=data_start, start_column=1,
                   end_row=data_start, end_column=1 + num_personen)
    data_start += 2

    headers = ["Wurf-ID"]
    headers += [f"Messung Person {p} (cm)" for p in range(1, num_personen + 1)]
    for col_idx, h in enumerate(headers, start=1):
        _msa_apply_header(ws.cell(row=data_start, column=col_idx, value=h))

    for i in range(1, num_wuerfe + 1):
        row = data_start + i
        _msa_apply_data(ws.cell(row=row, column=1, value=i))
        for p in range(num_personen):
            _msa_apply_data(ws.cell(row=row, column=2 + p, value=None))

    ws.column_dimensions["A"].width = 14
    for p in range(num_personen):
        ws.column_dimensions[get_column_letter(2 + p)].width = 24


def generate_msa_template(gruppenname="", num_personen=3, num_messungen=10,
                          num_wuerfe=10, messmodus="1D", output_path=None) -> str:
    """Erzeugt MSA_Messung_Template.xlsx mit Type-1 und Reproduzierbarkeit."""
    from openpyxl import Workbook
    if num_personen < 2:
        raise ValueError("Es werden mindestens 2 Personen benötigt.")
    messmodus = messmodus.upper()
    if messmodus not in ("1D", "2D"):
        raise ValueError("Messmodus muss '1D' oder '2D' sein.")

    wb = Workbook()
    _msa_build_type1(wb, gruppenname, num_personen, num_messungen, messmodus)
    _msa_build_reproduzierbarkeit(wb, gruppenname, num_personen, num_wuerfe, messmodus)

    if output_path is None:
        output_path = os.path.join(os.getcwd(), "MSA_Messung_Template.xlsx")
    else:
        output_path = os.path.abspath(output_path)
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    print(f"MSA-Template gespeichert: {output_path}")
    return output_path


# ═══════════════════════════════════════════════════════════════════
# 4. MEASURE
# ═══════════════════════════════════════════════════════════════════

# --- 4a. Type-1 Analyse ---

def analysiere_type1(messdaten: pd.DataFrame, referenzwert: float) -> Dict:
    """
    Type-1 Gage Study: Bias und Repeatability.

    messdaten: DataFrame mit Spalten pro Person, Index = Messung-Nr.
    referenzwert: Bekannter Referenzwert (cm)
    """
    ergebnisse = {}
    for col in messdaten.columns:
        werte = messdaten[col].dropna().values.astype(float)
        bias = np.mean(werte) - referenzwert
        repeatability = np.std(werte, ddof=1)
        ergebnisse[col] = {
            "bias": bias,
            "repeatability": repeatability,
            "mean": np.mean(werte),
            "n": len(werte),
        }
    return ergebnisse


def zeige_type1(ergebnisse: Dict):
    """Zeigt die Type-1-Ergebnisse als Tabelle."""
    rows = ""
    for person, data in ergebnisse.items():
        rows += f"""<tr>
            <td style="padding:6px;">{person}</td>
            <td style="padding:6px; text-align:center;">{data['mean']:.2f} cm</td>
            <td style="padding:6px; text-align:center;">{data['bias']:.2f} cm</td>
            <td style="padding:6px; text-align:center;">{data['repeatability']:.2f} cm</td>
        </tr>"""

    html = f"""
    <h4>Type-1 Gage Study: Ergebnisse</h4>
    <table style="border-collapse:collapse; width:100%; border:1px solid #E5E7EB;">
        <tr style="background:{LIGHT_BLUE};">
            <th style="padding:8px;">Person</th>
            <th style="padding:8px;">Mittelwert</th>
            <th style="padding:8px;">Bias (Abweichung)</th>
            <th style="padding:8px;">Repeatability (σ)</th>
        </tr>
        {rows}
    </table>"""
    display(HTML(html))


# --- 4b. Gage R&R (ANOVA-Methode, AIAG-konform) ---

def analysiere_gage_rr(daten: pd.DataFrame) -> Dict:
    """
    AIAG-konforme Gage R&R Analyse (ANOVA-Methode).

    daten: DataFrame mit Spalten:
        - 'Wurf_ID': Auftreffpunkt-Nr.  (=Teil)
        - 'Person':  Name/Nr. des Messenden
        - 'Messwert': Gemessener Wert

    Rückgabe: Dict mit Varianzkomponenten, %GRR, Bewertung.
    """
    import statsmodels.api as sm
    from statsmodels.formula.api import ols

    df = daten.copy()
    df = df.dropna(subset=["Messwert"])
    df["Messwert"] = pd.to_numeric(df["Messwert"], errors="coerce")
    df = df.dropna(subset=["Messwert"])
    df["Wurf_ID"] = df["Wurf_ID"].astype(str)
    df["Person"] = df["Person"].astype(str)

    n_teile = df["Wurf_ID"].nunique()
    n_personen = df["Person"].nunique()

    if len(df) < 3 or n_personen < 2:
        return {"fehler": "Mindestens 2 Personen und 3 Messwerte benötigt."}

    n_wiederholungen = max(1, len(df) // max(1, n_teile * n_personen))

    # Bei nur 1 Messung pro Person/Wurf: additives Modell (ohne Interaktion),
    # da das volle Modell saturiert wäre (df_residual = 0).
    try:
        if n_wiederholungen <= 1:
            model = ols("Messwert ~ C(Wurf_ID) + C(Person)", data=df).fit()
            _hat_interaktion = False
        else:
            model = ols("Messwert ~ C(Wurf_ID) + C(Person) + C(Wurf_ID):C(Person)", data=df).fit()
            _hat_interaktion = True
        anova_table = sm.stats.anova_lm(model, typ=2)
    except Exception as e:
        return {"fehler": str(e)}

    def _safe_ms(source, col="sum_sq"):
        """Sichere Mean-Square-Berechnung (NaN/Inf → 0)."""
        if source not in anova_table.index:
            return 0.0
        ss = anova_table.loc[source, col]
        dof = anova_table.loc[source, "df"]
        if not np.isfinite(ss) or not np.isfinite(dof) or dof <= 0:
            return 0.0
        val = ss / dof
        return val if np.isfinite(val) else 0.0

    ms_teil = _safe_ms("C(Wurf_ID)")
    ms_person = _safe_ms("C(Person)")
    ms_residual = _safe_ms("Residual")

    if _hat_interaktion:
        ms_interact = _safe_ms("C(Wurf_ID):C(Person)")
    else:
        # Ohne Interaktionsterm: Residual enthält Repeatability + Interaktion
        ms_interact = 0

    # Varianzkomponenten (negative auf 0 setzen)
    var_repeatability = max(0, ms_residual)
    var_interact = max(0, (ms_interact - ms_residual) / n_wiederholungen) \
        if ms_interact > 0 else 0
    var_reproducibility = max(0, (ms_person - ms_interact) / (n_teile * n_wiederholungen))
    var_teil = max(0, (ms_teil - ms_interact) / (n_personen * n_wiederholungen))

    var_grr = var_repeatability + var_reproducibility + var_interact
    var_total = var_teil + var_grr

    # %GRR auf σ-Basis (AIAG-konform)
    sigma_grr = np.sqrt(var_grr) if var_grr > 0 else 0
    sigma_total = np.sqrt(var_total) if var_total > 0 else 0
    pct_grr = (sigma_grr / sigma_total * 100) if sigma_total > 0 else 0

    # Bewertung
    if pct_grr < 10:
        bewertung = "✅ Messsystem akzeptabel"
    elif pct_grr < 30:
        bewertung = "⚠️ Messsystem bedingt akzeptabel"
    else:
        bewertung = "❌ Messsystem nicht akzeptabel"

    ergebnis = {
        "var_teil": var_teil,
        "var_repeatability": var_repeatability,
        "var_reproducibility": var_reproducibility,
        "var_interact": var_interact,
        "var_grr": var_grr,
        "var_total": var_total,
        "sigma_grr": sigma_grr,
        "sigma_total": sigma_total,
        "pct_grr": pct_grr,
        "bewertung": bewertung,
        "anova_table": anova_table,
        "n_teile": n_teile,
        "n_personen": n_personen,
        "n_wiederholungen": n_wiederholungen,
    }
    if not _hat_interaktion:
        ergebnis["hinweis"] = (
            "Vereinfachtes Modell (ohne Interaktion): Bei nur 1 Messung pro "
            "Person/Messpunkt kann die Interaktion nicht vom Messfehler getrennt "
            "werden. Die Repeatability enthält daher auch den Interaktionsanteil."
        )
    return ergebnis


def zeige_gage_rr(ergebnis: Dict):
    """Zeigt die Gage R&R Ergebnisse übersichtlich an."""
    if "fehler" in ergebnis:
        display(HTML(f"<div style='color:red;'>Fehler: {ergebnis['fehler']}</div>"))
        return

    # Ampel
    grr_schwellen = [
        (10, "✅", "Messsystem akzeptabel"),
        (30, "⚠️", "Bedingt akzeptabel – abhängig von Anwendung"),
        (float("inf"), "❌", "Nicht akzeptabel – Messmethode überdenken"),
    ]
    zeige_ampel(ergebnis["pct_grr"], grr_schwellen, einheit="%", titel="%GRR:")

    # Varianzkomponenten-Tabelle
    vk = [
        ("Teil-zu-Teil", ergebnis["var_teil"], ergebnis["var_teil"] / ergebnis["var_total"] * 100
         if ergebnis["var_total"] > 0 else 0),
        ("Wiederholbarkeit (Repeatability)", ergebnis["var_repeatability"],
         ergebnis["var_repeatability"] / ergebnis["var_total"] * 100
         if ergebnis["var_total"] > 0 else 0),
        ("Reproduzierbarkeit (Reproducibility)", ergebnis["var_reproducibility"],
         ergebnis["var_reproducibility"] / ergebnis["var_total"] * 100
         if ergebnis["var_total"] > 0 else 0),
        ("Interaktion", ergebnis["var_interact"],
         ergebnis["var_interact"] / ergebnis["var_total"] * 100
         if ergebnis["var_total"] > 0 else 0),
    ]
    rows = ""
    for name, var, pct in vk:
        rows += f"<tr><td style='padding:6px;'>{name}</td>"
        rows += f"<td style='padding:6px; text-align:right;'>{var:.4f}</td>"
        rows += f"<td style='padding:6px; text-align:right;'>{pct:.1f}%</td></tr>"

    html = f"""
    <h4>Varianzkomponenten</h4>
    <table style="border-collapse:collapse; width:80%; border:1px solid #E5E7EB;">
        <tr style="background:{LIGHT_BLUE};">
            <th style="padding:8px; text-align:left;">Quelle</th>
            <th style="padding:8px;">Varianz (σ²)</th>
            <th style="padding:8px;">Anteil</th>
        </tr>
        {rows}
        <tr style="background:#F3F4F6; font-weight:bold;">
            <td style="padding:6px;">Gesamt</td>
            <td style="padding:6px; text-align:right;">{ergebnis['var_total']:.4f}</td>
            <td style="padding:6px; text-align:right;">100%</td>
        </tr>
    </table>
    <p><strong>%GRR (σ-Basis, AIAG):</strong> {ergebnis['pct_grr']:.1f}%</p>
    """
    if "hinweis" in ergebnis:
        html += f"""
        <div style="padding:8px; border-left:3px solid {ORANGE}; background:{LIGHT_YELLOW};
                    border-radius:4px; margin:8px 0; font-size:0.9em;">
            ℹ️ {ergebnis['hinweis']}
        </div>"""
    display(HTML(html))

    # ANOVA-Tabelle als Prio 2
    anova_html = ergebnis["anova_table"].to_html(
        float_format=lambda x: f"{x:.4f}", classes="table"
    )
    zeige_prio2("ANOVA-Detailtabelle der Varianzkomponenten", anova_html)


def plot_msa_boxplot(daten: pd.DataFrame) -> plt.Figure:
    """Boxplot: Messwerte je Person (Reproduzierbarkeit sichtbar)."""
    fig, ax = plt.subplots(figsize=(8, 5))
    personen = daten["Person"].unique()
    box_data = [daten[daten["Person"] == p]["Messwert"].values for p in personen]
    bp = ax.boxplot(box_data, labels=personen, patch_artist=True)
    for patch in bp["boxes"]:
        patch.set_facecolor(LIGHT_BLUE)
        patch.set_edgecolor(BLUE)
    ax.set_xlabel("Messende Person")
    ax.set_ylabel("Messwert (cm)")
    ax.set_title("MSA: Messwerte je Person", fontsize=13, fontweight="bold")
    ax.grid(True, alpha=0.3, axis="y")
    fig.tight_layout()
    return fig


def plot_msa_interaktion(daten: pd.DataFrame) -> plt.Figure:
    """Interaktionsplot: Messwerte je Person je Wurf."""
    fig, ax = plt.subplots(figsize=(10, 5))
    personen = daten["Person"].unique()
    farben = [BLUE, RED, GREEN, ORANGE, GRAY]

    for i, person in enumerate(personen):
        subset = daten[daten["Person"] == person]
        mittel = subset.groupby("Wurf_ID")["Messwert"].mean()
        ax.plot(mittel.index, mittel.values, "o-",
                color=farben[i % len(farben)], label=person, markersize=5)

    ax.set_xlabel("Wurf-ID (Auftreffpunkt)")
    ax.set_ylabel("Messwert (cm)")
    ax.set_title("MSA: Interaktionsplot (Person × Wurf)", fontsize=13, fontweight="bold")
    ax.legend(title="Person", fontsize=9)
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    return fig


def plot_msa_zielscheiben_pro_messer(daten: pd.DataFrame, modus: str = "1D") -> plt.Figure:
    """Zielscheibe pro messende Person (Reproduzierbarkeit sichtbar)."""
    personen = daten["Person"].unique()
    n_pers = len(personen)
    fig, axes = plt.subplots(1, n_pers, figsize=(5 * n_pers, 5))
    if n_pers == 1:
        axes = [axes]

    farben = [BLUE, RED, GREEN, ORANGE, GRAY]

    for ax, person, farbe in zip(axes, personen, farben):
        subset = daten[daten["Person"] == person]
        werte = subset["Messwert"].values
        mu = np.mean(werte)
        sigma = np.std(werte, ddof=1) if len(werte) > 1 else 0

        # 1D-Streifendiagramm
        gesamt_mean = daten["Messwert"].mean()
        ax.axvline(gesamt_mean, color=GRAY, linewidth=2, linestyle="--", label="Gesamtmittel")
        y_jitter = np.random.default_rng(42).uniform(-0.3, 0.3, len(werte))
        ax.scatter(werte, y_jitter, c=farbe, s=50, zorder=5,
                   edgecolors="white", linewidths=0.5)
        ax.axvline(mu, color=farbe, linewidth=1.5, linestyle=":", label=f"\u03bc={mu:.1f}")
        ax.set_xlabel("Messwert (cm)")
        ax.set_yticks([])
        ax.set_ylim(-1, 1)
        ax.set_title(f"{person}\n\u03bc={mu:.1f}, \u03c3={sigma:.1f}", fontsize=11, fontweight="bold")
        ax.legend(fontsize=8, loc="upper right")

    fig.suptitle("MSA: Messwerte pro Person (Zielscheibe)", fontsize=13, fontweight="bold")
    fig.tight_layout(rect=[0, 0, 1, 0.93])
    return fig


# --- 4c. Baseline ---

def analysiere_baseline(wuerfe: np.ndarray) -> Dict:
    """Berechnet Baseline-Statistiken inkl. Shapiro-Wilk."""
    mu = np.mean(wuerfe)
    sigma = np.std(wuerfe, ddof=1)
    shapiro_stat, shapiro_p = stats.shapiro(wuerfe) if len(wuerfe) >= 8 else (np.nan, np.nan)
    return {
        "mean": mu,
        "std": sigma,
        "n": len(wuerfe),
        "shapiro_stat": shapiro_stat,
        "shapiro_p": shapiro_p,
    }


def plot_baseline_histogramm(wuerfe: np.ndarray, ziel: float, toleranz: float) -> plt.Figure:
    """Histogramm mit Normalverteilungskurve + Shapiro-Wilk."""
    fig, ax = plt.subplots(figsize=(9, 5))
    mu, sigma = np.mean(wuerfe), np.std(wuerfe, ddof=1)

    ax.hist(wuerfe, bins="auto", density=True, color=LIGHT_BLUE,
            edgecolor=BLUE, linewidth=1.2, alpha=0.8)

    x = np.linspace(mu - 4 * sigma, mu + 4 * sigma, 200)
    ax.plot(x, stats.norm.pdf(x, mu, sigma), color=RED, linewidth=2,
            label=f"N(μ={mu:.1f}, σ={sigma:.1f})")

    # Spezifikationsgrenzen
    ax.axvline(ziel, color=GREEN, linewidth=2, linestyle="--", label=f"Ziel = {ziel:.0f} cm")
    ax.axvline(ziel - toleranz, color=RED, linewidth=1.5, linestyle=":",
               label=f"LSL = {ziel - toleranz:.0f}")
    ax.axvline(ziel + toleranz, color=RED, linewidth=1.5, linestyle=":",
               label=f"USL = {ziel + toleranz:.0f}")

    _, shapiro_p = stats.shapiro(wuerfe) if len(wuerfe) >= 8 else (np.nan, np.nan)
    shapiro_str = f"p = {shapiro_p:.3f}" if not np.isnan(shapiro_p) else "n zu klein"
    ax.set_title(f"Baseline-Histogramm (n={len(wuerfe)}) | Shapiro-Wilk: {shapiro_str}",
                 fontsize=13, fontweight="bold")
    ax.set_xlabel("Wurfweite (cm)")
    ax.set_ylabel("Dichte")
    ax.legend(fontsize=9)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    fig.tight_layout()
    return fig


# ═══════════════════════════════════════════════════════════════════
# 5. ANALYZE
# ═══════════════════════════════════════════════════════════════════

# --- 5a. DoE-Generierung ---

def _berechne_konfundierung(k: int, p: int) -> str:
    """Erzeugt Konfundierungshinweis für 2^(k-p) Designs."""
    buchstaben = [chr(65 + i) for i in range(k)]
    if p == 1:
        gen = "".join(buchstaben[:k - 1])
        resolution = k if k <= 5 else 4
        return (
            f"Design: 2^({k}-1), Resolution {['', '', '', 'III', 'IV', 'V'][min(resolution, 5)]}.\n"
            f"Generator: {buchstaben[-1]} = {gen}.\n"
            f"Haupteffekte sind frei von Zweifach-Interaktionen."
            if resolution >= 4 else
            f"Design: 2^({k}-1), Resolution III.\n"
            f"Generator: {buchstaben[-1]} = {gen}.\n"
            f"⚠️ Haupteffekte sind mit Zweifach-Interaktionen konfundiert!"
        )
    elif p == 2:
        resolution = 3 if k <= 5 else 4
        return (
            f"Design: 2^({k}-2), Resolution III.\n"
            f"⚠️ Haupteffekte sind mit Zweifach-Interaktionen konfundiert!\n"
            f"Dieses Design eignet sich nur zum Screening (viele Faktoren, wenig Runs)."
        )
    return f"2^({k}-{p}) fraktionelles Design."


def generiere_versuchsplan(
    faktoren: List[Dict],
    wiederholungen: int = 3,
    blocking: bool = False,
    centerpoints: int = 3,
    seed: int = 42,
    design: str = "voll",
) -> pd.DataFrame:
    """
    Generiert einen randomisierten Versuchsplan.

    faktoren: Liste von Dicts mit 'name', 'einheit', 'low', 'high'
    wiederholungen: Anzahl Wiederholungen (1-10)
    blocking: Ob in 2 Blöcke aufgeteilt wird
    centerpoints: Anzahl Centerpoints (0+)
    design: "voll" (2^k), "halb" (2^(k-1)), "viertel" (2^(k-2))

    Rückgabe: DataFrame mit Versuchsplan (kodiert und Original)
    """
    from itertools import product as _product

    k = len(faktoren)
    if k < 3:
        raise ValueError("Mindestens 3 Faktoren erforderlich.")

    for f in faktoren:
        if f["low"] == f["high"]:
            raise ValueError(f"Faktor '{f['name']}': Low ({f['low']}) und High ({f['high']}) sind identisch.")
        if f["low"] > f["high"]:
            f["low"], f["high"] = f["high"], f["low"]

    konfundierung = ""

    # Design-Auswahl
    if design == "voll":
        design_coded = np.array(list(_product([-1, 1], repeat=k)))
        design_typ = f"2^{k} vollfaktoriell ({2**k} Runs)"
    elif design == "halb":
        # 2^(k-1): letzter Faktor = Produkt aller Basisfaktoren
        basis = np.array(list(_product([-1, 1], repeat=k - 1)))
        generated = np.prod(basis, axis=1, keepdims=True)
        design_coded = np.hstack([basis, generated])
        design_typ = f"2^({k}-1) halbfraktionell ({2**(k-1)} Runs)"
        konfundierung = _berechne_konfundierung(k, 1)
    elif design == "viertel":
        if k < 4:
            print("⚠️ Viertelfraktionell braucht ≥4 Faktoren. Nutze vollfaktoriell.")
            design_coded = np.array(list(_product([-1, 1], repeat=k)))
            design_typ = f"2^{k} vollfaktoriell ({2**k} Runs)"
        else:
            # 2^(k-2): zwei Generatoren
            n_basis = k - 2
            basis = np.array(list(_product([-1, 1], repeat=n_basis)))
            # Generator 1: vorletzter Faktor = Produkt aller Basisfaktoren
            gen1 = np.prod(basis, axis=1, keepdims=True)
            # Generator 2: letzter Faktor = Produkt der letzten (n_basis-1) Basisfaktoren
            gen2 = np.prod(basis[:, 1:], axis=1, keepdims=True)
            design_coded = np.hstack([basis, gen1, gen2])
            design_typ = f"2^({k}-2) viertelfraktionell ({2**(k-2)} Runs)"
            konfundierung = _berechne_konfundierung(k, 2)
    else:
        raise ValueError(f"Unbekanntes Design: '{design}'. Erlaubt: 'voll', 'halb', 'viertel'.")

    n_basis = len(design_coded)

    # Centerpoints: nur für Faktoren mit centerpoint_moeglich=True
    # Binäre Faktoren werden im Centerpoint auf Low (-1) gesetzt
    cp_mask = [f.get("centerpoint_moeglich", True) for f in faktoren]
    hat_cp_faktoren = any(cp_mask)

    if centerpoints > 0 and hat_cp_faktoren:
        cp_row = np.array([-1.0 if not cp_ok else 0.0 for cp_ok in cp_mask])
        cp = np.tile(cp_row, (centerpoints, 1))
        design_with_cp = np.vstack([design_coded, cp])
    else:
        centerpoints = 0
        design_with_cp = design_coded

    # Wiederholungen
    design_repeated = np.tile(design_with_cp, (wiederholungen, 1))

    # DataFrame erstellen
    columns_coded = [f"{f['name']}_coded" for f in faktoren]
    columns_original = [f"{f['name']} ({f['einheit']})" for f in faktoren]

    df = pd.DataFrame(design_repeated, columns=columns_coded)

    # Originale Werte berechnen
    for i, f in enumerate(faktoren):
        col_coded = columns_coded[i]
        col_orig = columns_original[i]
        low, high = f["low"], f["high"]
        center = (low + high) / 2
        half_range = (high - low) / 2
        df[col_orig] = center + df[col_coded] * half_range

    # Versuchsnummer
    df.insert(0, "Versuch_Nr", range(1, len(df) + 1))

    # Blocking
    if blocking:
        block_size = len(df) // 2
        df["Block"] = [1] * block_size + [2] * (len(df) - block_size)
    else:
        df["Block"] = 1

    # Randomisierung innerhalb Blöcke
    rng = np.random.RandomState(seed)
    randomized_parts = []
    for block in df["Block"].unique():
        block_df = df[df["Block"] == block].copy()
        block_df = block_df.sample(frac=1, random_state=rng).reset_index(drop=True)
        randomized_parts.append(block_df)
    df = pd.concat(randomized_parts, ignore_index=True)
    df["Versuch_Nr"] = range(1, len(df) + 1)

    # Ist-Centerpoint markieren
    if hat_cp_faktoren and centerpoints > 0:
        # Centerpoints: stetige Faktoren auf 0, binäre auf -1
        cp_check = pd.Series(True, index=df.index)
        for i, f in enumerate(faktoren):
            col = columns_coded[i]
            if f.get("centerpoint_moeglich", True):
                cp_check &= (df[col] == 0)
            else:
                cp_check &= (df[col] == -1)
        df["Typ"] = np.where(cp_check, "Centerpoint", "Eckpunkt")
    else:
        df["Typ"] = "Eckpunkt"

    # Metadaten
    df.attrs["design_typ"] = design_typ
    df.attrs["n_basis"] = n_basis
    df.attrs["wiederholungen"] = wiederholungen
    df.attrs["centerpoints"] = centerpoints
    df.attrs["blocking"] = blocking
    df.attrs["n_gesamt"] = len(df)
    df.attrs["konfundierung"] = konfundierung

    return df


def zeige_versuchsplan_info(df: pd.DataFrame, faktoren: List[Dict]):
    """Zeigt eine Zusammenfassung des Versuchsplans."""
    attrs = df.attrs
    html = f"""
    <div style="padding:12px; background:{LIGHT_BLUE}; border-radius:6px; margin:10px 0;">
        <h4 style="margin:0 0 8px 0;">Versuchsplan: {attrs.get('design_typ', '?')}</h4>
        <ul style="margin:0;">
            <li><strong>Faktoren:</strong> {len(faktoren)} ({', '.join(f['name'] for f in faktoren)})</li>
            <li><strong>Basisversuche:</strong> {attrs.get('n_basis', '?')} + {attrs.get('centerpoints', '?')} Centerpoints</li>
            <li><strong>Wiederholungen:</strong> {attrs.get('wiederholungen', '?')}</li>
            <li><strong>Blocking:</strong> {'Ja (2 Blöcke)' if attrs.get('blocking') else 'Nein'}</li>
            <li><strong>Gesamtversuche:</strong> {attrs.get('n_gesamt', len(df))}</li>
            <li><strong>Geschätzte Dauer:</strong> ~{len(df) * 2} Minuten ({len(df)} × 2 min/Wurf)</li>
        </ul>
    </div>"""
    display(HTML(html))

    # Zeitwarnung
    if len(df) > 40:
        display(HTML(f"""
        <div style="padding:10px; border-left:4px solid {ORANGE}; background:{LIGHT_YELLOW};
                     border-radius:4px; margin:8px 0;">
            ⚠️ <strong>Hinweis:</strong> {len(df)} Versuche erfordern ca. {len(df)*2} Minuten.
            Prüft, ob das zeitlich machbar ist!
        </div>"""))

    # Konfundierungsmatrix bei fraktionellen Designs
    konfundierung = df.attrs.get("konfundierung", "")
    if konfundierung:
        display(HTML(f"""
        <details style="margin:10px 0; padding:8px; background:#F9FAFB; border:1px solid #E5E7EB; border-radius:6px;">
            <summary style="cursor:pointer; font-weight:bold; color:{BLUE};">
                \U0001f50d Konfundierungsmatrix (fraktionelles Design)
            </summary>
            <div style="margin-top:8px; padding:8px; font-size:0.95em;">
                <p>Bei einem fraktionellen Design sind bestimmte Effekte nicht unabh\u00e4ngig voneinander sch\u00e4tzbar (konfundiert):</p>
                <p><code>{konfundierung}</code></p>
                <p><strong>Bedeutung:</strong> Falls eine konfundierte Interaktion signifikant erscheint, kann das Notebook nicht unterscheiden, welcher der konfundierten Effekte der wahre Verursacher ist. Das sollte im Bericht diskutiert werden.</p>
            </div>
        </details>"""))


def erstelle_doe_excel(versuchsplan: pd.DataFrame, faktoren: List[Dict],
                       output_path: str = "DoE_Versuchsergebnisse.xlsx") -> str:
    """Erstellt die DoE-Versuchsergebnisse-Excel-Vorlage."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Versuchsergebnisse"

    # Header-Style
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Spalten: Versuch_Nr, Block, Faktor1_coded, ..., Faktor1_orig, ..., Ergebnis (Weite cm)
    headers = ["Versuch_Nr", "Block"]
    for f in faktoren:
        headers.append(f"{f['name']} (kodiert)")
    for f in faktoren:
        headers.append(f"{f['name']} ({f['einheit']})")
    headers.append("Ergebnis: Weite (cm)")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Daten
    coded_cols = [f"{f['name']}_coded" for f in faktoren]
    orig_cols = [f"{f['name']} ({f['einheit']})" for f in faktoren]

    for row_idx, (_, row) in enumerate(versuchsplan.iterrows(), start=2):
        ws.cell(row=row_idx, column=1, value=int(row["Versuch_Nr"])).border = thin_border
        ws.cell(row=row_idx, column=2, value=int(row["Block"])).border = thin_border

        for ci, col in enumerate(coded_cols):
            c = ws.cell(row=row_idx, column=3 + ci, value=float(row[col]))
            c.border = thin_border
        for ci, col in enumerate(orig_cols):
            c = ws.cell(row=row_idx, column=3 + len(coded_cols) + ci, value=float(row[col]))
            c.border = thin_border

        # Ergebnis-Spalte leer (Studis füllen aus)
        c = ws.cell(row=row_idx, column=3 + 2 * len(coded_cols))
        c.border = thin_border

    # Spaltenbreiten
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + min(col_idx, 26))].width = 18

    wb.save(output_path)
    return os.path.abspath(output_path)


# --- 5b. Modell-Fitting ---

def _finde_kodierte_spalte(df_columns: list, faktor_name: str) -> Optional[str]:
    """Findet die kodierte Spalte für einen Faktor, unabhängig vom Benennungsmuster."""
    # Mögliche Muster: "Name_coded", "Name (kodiert)", "Name_kodiert"
    kandidaten = [
        f"{faktor_name}_coded",
        f"{faktor_name} (kodiert)",
        f"{faktor_name}_kodiert",
    ]
    for k in kandidaten:
        if k in df_columns:
            return k
    # Fuzzy: Spalte die den Faktornamen enthält + "cod" oder "kod"
    for col in df_columns:
        if faktor_name.lower() in col.lower() and ("cod" in col.lower() or "kod" in col.lower()):
            return col
    return None


def _parse_faktoren_aus_excel(daten: pd.DataFrame) -> List[Dict]:
    """Erkennt Faktoren automatisch aus den Spaltenheadern einer DoE-Excel-Datei."""
    faktoren = []
    for col in daten.columns:
        # Pattern: "Name (kodiert)" oder "Name_coded"
        if "(kodiert)" in col:
            name = col.replace(" (kodiert)", "").strip()
        elif col.endswith("_coded"):
            name = col[:-6].strip()
        else:
            continue
        # Einheit und Originalwerte aus korrespondierender Spalte suchen
        einheit = "?"
        low, high = -1.0, 1.0
        for orig_col in daten.columns:
            if orig_col.startswith(name) and orig_col != col and "kodiert" not in orig_col and "coded" not in orig_col:
                # Einheit aus Klammer extrahieren: "Winkel (Grad)" → "Grad"
                import re
                m = re.search(r"\(([^)]+)\)", orig_col)
                if m:
                    einheit = m.group(1)
                vals = daten[orig_col].dropna()
                if len(vals) > 0:
                    low = float(vals.min())
                    high = float(vals.max())
                break
        faktoren.append({"name": name, "einheit": einheit, "low": low, "high": high})
    return faktoren


def fitte_modell(daten: pd.DataFrame, faktoren: List[Dict],
                 mit_interaktionen: bool = True,
                 mit_drei_faktor_interaktionen: bool = False,
                 mit_quadratischen_termen: str = "auto") -> Any:
    """
    Fittet ein OLS-Regressionsmodell.

    Standard (2FI): ŷ = β₀ + Σβᵢxᵢ + Σβᵢⱼxᵢxⱼ + ε
    RSM (mit x²):  ŷ = β₀ + Σβᵢxᵢ + Σβᵢᵢxᵢ² + Σβᵢⱼxᵢxⱼ + ε

    daten: DataFrame mit kodierten Faktorspalten + 'Ergebnis' Spalte
    faktoren: Liste der Faktor-Dicts
    mit_interaktionen: 2-Faktor-Interaktionen einbeziehen (Standard: True)
    mit_drei_faktor_interaktionen: 3-Faktor-Interaktionen einbeziehen
    mit_quadratischen_termen: Quadratische Terme xᵢ² einbeziehen.
        - "auto" (Standard): Prüft ob Centerpoints vorhanden sind. Wenn ja,
          fittet beide Modelle (linear + quadratisch) und wählt das bessere
          anhand R²_adj. Wenn keine Centerpoints: nur lineares Modell.
        - True: Erzwingt quadratische Terme (Warnung wenn keine Centerpoints)
        - False: Kein quadratisches Modell
    """
    from statsmodels.formula.api import ols

    # Auto-Detect: Faktoren aus Excel-Headern parsen wenn Liste leer
    if not faktoren:
        faktoren = _parse_faktoren_aus_excel(daten)
        if not faktoren:
            raise ValueError(
                "Keine Faktoren angegeben und keine kodierten Spalten "
                "(z.B. 'Name (kodiert)') in den Daten gefunden.\n"
                f"Vorhandene Spalten: {daten.columns.tolist()}"
            )
        print(f"ℹ️ {len(faktoren)} Faktoren automatisch aus Excel erkannt: "
              f"{', '.join(f['name'] for f in faktoren)}")

    # Faktor-Spaltennamen (kodiert) robust erkennen, lesbar benennen
    import re
    faktor_namen = []
    rename_map = {}
    fehlende = []
    _used_names = set()
    for i, f in enumerate(faktoren):
        # Lesbarer Name: Sonderzeichen → Unterstrich, aber Grundname beibehalten
        clean = re.sub(r"[^A-Za-z0-9äöüÄÖÜß]", "_", f["name"]).strip("_")
        clean = re.sub(r"_+", "_", clean)
        if not clean or not clean[0].isalpha():
            clean = f"F{i+1}_{clean}"
        # Duplikate vermeiden
        base = clean
        cnt = 2
        while clean in _used_names:
            clean = f"{base}_{cnt}"
            cnt += 1
        _used_names.add(clean)

        gefunden = _finde_kodierte_spalte(daten.columns.tolist(), f["name"])
        if gefunden:
            rename_map[gefunden] = clean
        else:
            fehlende.append(f["name"])
        faktor_namen.append(clean)

    if fehlende:
        raise ValueError(
            f"Kodierte Spalten nicht gefunden für: {fehlende}\n"
            f"Vorhandene Spalten: {daten.columns.tolist()}"
        )

    df = daten.rename(columns=rename_map).copy()

    # Ergebnis-Spalte finden
    y_col = None
    for col in df.columns:
        col_lower = col.lower()
        if "ergebnis" in col_lower or "result" in col_lower or "weite" in col_lower:
            y_col = col
            break
    if y_col:
        df = df.rename(columns={y_col: "Y"})
    else:
        # Letzte numerische Spalte als Y
        num_cols = df.select_dtypes(include=[np.number]).columns
        df = df.rename(columns={num_cols[-1]: "Y"})

    # NaN-Zeilen entfernen
    df = df.dropna(subset=["Y"] + faktor_namen)

    # --- Centerpoint-Erkennung ---
    # Centerpoints haben kodierte Werte nahe 0 (nicht ±1)
    hat_centerpoints = False
    if len(df) > 0:
        coded_abs = df[faktor_namen].abs()
        cp_mask = (coded_abs < 0.5).all(axis=1)
        n_centerpoints = cp_mask.sum()
        hat_centerpoints = n_centerpoints >= 2

    # --- Entscheidung: quadratische Terme testen? ---
    teste_quad = False
    if mit_quadratischen_termen == "auto":
        teste_quad = hat_centerpoints
    elif mit_quadratischen_termen is True:
        if not hat_centerpoints:
            print("⚠️ Quadratische Terme angefordert, aber keine Centerpoints "
                  "gefunden. x²-Koeffizienten sind ohne Centerpoints nicht "
                  "zuverlässig schätzbar.")
        teste_quad = True

    # --- Quadratische Spalten für ALLE Faktoren vorbereiten ---
    alle_quad = []
    for name in faktor_namen:
        sq_name = f"{name}_sq"
        df[sq_name] = df[name] ** 2
        alle_quad.append(sq_name)

    # --- Formel aufbauen ---
    def _build_formula(fn, quad, mit_2fi, mit_3fi):
        terme = [" + ".join(fn)]
        if quad:
            terme.append(" + ".join(quad))
        if mit_2fi:
            zweier = [f"{fn[i]}:{fn[j]}"
                      for i in range(len(fn)) for j in range(i + 1, len(fn))]
            if zweier:
                terme.append(" + ".join(zweier))
        if mit_3fi and len(fn) >= 3:
            dreier = [f"{fn[i]}:{fn[j]}:{fn[k_]}"
                      for i in range(len(fn)) for j in range(i + 1, len(fn))
                      for k_ in range(j + 1, len(fn))]
            if dreier:
                terme.append(" + ".join(dreier))
        return "Y ~ " + " + ".join(terme)

    # --- Modell fitten ---
    # Schritt 1: Lineares Basismodell (immer)
    formel_lin = _build_formula(faktor_namen, [], mit_interaktionen,
                                mit_drei_faktor_interaktionen)
    model_lin = ols(formel_lin, data=df).fit()

    # Schritt 2: Per-Faktor quadratische Terme testen
    quad_namen = []
    if teste_quad:
        # Volles quadratisches Modell fitten
        formel_voll = _build_formula(faktor_namen, alle_quad, mit_interaktionen,
                                     mit_drei_faktor_interaktionen)
        model_voll = ols(formel_voll, data=df).fit()

        # Jeden x²-Term einzeln bewerten: behalten wenn p < 0.10
        alpha_quad = 0.10
        behalten = []
        verworfen = []
        for sq_name in alle_quad:
            if sq_name in model_voll.pvalues:
                p = model_voll.pvalues[sq_name]
                if p < alpha_quad:
                    behalten.append(sq_name)
                else:
                    verworfen.append(sq_name)

        if behalten:
            quad_namen = behalten
            # Finales Modell nur mit signifikanten x²-Termen
            formel_final = _build_formula(faktor_namen, quad_namen,
                                          mit_interaktionen,
                                          mit_drei_faktor_interaktionen)
            model = ols(formel_final, data=df).fit()
            # Faktor-Namen aus den sq-Namen ableiten (z.B. "X_sq" → "X")
            namen_kurz = [s.replace("_sq", "") for s in behalten]
            print(f"ℹ️ Krümmung erkannt bei: {', '.join(namen_kurz)} "
                  f"(R²_adj: {model.rsquared_adj:.4f} "
                  f"vs. linear {model_lin.rsquared_adj:.4f})")
            if verworfen:
                namen_verw = [s.replace("_sq", "") for s in verworfen]
                print(f"   Kein x²-Term nötig für: {', '.join(namen_verw)}")
        else:
            model = model_lin
            if mit_quadratischen_termen == "auto":
                print(f"ℹ️ Kein Faktor zeigt signifikante Krümmung — "
                      f"lineares Modell beibehalten (R²_adj: {model_lin.rsquared_adj:.4f})")
    else:
        model = model_lin

    # Metadaten anhängen
    model._faktor_namen = faktor_namen
    model._faktor_details = faktoren
    model._rename_map = {v: k for k, v in rename_map.items()}
    model._quad_namen = quad_namen
    model._daten = df

    return model


def hierarchisches_pruning(modell, alpha: float = 0.05) -> Tuple[Any, List[str]]:
    """
    Backward Elimination mit Hierarchie-Schutz.

    Regel: Ein Haupteffekt bleibt im Modell, wenn eine signifikante
    Interaktion existiert, die diesen Haupteffekt enthält.

    Rückgabe: (gepruntes_modell, log_nachrichten)
    """
    from statsmodels.formula.api import ols

    df = modell._daten.copy()
    faktor_namen = modell._faktor_namen
    # Metadaten vom Original-Modell sichern
    _orig_faktor_details = getattr(modell, "_faktor_details", [])
    _orig_rename_map = getattr(modell, "_rename_map", {})
    _orig_quad_namen = getattr(modell, "_quad_namen", [])
    log = []

    # Aktuelle Terme sammeln
    current_terms = [t for t in modell.params.index if t != "Intercept"]

    while True:
        # Finde Term mit höchstem p-Wert
        p_values = modell.pvalues.drop("Intercept", errors="ignore")
        p_values = p_values[p_values.index.isin(current_terms)]

        if p_values.empty:
            break

        worst_term = p_values.idxmax()
        worst_p = p_values[worst_term]

        if worst_p <= alpha:
            break  # Alle signifikant

        # Hierarchie-Check: Ist worst_term ein Haupteffekt?
        is_haupteffekt = ":" not in worst_term
        if is_haupteffekt:
            # Prüfe ob eine signifikante Interaktion diesen Haupteffekt enthält
            protected = False
            for term in current_terms:
                if ":" in term and worst_term in term.split(":"):
                    if modell.pvalues.get(term, 1.0) <= alpha:
                        protected = True
                        log.append(
                            f"⚙️ {worst_term} bleibt im Modell (p={worst_p:.3f}): "
                            f"geschützt durch signifikante Interaktion {term}"
                        )
                        break
            if protected:
                # Markiere als geprüft, versuche nächsten
                p_values = p_values.drop(worst_term)
                if p_values.empty:
                    break
                # Versuche den nächsthöheren p-Wert
                worst_term = p_values.idxmax()
                worst_p = p_values[worst_term]
                if worst_p <= alpha:
                    break
                if ":" not in worst_term:
                    # Auch diesen prüfen...
                    continue

        # Term entfernen
        current_terms.remove(worst_term)
        log.append(f"❌ {worst_term} entfernt (p={worst_p:.3f})")

        if not current_terms:
            log.append("⚠️ Alle Terme entfernt – Modell nur mit Intercept")
            break

        # Neues Modell fitten
        formel = "Y ~ " + " + ".join(current_terms)
        modell = ols(formel, data=df).fit()
        modell._faktor_namen = faktor_namen
        modell._faktor_details = _orig_faktor_details
        modell._rename_map = _orig_rename_map
        modell._quad_namen = _orig_quad_namen
        modell._daten = df

    if not log:
        log.append("✅ Alle Terme sind signifikant – kein Pruning nötig")

    return modell, log


# --- 5c. Modell-Ausgaben ---

def plot_pareto_effekte(modell, titel="Pareto-Diagramm der standardisierten Effekte") -> plt.Figure:
    """Erstellt ein Pareto-Diagramm der standardisierten Effekte."""
    # t-Werte (standardisierte Koeffizienten)
    t_values = modell.tvalues.drop("Intercept", errors="ignore")
    abs_t = t_values.abs().sort_values(ascending=True)

    fig, ax = plt.subplots(figsize=(9, max(4, len(abs_t) * 0.5 + 1)))

    # Signifikanzschwelle
    from scipy.stats import t as t_dist
    df_resid = modell.df_resid
    t_crit = t_dist.ppf(0.975, df_resid) if df_resid > 0 else 2.0

    # Farben: signifikant = blau, nicht signifikant = grau
    colors = [BLUE if v > t_crit else GRAY for v in abs_t.values]

    bars = ax.barh(range(len(abs_t)), abs_t.values, color=colors, edgecolor="white")
    ax.axvline(t_crit, color=RED, linewidth=2, linestyle="--",
               label=f"Signifikanzschwelle (α=0,05, t={t_crit:.2f})")

    # Labels
    labels = []
    faktor_details = getattr(modell, "_faktor_details", [])
    for term in abs_t.index:
        if ":" in term:
            parts = term.split(":")
            label = " × ".join(parts)
        else:
            label = term
        # Versuche lesbaren Namen
        for i, fn in enumerate(getattr(modell, "_faktor_namen", [])):
            if fn in label and i < len(faktor_details):
                label = label.replace(fn, faktor_details[i]["name"])
        labels.append(label)

    ax.set_yticks(range(len(abs_t)))
    ax.set_yticklabels(labels, fontsize=10)
    ax.set_xlabel("|t-Wert| (standardisierter Effekt)")
    ax.set_title(titel, fontsize=13, fontweight="bold")
    ax.legend(loc="lower right", fontsize=9)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    fig.tight_layout()
    return fig


def koeffizienten_tabelle(modell) -> str:
    """Erzeugt eine HTML-Tabelle der Koeffizienten mit Ampel."""
    params = modell.params.drop("Intercept", errors="ignore")
    pvalues = modell.pvalues.drop("Intercept", errors="ignore")
    faktor_details = getattr(modell, "_faktor_details", [])
    faktor_namen = getattr(modell, "_faktor_namen", [])

    rows = ""
    for term in params.index:
        coef = params[term]
        p = pvalues[term]
        sig = "✅" if p < 0.05 else "⚠️"
        farbe = GREEN if p < 0.05 else ORANGE

        # Lesbarer Name
        label = term
        for i, fn in enumerate(faktor_namen):
            if fn in label and i < len(faktor_details):
                label = label.replace(fn, faktor_details[i]["name"])
        label = label.replace(":", " × ")

        rows += f"""<tr>
            <td style="padding:6px;">{label}</td>
            <td style="padding:6px; text-align:right;">{coef:+.2f}</td>
            <td style="padding:6px; text-align:right;">{p:.4f}</td>
            <td style="padding:6px; text-align:center; color:{farbe};">{sig}</td>
        </tr>"""

    # Intercept
    rows = f"""<tr style="background:#F3F4F6;">
        <td style="padding:6px;">Intercept (β₀)</td>
        <td style="padding:6px; text-align:right;">{modell.params['Intercept']:+.2f}</td>
        <td style="padding:6px; text-align:right;">{modell.pvalues['Intercept']:.4f}</td>
        <td style="padding:6px; text-align:center;">–</td>
    </tr>""" + rows

    return f"""
    <h4>Koeffiziententabelle</h4>
    <table style="border-collapse:collapse; width:90%; border:1px solid #E5E7EB;">
        <tr style="background:{LIGHT_BLUE};">
            <th style="padding:8px; text-align:left;">Term</th>
            <th style="padding:8px;">Koeffizient</th>
            <th style="padding:8px;">p-Wert</th>
            <th style="padding:8px;">Signifikant?</th>
        </tr>
        {rows}
    </table>
    <p style="font-size:0.9em; color:{GRAY};">
        Positiver Koeffizient: Faktor erhöht die Wurfweite (Low→High).
        Negativer Koeffizient: Faktor verringert die Wurfweite.
    </p>"""


def zeige_koeffizienten(modell):
    """Zeigt die Koeffiziententabelle an."""
    display(HTML(koeffizienten_tabelle(modell)))


def zeige_modellguete(modell):
    """Zeigt R², Adj. R², RMSE mit Ampel-Bewertung."""
    r2 = modell.rsquared
    adj_r2 = modell.rsquared_adj
    rmse = np.sqrt(modell.mse_resid)

    r2_schwellen = [
        (0.5, "❌", "Modell erklärt weniger als die Hälfte der Variation"),
        (0.8, "⚠️", "Akzeptabel – wichtige Faktoren gefunden, aber Rauschen vorhanden"),
        (float("inf"), "✅", "Sehr gutes Modell – Faktoren erklären die Wurfweite gut"),
    ]
    zeige_ampel(r2, r2_schwellen, titel="R²:")

    html = f"""
    <table style="border-collapse:collapse; width:60%; border:1px solid #E5E7EB; margin:8px 0;">
        <tr style="background:{LIGHT_BLUE};">
            <th style="padding:8px;">Kennzahl</th>
            <th style="padding:8px;">Wert</th>
        </tr>
        <tr><td style="padding:6px;">R²</td><td style="padding:6px; text-align:right;">{r2:.4f}</td></tr>
        <tr><td style="padding:6px;">Adj. R²</td><td style="padding:6px; text-align:right;">{adj_r2:.4f}</td></tr>
        <tr><td style="padding:6px;">RMSE</td><td style="padding:6px; text-align:right;">{rmse:.2f} cm</td></tr>
    </table>"""
    display(HTML(html))


def pruefe_vif(modell) -> Dict:
    """Berechnet VIF (Variance Inflation Factor) für jeden Faktor."""
    from statsmodels.stats.outliers_influence import variance_inflation_factor

    df = modell._daten.copy()
    faktor_namen = modell._faktor_namen
    X = df[faktor_namen].values
    vifs = {}
    warnungen = []

    for i, name in enumerate(faktor_namen):
        try:
            vif = variance_inflation_factor(X, i)
        except Exception:
            vif = np.nan
        vifs[name] = vif
        if vif > 5:
            warnungen.append(name)

    if warnungen:
        namen = ", ".join(warnungen)
        display(HTML(f"""
        <div style="padding:10px; border-left:4px solid {ORANGE}; background:{LIGHT_YELLOW};
                     border-radius:4px; margin:8px 0;">
            ⚠️ <strong>VIF-Warnung:</strong> Faktor(en) {namen} zeigen VIF > 5
            (Multikollinearität). Prüft, ob alle Versuche korrekt durchgeführt und
            eingetragen wurden.
        </div>"""))

    return vifs


def pruefe_lack_of_fit(modell, daten: pd.DataFrame) -> Dict:
    """Prüft Lack-of-Fit über Centerpoints."""
    df = daten.copy()
    faktor_namen = modell._faktor_namen

    # Centerpoints identifizieren
    if "Typ" in df.columns:
        cp_data = df[df["Typ"] == "Centerpoint"]
    else:
        cp_mask = (df[faktor_namen] == 0).all(axis=1)
        cp_data = df[cp_mask]

    if len(cp_data) < 2:
        return {"test_moeglich": False, "grund": "Zu wenig Centerpoints"}

    # Vorhersage am Centerpoint
    pred_at_center = modell.params["Intercept"]  # Alle kodierten Werte = 0

    # Tatsächliche Centerpoint-Werte
    if "Y" in cp_data.columns:
        cp_values = cp_data["Y"].values
    else:
        cp_values = cp_data.iloc[:, -1].values

    cp_mean = np.mean(cp_values)
    cp_std = np.std(cp_values, ddof=1) if len(cp_values) > 1 else 0

    # Einfacher t-Test: Centerpoint-Mittelwert vs. Vorhersage
    if cp_std > 0:
        t_stat = (cp_mean - pred_at_center) / (cp_std / np.sqrt(len(cp_values)))
        p_value = 2 * (1 - stats.t.cdf(abs(t_stat), len(cp_values) - 1))
    else:
        t_stat, p_value = 0, 1.0

    kruemmung = p_value < 0.05

    if kruemmung:
        display(HTML(f"""
        <div style="padding:10px; border-left:4px solid {ORANGE}; background:{LIGHT_YELLOW};
                     border-radius:4px; margin:8px 0;">
            ⚠️ <strong>Krümmungshinweis:</strong> Der Lack-of-Fit-Test zeigt signifikante
            Krümmung (p={p_value:.3f}). Euer Modell zeigt Hinweise auf nichtlineare
            Zusammenhänge. Optional: Design um Sternpunkte erweitern (CCD).
        </div>"""))

    return {
        "test_moeglich": True,
        "cp_mean": cp_mean,
        "pred_at_center": pred_at_center,
        "t_stat": t_stat,
        "p_value": p_value,
        "kruemmung": kruemmung,
    }


def plot_residuen(modell) -> plt.Figure:
    """Residuenplots (Prio 2): Residuen vs. Fitted, Q-Q, vs. Reihenfolge."""
    residuals = modell.resid
    fitted = modell.fittedvalues

    fig, axes = plt.subplots(1, 3, figsize=(15, 4.5))

    # 1. Residuen vs. Fitted
    ax = axes[0]
    ax.scatter(fitted, residuals, c=BLUE, s=30, alpha=0.7, edgecolors="white")
    ax.axhline(0, color=RED, linewidth=1, linestyle="--")
    ax.set_xlabel("Vorhergesagt (ŷ)")
    ax.set_ylabel("Residuum (e)")
    ax.set_title("Residuen vs. Fitted", fontsize=11, fontweight="bold")

    # 2. Q-Q-Plot
    ax = axes[1]
    stats.probplot(residuals, plot=ax)
    ax.get_lines()[0].set(color=BLUE, markersize=4)
    ax.get_lines()[1].set(color=RED)
    ax.set_title("Q-Q-Plot", fontsize=11, fontweight="bold")

    # 3. Residuen vs. Reihenfolge
    ax = axes[2]
    ax.plot(range(1, len(residuals) + 1), residuals, "o-", color=BLUE, markersize=4)
    ax.axhline(0, color=RED, linewidth=1, linestyle="--")
    ax.set_xlabel("Versuchsreihenfolge")
    ax.set_ylabel("Residuum")
    ax.set_title("Residuen vs. Reihenfolge", fontsize=11, fontweight="bold")

    for ax in axes:
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)

    fig.suptitle("Residuenanalyse (Prio 2)", fontsize=13, fontweight="bold")
    fig.tight_layout(rect=[0, 0, 1, 0.93])
    return fig


def zeige_anova_tabelle(modell):
    """Zeigt die ANOVA-Tabelle des Regressionsmodells (Prio 2)."""
    import statsmodels.api as sm
    try:
        anova = sm.stats.anova_lm(modell, typ=2)
        anova_html = anova.to_html(float_format=lambda x: f"{x:.4f}", classes="table")
        zeige_prio2("ANOVA-Tabelle (Sum of Squares, df, Mean Squares, F, p)",
                    f"""
                    <p>Die ANOVA-Tabelle zerlegt die Gesamtvariation in Anteile pro Faktor:</p>
                    <ul>
                        <li><strong>SS</strong> (Sum of Squares): Anteil der erkl\u00e4rten Variation</li>
                        <li><strong>df</strong> (Degrees of Freedom): Freiheitsgrade</li>
                        <li><strong>MS</strong> = SS/df (Mean Squares)</li>
                        <li><strong>F</strong> = MS_Faktor / MS_Residual (Testvariable)</li>
                        <li><strong>p</strong>: Signifikanz (&lt; 0,05 = signifikant)</li>
                    </ul>
                    {anova_html}
                    """)
    except Exception as e:
        print(f"ANOVA-Tabelle konnte nicht erstellt werden: {e}")


# ═══════════════════════════════════════════════════════════════════
# 6. IMPROVE
# ═══════════════════════════════════════════════════════════════════

def plot_kontur(modell, faktoren: List[Dict], zielweite: float,
                faktor_idx: Tuple[int, int] = (0, 1),
                n_grid: int = 50) -> plt.Figure:
    """Konturplot / Heatmap für die zwei wichtigsten Faktoren."""
    f1_idx, f2_idx = faktor_idx
    fn = modell._faktor_namen

    # Grid erstellen
    x1 = np.linspace(-1.5, 1.5, n_grid)
    x2 = np.linspace(-1.5, 1.5, n_grid)
    X1, X2 = np.meshgrid(x1, x2)

    # Vorhersage-Grid (alle anderen Faktoren = 0)
    grid_df = pd.DataFrame({name: np.zeros(n_grid * n_grid) for name in fn})
    grid_df[fn[f1_idx]] = X1.ravel()
    grid_df[fn[f2_idx]] = X2.ravel()

    # Interaktionsterme hinzufügen
    for i in range(len(fn)):
        for j in range(i + 1, len(fn)):
            col_name = f"{fn[i]}:{fn[j]}"
            if col_name in modell.params.index:
                grid_df[col_name] = grid_df[fn[i]] * grid_df[fn[j]]

    grid_df = grid_df.reindex(columns=[c for c in modell.params.index if c != "Intercept"],
                               fill_value=0)
    import statsmodels.api as sm
    grid_df_const = sm.add_constant(grid_df)
    Z = modell.predict(grid_df_const).values.reshape(n_grid, n_grid)

    # Plot
    fig, ax = plt.subplots(figsize=(8, 6))
    cf = ax.contourf(X1, X2, Z, levels=20, cmap="RdYlGn_r", alpha=0.8)
    cs = ax.contour(X1, X2, Z, levels=[zielweite], colors=[GREEN], linewidths=2, linestyles="--")
    ax.clabel(cs, fmt=f"Ziel: {zielweite:.0f}", fontsize=10)
    fig.colorbar(cf, ax=ax, label="Vorhergesagte Wurfweite (cm)")

    # Achsenbeschriftungen mit Originalnamen
    f1 = faktoren[f1_idx]
    f2 = faktoren[f2_idx]
    ax.set_xlabel(f"{f1['name']} (kodiert: -1 = {f1['low']}, +1 = {f1['high']} {f1['einheit']})")
    ax.set_ylabel(f"{f2['name']} (kodiert: -1 = {f2['low']}, +1 = {f2['high']} {f2['einheit']})")
    # Versuchsraum einrahmen
    rect = plt.Rectangle((-1, -1), 2, 2, linewidth=2.5, edgecolor="black",
                          facecolor="none", linestyle="-", zorder=8,
                          label="Versuchsraum (Low/High)")
    ax.add_patch(rect)
    ax.legend(loc="lower left", fontsize=9)

    ax.set_title("Konturplot: Vorhergesagte Wurfweite", fontsize=13, fontweight="bold")
    fig.tight_layout()
    return fig


def _predict_grid(modell, faktor_idx: Tuple[int, int],
                  n_grid: int = 50) -> Tuple[np.ndarray, np.ndarray, np.ndarray]:
    """Erzeugt ein Vorhersage-Grid für zwei Faktoren. Gibt (X1, X2, Z) zurück."""
    import statsmodels.api as sm
    f1_idx, f2_idx = faktor_idx
    fn = modell._faktor_namen

    x1 = np.linspace(-1.5, 1.5, n_grid)
    x2 = np.linspace(-1.5, 1.5, n_grid)
    X1, X2 = np.meshgrid(x1, x2)

    grid_df = pd.DataFrame({name: np.zeros(n_grid * n_grid) for name in fn})
    grid_df[fn[f1_idx]] = X1.ravel()
    grid_df[fn[f2_idx]] = X2.ravel()

    for i in range(len(fn)):
        for j in range(i + 1, len(fn)):
            col_name = f"{fn[i]}:{fn[j]}"
            if col_name in modell.params.index:
                grid_df[col_name] = grid_df[fn[i]] * grid_df[fn[j]]

    grid_df = grid_df.reindex(columns=[c for c in modell.params.index if c != "Intercept"],
                               fill_value=0)
    grid_df_const = sm.add_constant(grid_df)
    Z = modell.predict(grid_df_const).values.reshape(n_grid, n_grid)
    return X1, X2, Z


def plot_kontur_varianz_dispersion(modell, doe_X: np.ndarray, doe_response: np.ndarray,
                                   faktoren: List[Dict],
                                   faktor_idx: Tuple[int, int] = (0, 1),
                                   n_grid: int = 50) -> plt.Figure:
    """Konturplot der geschätzten Streuung (Dispersionsmodell auf ln(s²))."""
    import statsmodels.api as sm
    f1_idx, f2_idx = faktor_idx
    fn = modell._faktor_namen

    # Varianz an jedem Designpunkt (aus Wiederholungen)
    unique_rows = np.unique(doe_X, axis=0)
    ln_s2_list, X_unique_list = [], []
    for row in unique_rows:
        mask = np.all(doe_X == row, axis=1)
        vals = doe_response[mask]
        s2 = np.var(vals, ddof=1) if len(vals) > 1 else np.nan
        if s2 > 0 and not np.isnan(s2):
            ln_s2_list.append(np.log(s2))
            X_unique_list.append(row)
    ln_s2 = np.array(ln_s2_list)
    X_unique = np.array(X_unique_list)

    # Dispersionsmodell: ln(s²) ~ Haupteffekte
    X_disp = pd.DataFrame({fn[i]: X_unique[:, i] for i in range(len(fn))})
    X_disp_const = sm.add_constant(X_disp)
    disp_model = sm.OLS(ln_s2, X_disp_const).fit()

    # Grid für Konturplot
    x1 = np.linspace(-1.5, 1.5, n_grid)
    x2 = np.linspace(-1.5, 1.5, n_grid)
    X1, X2 = np.meshgrid(x1, x2)

    grid_df = pd.DataFrame({fn[i]: np.zeros(n_grid * n_grid) for i in range(len(fn))})
    grid_df[fn[f1_idx]] = X1.ravel()
    grid_df[fn[f2_idx]] = X2.ravel()
    grid_const = sm.add_constant(grid_df)
    Z_ln_s2 = disp_model.predict(grid_const).values.reshape(n_grid, n_grid)
    Z_s = np.exp(Z_ln_s2 / 2)  # σ = sqrt(exp(ln(s²)))

    fig, ax = plt.subplots(figsize=(8, 6))
    cf = ax.contourf(X1, X2, Z_s, levels=20, cmap="YlOrRd", alpha=0.8)
    fig.colorbar(cf, ax=ax, label="Geschätzte Streuung σ (cm)")

    rect = plt.Rectangle((-1, -1), 2, 2, linewidth=2.5, edgecolor="black",
                          facecolor="none", linestyle="-", zorder=8,
                          label="Versuchsraum")
    ax.add_patch(rect)

    f1, f2 = faktoren[f1_idx], faktoren[f2_idx]
    ax.set_xlabel(f"{f1['name']} (kodiert: -1 = {f1['low']}, +1 = {f1['high']} {f1['einheit']})")
    ax.set_ylabel(f"{f2['name']} (kodiert: -1 = {f2['low']}, +1 = {f2['high']} {f2['einheit']})")
    ax.set_title("Dispersionsmodell: Geschätzte Streuung\n"
                 "(Modell auf ln(s²) der Wiederholungen)", fontsize=13, fontweight="bold")
    ax.legend(loc="lower left", fontsize=9)
    fig.tight_layout()
    return fig


def plot_kontur_varianz_transmitted(modell, faktoren: List[Dict],
                                     faktor_idx: Tuple[int, int] = (0, 1),
                                     sigma_setting: float = 0.1,
                                     n_grid: int = 50) -> plt.Figure:
    """Konturplot der übertragenen Varianz (Fehlerfortpflanzung)."""
    f1_idx, f2_idx = faktor_idx
    fn = modell._faktor_namen
    params = modell.params

    def get_beta(name):
        return params.get(name, 0.0)

    x1 = np.linspace(-1.5, 1.5, n_grid)
    x2 = np.linspace(-1.5, 1.5, n_grid)
    X1, X2 = np.meshgrid(x1, x2)

    # Partielle Ableitungen für alle Faktoren
    # Für jeden Faktor k: ∂ŷ/∂x_k = β_k + Σ_{j≠k} β_{k:j} · x_j
    # Auf dem Grid: Faktoren außerhalb (f1, f2) sind 0
    transmitted_var = np.zeros_like(X1)
    for k in range(len(fn)):
        dy_dxk = get_beta(fn[k])
        for j in range(len(fn)):
            if j == k:
                continue
            # Interaktionsterm suchen (beide Reihenfolgen)
            col1 = f"{fn[k]}:{fn[j]}"
            col2 = f"{fn[j]}:{fn[k]}"
            beta_int = get_beta(col1) if col1 in params.index else get_beta(col2)
            if beta_int != 0:
                if j == f1_idx:
                    dy_dxk = dy_dxk + beta_int * X1
                elif j == f2_idx:
                    dy_dxk = dy_dxk + beta_int * X2
                # else: Faktor j = 0 (Mitte), kein Beitrag
        transmitted_var += dy_dxk**2 * sigma_setting**2

    sigma_residual = np.sqrt(modell.mse_resid)
    total_sigma = np.sqrt(transmitted_var + modell.mse_resid)

    fig, ax = plt.subplots(figsize=(8, 6))
    cf = ax.contourf(X1, X2, total_sigma, levels=20, cmap="YlOrRd", alpha=0.8)
    fig.colorbar(cf, ax=ax, label="Erwartete Streuung σ_total (cm)")

    rect = plt.Rectangle((-1, -1), 2, 2, linewidth=2.5, edgecolor="black",
                          facecolor="none", linestyle="-", zorder=8,
                          label="Versuchsraum")
    ax.add_patch(rect)

    f1, f2 = faktoren[f1_idx], faktoren[f2_idx]
    ax.set_xlabel(f"{f1['name']} (kodiert: -1 = {f1['low']}, +1 = {f1['high']} {f1['einheit']})")
    ax.set_ylabel(f"{f2['name']} (kodiert: -1 = {f2['low']}, +1 = {f2['high']} {f2['einheit']})")
    ax.set_title("Fehlerfortpflanzung: Erwartete Streuung\n"
                 f"(σ_Einstellung = {sigma_setting}, σ_Residual = {sigma_residual:.1f} cm)",
                 fontsize=13, fontweight="bold")
    ax.legend(loc="lower left", fontsize=9)
    fig.tight_layout()
    return fig


def _transmitted_variance(x_coded, modell, sigma_setting=0.1):
    """Berechnet die übertragene Varianz an einem Punkt (für Optimierung)."""
    fn = modell._faktor_namen
    params = modell.params

    def get_beta(name):
        return params.get(name, 0.0)

    total = 0.0
    for k in range(len(fn)):
        # dy/dx_k = beta_k + 2*beta_{k_sq}*x_k + sum(beta_{k:j} * x_j)
        dy_dxk = get_beta(fn[k])
        # Quadratischer Term: dy/dx_k += 2 * beta_{k_sq} * x_k
        sq_name = f"{fn[k]}_sq"
        if sq_name in params.index:
            dy_dxk += 2.0 * get_beta(sq_name) * x_coded[k]
        # Interaktionsterme
        for j in range(len(fn)):
            if j == k:
                continue
            col1 = f"{fn[k]}:{fn[j]}"
            col2 = f"{fn[j]}:{fn[k]}"
            beta_int = get_beta(col1) if col1 in params.index else get_beta(col2)
            dy_dxk += beta_int * x_coded[j]
        total += dy_dxk**2 * sigma_setting**2
    return total + modell.mse_resid


def optimiere_einstellungen(modell, zielweite: float,
                            faktoren: List[Dict],
                            strategie: str = "mittelwert",
                            lambda_gewicht: float = 0.01,
                            sigma_setting: float = 0.1) -> Dict:
    """
    Optimiert die Faktoreinstellungen.

    Strategien:
        "mittelwert": Minimiere |ŷ - Ziel|  (nur Accuracy)
        "varianz":    Minimiere σ²(x)        (nur Precision/Robustheit)
        "dual":       Minimiere (ŷ - Ziel)² + λ · σ²(x)  (beides)
    """
    fn = modell._faktor_namen

    def _predict_at(x_coded):
        """Vorhersage an einem Punkt (mit 2FI, optionalen 3FI und quadratischen Termen)."""
        pred_dict = {name: val for name, val in zip(fn, x_coded)}
        # Quadratische Terme
        for i, name in enumerate(fn):
            sq_name = f"{name}_sq"
            if sq_name in modell.params.index:
                pred_dict[sq_name] = x_coded[i] ** 2
        # 2-Faktor-Interaktionen
        for i in range(len(fn)):
            for j in range(i + 1, len(fn)):
                col_name = f"{fn[i]}:{fn[j]}"
                if col_name in modell.params.index:
                    pred_dict[col_name] = x_coded[i] * x_coded[j]
        # 3-Faktor-Interaktionen
        for i in range(len(fn)):
            for j in range(i + 1, len(fn)):
                for k_ in range(j + 1, len(fn)):
                    col_name = f"{fn[i]}:{fn[j]}:{fn[k_]}"
                    if col_name in modell.params.index:
                        pred_dict[col_name] = x_coded[i] * x_coded[j] * x_coded[k_]
        pred_df = pd.DataFrame([pred_dict])
        pred_df = pred_df.reindex(columns=[c for c in modell.params.index if c != "Intercept"],
                                   fill_value=0)
        import statsmodels.api as sm
        pred_df = sm.add_constant(pred_df)
        return modell.predict(pred_df).values[0]

    def objective(x_coded):
        pred = _predict_at(x_coded)
        mean_loss = (pred - zielweite) ** 2
        var_loss = _transmitted_variance(x_coded, modell, sigma_setting)

        if strategie == "mittelwert":
            return mean_loss
        elif strategie == "varianz":
            return var_loss
        elif strategie == "dual":
            return mean_loss + lambda_gewicht * var_loss
        else:
            raise ValueError(f"Unbekannte Strategie: {strategie}")

    # Startpunkt: Mittelwert (kodiert = 0)
    x0 = np.zeros(len(fn))
    bounds = [(-1.2, 1.2)] * len(fn)

    result = optimize.minimize(objective, x0, method="L-BFGS-B", bounds=bounds)
    optimal_coded = result.x

    # Rück-Dekodierung in Originalwerte
    einstellungen = {}
    for i, f in enumerate(faktoren):
        center = (f["low"] + f["high"]) / 2
        half_range = (f["high"] - f["low"]) / 2
        original = center + optimal_coded[i] * half_range
        einstellungen[f["name"]] = {
            "coded": optimal_coded[i],
            "original": original,
            "einheit": f["einheit"],
        }

    # Vorhersage + Intervall
    pred_dict = {name: val for name, val in zip(fn, optimal_coded)}
    for i, name in enumerate(fn):
        sq_name = f"{name}_sq"
        if sq_name in modell.params.index:
            pred_dict[sq_name] = optimal_coded[i] ** 2
    for i in range(len(fn)):
        for j in range(i + 1, len(fn)):
            col_name = f"{fn[i]}:{fn[j]}"
            if col_name in modell.params.index:
                pred_dict[col_name] = optimal_coded[i] * optimal_coded[j]
    for i in range(len(fn)):
        for j in range(i + 1, len(fn)):
            for k_ in range(j + 1, len(fn)):
                col_name = f"{fn[i]}:{fn[j]}:{fn[k_]}"
                if col_name in modell.params.index:
                    pred_dict[col_name] = optimal_coded[i] * optimal_coded[j] * optimal_coded[k_]
    pred_df = pd.DataFrame([pred_dict])
    pred_df = pred_df.reindex(columns=[c for c in modell.params.index if c != "Intercept"],
                               fill_value=0)
    import statsmodels.api as sm
    pred_df = sm.add_constant(pred_df)

    prediction = modell.get_prediction(pred_df)
    pred_value = prediction.predicted_mean[0]
    pred_ci = prediction.conf_int(alpha=0.05)[0]
    pred_pi = prediction.summary_frame(alpha=0.05)

    # Erwartete Streuung am optimalen Punkt
    sigma_opt = np.sqrt(_transmitted_variance(optimal_coded, modell, sigma_setting))

    strategie_label = {
        "mittelwert": "Nur Mittelwert (Accuracy)",
        "varianz": "Nur Varianz (Precision)",
        "dual": f"Dual (λ = {lambda_gewicht})",
    }.get(strategie, strategie)

    return {
        "einstellungen": einstellungen,
        "vorhersage": pred_value,
        "ci_low": pred_ci[0],
        "ci_high": pred_ci[1],
        "pi_low": pred_pi["obs_ci_lower"].values[0],
        "pi_high": pred_pi["obs_ci_upper"].values[0],
        "zielweite": zielweite,
        "strategie": strategie,
        "strategie_label": strategie_label,
        "sigma_transmitted": sigma_opt,
    }


def zeige_optimierung(ergebnis: Dict):
    """Zeigt die Optimierungsergebnisse an."""
    rows = ""
    for name, data in ergebnis["einstellungen"].items():
        rows += f"""<tr>
            <td style="padding:6px; font-weight:bold;">{name}</td>
            <td style="padding:6px; text-align:right;">{data['original']:.1f} {data['einheit']}</td>
            <td style="padding:6px; text-align:right; color:{GRAY};">({data['coded']:+.2f})</td>
        </tr>"""

    # Strategie-Info
    strategie_html = ""
    if "strategie_label" in ergebnis:
        strategie_html = f"""<p style="margin:4px 0;">
            <strong>Strategie:</strong> {ergebnis['strategie_label']}</p>"""
    sigma_html = ""
    if "sigma_transmitted" in ergebnis:
        sigma_html = f"""<p style="margin:4px 0;">
            <strong>Erwartete Streuung (Fehlerfortpflanzung):</strong>
            σ ≈ {ergebnis['sigma_transmitted']:.1f} cm</p>"""

    html = f"""
    <div style="padding:12px; background:{LIGHT_GREEN}; border-radius:6px; margin:10px 0;
                border:2px solid {GREEN};">
        <h4 style="margin:0 0 8px 0;">🎯 Empfohlene Einstellungen</h4>
        <table style="border-collapse:collapse; width:80%;">
            <tr style="background:white;">
                <th style="padding:8px; text-align:left;">Faktor</th>
                <th style="padding:8px;">Einstellung</th>
                <th style="padding:8px;">Kodiert</th>
            </tr>
            {rows}
        </table>
        {strategie_html}
        <p style="margin:8px 0 0 0;">
            <strong>Vorhergesagte Wurfweite:</strong> {ergebnis['vorhersage']:.1f} cm<br>
            <strong>Vorhersageintervall (95%):</strong>
            [{ergebnis['pi_low']:.1f}, {ergebnis['pi_high']:.1f}] cm<br>
            <strong>Zielweite:</strong> {ergebnis['zielweite']:.0f} cm
        </p>
        {sigma_html}
    </div>"""
    display(HTML(html))


def zeige_regressionsformel(modell, faktoren: List[Dict]):
    """
    Zeigt die Regressionsformel vollständig gerendert als HTML/LaTeX an,
    mit echten Faktornamen und Koeffizienten.
    """
    fn = modell._faktor_namen
    params = modell.params
    intercept = params.get("Intercept", 0.0)

    # Mapping kodierter Name → lesbarer Name mit Einheit
    name_map = {}
    for i, f in enumerate(faktoren):
        if i < len(fn):
            name_map[fn[i]] = f["name"]

    # Formel in LaTeX aufbauen
    latex = f"\\hat{{y}} = {intercept:.2f}"

    # Haupteffekte
    for name in fn:
        beta = params.get(name, 0.0)
        if abs(beta) < 1e-10:
            continue
        anzeige = name_map.get(name, name)
        vorzeichen = "+" if beta > 0 else "-"
        latex += f" {vorzeichen} {abs(beta):.2f} \\cdot x_{{\\text{{{anzeige}}}}}"

    # Interaktionen
    for i in range(len(fn)):
        for j in range(i + 1, len(fn)):
            col = f"{fn[i]}:{fn[j]}"
            beta = params.get(col, 0.0)
            if abs(beta) < 1e-10:
                continue
            n1 = name_map.get(fn[i], fn[i])
            n2 = name_map.get(fn[j], fn[j])
            vorzeichen = "+" if beta > 0 else "-"
            latex += (f" {vorzeichen} {abs(beta):.2f} \\cdot "
                      f"x_{{\\text{{{n1}}}}} \\cdot x_{{\\text{{{n2}}}}}")

    # Kodierungstabelle
    kodierung_rows = ""
    for i, f in enumerate(faktoren):
        if i < len(fn):
            center = (f["low"] + f["high"]) / 2
            half_range = (f["high"] - f["low"]) / 2
            kodierung_rows += f"""<tr>
                <td style="padding:4px 8px;">x<sub>{f['name']}</sub></td>
                <td style="padding:4px 8px;">{f['name']}</td>
                <td style="padding:4px 8px; text-align:center;">
                    ({f['name']} − {center:.1f}) / {half_range:.1f}</td>
                <td style="padding:4px 8px; text-align:center;">
                    −1 → {f['low']:.1f} {f['einheit']}</td>
                <td style="padding:4px 8px; text-align:center;">
                    +1 → {f['high']:.1f} {f['einheit']}</td>
            </tr>"""

    html = f"""
    <div style="padding:16px; background:#F9FAFB; border:1px solid #E5E7EB;
                border-radius:8px; margin:10px 0;">
        <h4 style="margin:0 0 12px 0;">📐 Regressionsmodell</h4>
        <div style="padding:12px; background:white; border:1px solid #D1D5DB;
                    border-radius:6px; font-size:1.1em; overflow-x:auto;">
            $${latex}$$
        </div>
        <p style="margin:12px 0 4px 0; font-size:0.9em; color:{GRAY};">
            Alle x-Werte sind <strong>kodiert</strong> (−1 = niedrig, +1 = hoch).
            ŷ = vorhergesagte Wurfweite in cm.
            R² = {modell.rsquared:.3f}, R²(adj) = {modell.rsquared_adj:.3f},
            MSE = {modell.mse_resid:.2f}
        </p>
        <details style="margin-top:8px;">
            <summary style="cursor:pointer; font-weight:bold; font-size:0.9em;
                           color:#2563EB;">Kodierungstabelle anzeigen</summary>
            <table style="border-collapse:collapse; margin-top:6px; font-size:0.9em;
                          border:1px solid #E5E7EB;">
                <tr style="background:{LIGHT_BLUE};">
                    <th style="padding:4px 8px;">Variable</th>
                    <th style="padding:4px 8px;">Faktor</th>
                    <th style="padding:4px 8px;">Kodierung</th>
                    <th style="padding:4px 8px;">Niedrig (−1)</th>
                    <th style="padding:4px 8px;">Hoch (+1)</th>
                </tr>
                {kodierung_rows}
            </table>
        </details>
    </div>"""
    display(HTML(html))


def prognostiziere(modell, faktoren: List[Dict],
                   werte: Dict[str, float],
                   sigma_setting: float = 0.1) -> Dict:
    """
    Prognosetool: Vorhersage für beliebige Faktoreinstellungen.

    werte: Dict mit Faktornamen als Keys und kodierten Werten (-1 bis +1)
           ODER Originalwerten (werden automatisch kodiert wenn |Wert| > 1.5).
    """
    fn = modell._faktor_namen

    # Werte-Vektor aufbauen und ggf. auto-kodieren
    x_coded = np.zeros(len(fn))
    for i, f in enumerate(faktoren):
        if i >= len(fn):
            break
        val = werte.get(f["name"], 0.0)
        # Auto-Kodierung: wenn der Wert außerhalb [-1.5, 1.5] liegt,
        # ist es wahrscheinlich ein Originalwert
        center = (f["low"] + f["high"]) / 2
        half_range = (f["high"] - f["low"]) / 2
        if half_range > 0 and abs(val) > 1.5:
            x_coded[i] = (val - center) / half_range
        else:
            x_coded[i] = val

    # Vorhersage via statsmodels
    import statsmodels.api as sm
    pred_dict = {name: x_coded[idx] for idx, name in enumerate(fn)}
    for i in range(len(fn)):
        for j in range(i + 1, len(fn)):
            col = f"{fn[i]}:{fn[j]}"
            if col in modell.params.index:
                pred_dict[col] = x_coded[i] * x_coded[j]
    pred_df = pd.DataFrame([pred_dict])
    pred_df = pred_df.reindex(
        columns=[c for c in modell.params.index if c != "Intercept"],
        fill_value=0
    )
    pred_df = sm.add_constant(pred_df)

    prediction = modell.get_prediction(pred_df)
    pred_value = prediction.predicted_mean[0]
    summary = prediction.summary_frame(alpha=0.05)

    # Erwartete Streuung
    sigma = np.sqrt(_transmitted_variance(x_coded, modell, sigma_setting))

    return {
        "vorhersage": pred_value,
        "ci_low": summary["mean_ci_lower"].values[0],
        "ci_high": summary["mean_ci_upper"].values[0],
        "pi_low": summary["obs_ci_lower"].values[0],
        "pi_high": summary["obs_ci_upper"].values[0],
        "sigma": sigma,
        "x_coded": x_coded,
    }


def zeige_prognose(ergebnis: Dict, faktoren: List[Dict],
                   werte: Dict[str, float], zielweite: float = None):
    """Zeigt das Prognoseergebnis als formatierte HTML-Box an."""
    pred = ergebnis["vorhersage"]
    x_coded = ergebnis["x_coded"]

    # Einstellungstabelle
    rows = ""
    for i, f in enumerate(faktoren):
        if i >= len(x_coded):
            break
        center = (f["low"] + f["high"]) / 2
        half_range = (f["high"] - f["low"]) / 2
        original = center + x_coded[i] * half_range
        rows += f"""<tr>
            <td style="padding:4px 8px; font-weight:bold;">{f['name']}</td>
            <td style="padding:4px 8px; text-align:right;">
                {original:.1f} {f['einheit']}</td>
            <td style="padding:4px 8px; text-align:right; color:{GRAY};">
                ({x_coded[i]:+.2f})</td>
        </tr>"""

    # Zielweiten-Vergleich
    ziel_html = ""
    if zielweite is not None:
        abw = pred - zielweite
        farbe = GREEN if abs(abw) < 5 else (ORANGE if abs(abw) < 15 else RED)
        ziel_html = f"""
        <p style="margin:6px 0;">
            <strong>Zielweite:</strong> {zielweite:.0f} cm &nbsp;
            <strong>Abweichung:</strong>
            <span style="color:{farbe}; font-weight:bold;">{abw:+.1f} cm</span>
        </p>"""

    html = f"""
    <div style="padding:14px; background:#F0F9FF; border:1px solid #BFDBFE;
                border-radius:8px; margin:10px 0;">
        <h4 style="margin:0 0 8px 0;">🔮 Prognose</h4>
        <table style="border-collapse:collapse; margin-bottom:8px;">
            <tr style="background:{LIGHT_BLUE};">
                <th style="padding:4px 8px; text-align:left;">Faktor</th>
                <th style="padding:4px 8px;">Einstellung</th>
                <th style="padding:4px 8px;">Kodiert</th>
            </tr>
            {rows}
        </table>
        <div style="padding:10px; background:white; border:1px solid #D1D5DB;
                    border-radius:6px;">
            <p style="margin:4px 0; font-size:1.15em;">
                <strong>Vorhergesagte Wurfweite: {pred:.1f} cm</strong></p>
            <p style="margin:4px 0;">
                <strong>Konfidenzintervall (95%):</strong>
                [{ergebnis['ci_low']:.1f}, {ergebnis['ci_high']:.1f}] cm
                <span style="color:{GRAY}; font-size:0.85em;">
                    &mdash; wo liegt der wahre Mittelwert?</span></p>
            <p style="margin:4px 0;">
                <strong>Vorhersageintervall (95%):</strong>
                [{ergebnis['pi_low']:.1f}, {ergebnis['pi_high']:.1f}] cm
                <span style="color:{GRAY}; font-size:0.85em;">
                    &mdash; wo landet der nächste Wurf?</span></p>
            <p style="margin:4px 0;">
                <strong>Erwartete Streuung:</strong>
                σ ≈ {ergebnis['sigma']:.1f} cm</p>
            {ziel_html}
        </div>
    </div>"""
    display(HTML(html))


def vergleiche_optimierungen(modell, zielweite: float, faktoren: List[Dict],
                              lambda_werte: List[float] = None,
                              sigma_setting: float = 0.1):
    """Vergleicht die drei Optimierungsstrategien in einer Übersichtstabelle."""
    if lambda_werte is None:
        lambda_werte = [0.005, 0.01, 0.05]

    ergebnisse = []
    # Nur Mittelwert
    e_mean = optimiere_einstellungen(modell, zielweite, faktoren,
                                     strategie="mittelwert", sigma_setting=sigma_setting)
    ergebnisse.append(e_mean)
    # Nur Varianz
    e_var = optimiere_einstellungen(modell, zielweite, faktoren,
                                    strategie="varianz", sigma_setting=sigma_setting)
    ergebnisse.append(e_var)
    # Dual mit verschiedenen λ
    for lam in lambda_werte:
        e_dual = optimiere_einstellungen(modell, zielweite, faktoren,
                                          strategie="dual", lambda_gewicht=lam,
                                          sigma_setting=sigma_setting)
        ergebnisse.append(e_dual)

    # Tabelle
    rows = ""
    for e in ergebnisse:
        abw = abs(e["vorhersage"] - zielweite)
        color = GREEN if abw < 2 else (ORANGE if abw < 10 else RED)
        rows += f"""<tr>
            <td style="padding:6px;">{e['strategie_label']}</td>
            <td style="padding:6px; text-align:right; color:{color};">{e['vorhersage']:.1f}</td>
            <td style="padding:6px; text-align:right;">{abw:.1f}</td>
            <td style="padding:6px; text-align:right;">{e['sigma_transmitted']:.1f}</td>
        </tr>"""

    html = f"""
    <div style="padding:12px; background:#F9FAFB; border:1px solid #E5E7EB;
                border-radius:6px; margin:10px 0;">
        <h4 style="margin:0 0 8px 0;">📊 Vergleich der Optimierungsstrategien</h4>
        <p>Zielweite: <strong>{zielweite:.0f} cm</strong></p>
        <table style="border-collapse:collapse; width:100%; border:1px solid #E5E7EB;">
            <tr style="background:{LIGHT_BLUE};">
                <th style="padding:8px; text-align:left;">Strategie</th>
                <th style="padding:8px;">ŷ (cm)</th>
                <th style="padding:8px;">|ŷ − Ziel|</th>
                <th style="padding:8px;">σ erwartet (cm)</th>
            </tr>
            {rows}
        </table>
        <p style="margin:8px 0 0 0; font-size:0.9em; color:{GRAY};">
            <strong>Trade-off:</strong> "Nur Mittelwert" trifft das Ziel am besten,
            hat aber ggf. höhere Streuung. "Dual" findet einen Kompromiss.
            Je größer λ, desto stärker wird die Streuung gewichtet.
        </p>
    </div>"""
    display(HTML(html))
    return ergebnisse


def erstelle_konfirmation_excel(einstellungen: Dict, zielweite: float,
                                output_path: str = "Konfirmation.xlsx") -> str:
    """Erstellt die Konfirmations-Excel-Vorlage (nur Wurfweite)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Konfirmation"

    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Metadaten
    ws.cell(row=1, column=1, value="Konfirmationsversuch").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Zielweite: {zielweite:.0f} cm").font = Font(size=11)
    row = 3
    ws.cell(row=row, column=1, value="Empfohlene Einstellungen:").font = Font(bold=True)
    row += 1
    for name, data in einstellungen.items():
        ws.cell(row=row, column=1, value=f"  {name}:").font = Font(size=11)
        ws.cell(row=row, column=2, value=f"{data['original']:.1f} {data['einheit']}").font = Font(size=11)
        row += 1
    row += 1

    # Datentabelle (nur Wurfweite)
    headers = ["Wurf-ID", "Weite (cm)"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for i in range(1, 21):  # 20 Zeilen vorbereiten
        data_row = row + i
        ws.cell(row=data_row, column=1, value=i).border = thin_border
        ws.cell(row=data_row, column=2).border = thin_border

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 16

    wb.save(output_path)
    return os.path.abspath(output_path)


def lade_konfirmation_aus_excel(filepath: str) -> np.ndarray:
    """Liest Wurfweiten aus einer ausgefüllten Konfirmations-Excel.

    Erwartet das vom ``erstelle_konfirmation_excel`` erzeugte Layout:
    Nach den Metadatenzeilen folgt eine Header-Zeile mit mindestens
    "Wurf-ID" und "Weite". Datenzeilen beginnen direkt darunter.
    Leere Zellen werden ignoriert.
    """
    import openpyxl

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["Konfirmation"] if "Konfirmation" in wb.sheetnames else wb.active

    # Header-Zeile finden (Zelle mit "Wurf-ID" in Spalte A)
    header_row = None
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if isinstance(val, str) and "wurf" in val.lower() and "id" in val.lower():
            header_row = r
            break
    if header_row is None:
        raise ValueError("Konnte Header-Zeile (Wurf-ID) in der Excel nicht finden.")

    weiten: list = []
    for r in range(header_row + 1, ws.max_row + 1):
        wid = ws.cell(row=r, column=1).value
        weite = ws.cell(row=r, column=2).value
        if wid is None and weite is None:
            continue
        if isinstance(weite, (int, float)) and weite > 0:
            weiten.append(float(weite))
    return np.array(weiten)


def analysiere_konfirmation(wuerfe: np.ndarray, vorhersage: float,
                            pi_low: float, pi_high: float,
                            zielweite: float, toleranz: float) -> Dict:
    """Analysiert die Konfirmationswürfe."""
    mu = np.mean(wuerfe)
    sigma = np.std(wuerfe, ddof=1) if len(wuerfe) > 1 else 0

    # Mittelwert im Vorhersagebereich?
    in_pred = pi_low <= mu <= pi_high

    # Anteil innerhalb Toleranz
    in_tol = np.sum(np.abs(wuerfe - zielweite) <= toleranz)
    pct_in_tol = in_tol / len(wuerfe) * 100

    # Bewertung
    if in_pred and pct_in_tol >= 80:
        bewertung = "✅ Konfirmation erfolgreich"
        detail = "Mittelwert im Vorhersagebereich und ≥80% der Würfe in Toleranz"
    elif in_pred:
        bewertung = "⚠️ Mittelwert passt, aber hohe Streuung"
        detail = f"Mittelwert im Vorhersagebereich, aber nur {pct_in_tol:.0f}% in Toleranz"
    else:
        bewertung = "❌ Modell hat nicht getroffen"
        detail = f"Mittelwert ({mu:.1f} cm) liegt außerhalb des Vorhersagebereichs [{pi_low:.1f}, {pi_high:.1f}]"

    return {
        "mean": mu,
        "std": sigma,
        "n": len(wuerfe),
        "in_pred": in_pred,
        "pct_in_tol": pct_in_tol,
        "bewertung": bewertung,
        "detail": detail,
        "vorhersage": vorhersage,
    }


def zeige_konfirmation(ergebnis: Dict):
    """Zeigt die Konfirmationsergebnisse an."""
    farbe = GREEN if "✅" in ergebnis["bewertung"] else ORANGE if "⚠️" in ergebnis["bewertung"] else RED
    html = f"""
    <div style="padding:12px; border-left:4px solid {farbe}; background:{farbe}11;
                border-radius:4px; margin:10px 0;">
        <h4 style="margin:0;">{ergebnis['bewertung']}</h4>
        <p>{ergebnis['detail']}</p>
        <p>
            Modell-Vorhersage: <strong>{ergebnis['vorhersage']:.1f} cm</strong><br>
            Tatsächlicher Mittelwert: <strong>{ergebnis['mean']:.1f} cm</strong> (σ = {ergebnis['std']:.1f} cm)<br>
            Anteil in Toleranz: <strong>{ergebnis['pct_in_tol']:.0f}%</strong>
        </p>
    </div>"""
    display(HTML(html))


# ═══════════════════════════════════════════════════════════════════
# 7. CONTROL
# ═══════════════════════════════════════════════════════════════════

def berechne_imr(daten: np.ndarray) -> Dict:
    """
    Berechnet I-MR-Kontrollkarte (Individuals + Moving Range).

    UCL/LCL basierend auf Moving Range:
      MR_bar = mittlere gleitende Spannweite
      UCL_I = x_bar + 2.66 * MR_bar
      LCL_I = x_bar - 2.66 * MR_bar
      UCL_MR = 3.267 * MR_bar
    """
    x_bar = np.mean(daten)
    mr = np.abs(np.diff(daten))
    mr_bar = np.mean(mr)

    # d2 = 1.128 für n=2 (Moving Range Subgruppengröße)
    ucl_i = x_bar + 2.66 * mr_bar
    lcl_i = x_bar - 2.66 * mr_bar
    ucl_mr = 3.267 * mr_bar

    # Stabilitätsprüfung
    ausserhalb_i = np.sum((daten > ucl_i) | (daten < lcl_i))
    ausserhalb_mr = np.sum(mr > ucl_mr)
    stabil = ausserhalb_i == 0 and ausserhalb_mr == 0

    return {
        "x_bar": x_bar,
        "mr": mr,
        "mr_bar": mr_bar,
        "ucl_i": ucl_i,
        "lcl_i": lcl_i,
        "ucl_mr": ucl_mr,
        "ausserhalb_i": ausserhalb_i,
        "ausserhalb_mr": ausserhalb_mr,
        "stabil": stabil,
    }


def plot_imr(daten: np.ndarray, titel="I-MR-Kontrollkarte") -> plt.Figure:
    """Plottet I-Chart und MR-Chart."""
    ergebnis = berechne_imr(daten)

    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(10, 7), sharex=True)

    n = len(daten)
    x_range = range(1, n + 1)

    # I-Chart
    colors_i = [RED if v > ergebnis["ucl_i"] or v < ergebnis["lcl_i"] else BLUE
                for v in daten]
    ax1.plot(x_range, daten, "o-", color=BLUE, markersize=6, linewidth=1.5, zorder=3)
    for i, (v, c) in enumerate(zip(daten, colors_i)):
        if c == RED:
            ax1.plot(i + 1, v, "o", color=RED, markersize=10, zorder=5)

    ax1.axhline(ergebnis["x_bar"], color=GREEN, linewidth=2, label=f"x̄ = {ergebnis['x_bar']:.1f}")
    ax1.axhline(ergebnis["ucl_i"], color=RED, linewidth=1.5, linestyle="--",
                label=f"UCL = {ergebnis['ucl_i']:.1f}")
    ax1.axhline(ergebnis["lcl_i"], color=RED, linewidth=1.5, linestyle="--",
                label=f"LCL = {ergebnis['lcl_i']:.1f}")
    ax1.fill_between(range(0, n + 2), ergebnis["lcl_i"], ergebnis["ucl_i"],
                     alpha=0.06, color=GREEN)
    ax1.set_ylabel("Wurfweite (cm)")
    ax1.set_title("I-Chart (Einzelwerte)", fontsize=12, fontweight="bold")
    ax1.legend(fontsize=9, loc="upper right")

    # MR-Chart
    mr = ergebnis["mr"]
    mr_range = range(2, n + 1)
    colors_mr = [RED if v > ergebnis["ucl_mr"] else BLUE for v in mr]
    ax2.plot(mr_range, mr, "o-", color=BLUE, markersize=6, linewidth=1.5, zorder=3)
    for i, (v, c) in enumerate(zip(mr, colors_mr)):
        if c == RED:
            ax2.plot(i + 2, v, "o", color=RED, markersize=10, zorder=5)

    ax2.axhline(ergebnis["mr_bar"], color=GREEN, linewidth=2,
                label=f"MR̄ = {ergebnis['mr_bar']:.1f}")
    ax2.axhline(ergebnis["ucl_mr"], color=RED, linewidth=1.5, linestyle="--",
                label=f"UCL = {ergebnis['ucl_mr']:.1f}")
    ax2.set_xlabel("Wurf-Nr.")
    ax2.set_ylabel("Moving Range")
    ax2.set_title("MR-Chart (Gleitende Spannweiten)", fontsize=12, fontweight="bold")
    ax2.legend(fontsize=9, loc="upper right")

    for ax in [ax1, ax2]:
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)

    fig.suptitle(titel, fontsize=14, fontweight="bold")
    fig.tight_layout(rect=[0, 0, 1, 0.95])
    return fig


def zeige_stabilitaet(ergebnis: Dict):
    """Zeigt die Stabilitätsbewertung an."""
    if ergebnis["stabil"]:
        display(HTML(f"""
        <div style="padding:10px; border-left:4px solid {GREEN}; background:{LIGHT_GREEN};
                     border-radius:4px; margin:8px 0;">
            ✅ <strong>Prozess ist statistisch stabil.</strong>
            Kein Punkt außerhalb der Kontrollgrenzen. Cpk-Berechnung ist aussagekräftig.
        </div>"""))
    else:
        display(HTML(f"""
        <div style="padding:10px; border-left:4px solid {ORANGE}; background:{LIGHT_YELLOW};
                     border-radius:4px; margin:8px 0;">
            ⚠️ <strong>Anzeichen von Instabilität.</strong>
            {ergebnis['ausserhalb_i']} Punkt(e) außerhalb der I-Kontrollgrenzen.
            Cpk wird berechnet, aber die Aussagekraft ist eingeschränkt.
            Diskutiert: Was könnte die Ursache sein?
        </div>"""))


def pruefe_normalverteilung(daten: np.ndarray) -> Dict:
    """Shapiro-Wilk-Test + Daten für Q-Q-Plot."""
    stat, p = stats.shapiro(daten) if len(daten) >= 8 else (np.nan, np.nan)
    return {"shapiro_stat": stat, "shapiro_p": p}


def plot_qq(daten: np.ndarray) -> plt.Figure:
    """Q-Q-Plot der Daten."""
    fig, ax = plt.subplots(figsize=(6, 6))
    stats.probplot(daten, plot=ax)
    ax.get_lines()[0].set(color=BLUE, markersize=5)
    ax.get_lines()[1].set(color=RED)
    ax.set_title("Q-Q-Plot (Normalverteilungsprüfung)", fontsize=13, fontweight="bold")
    fig.tight_layout()
    return fig


def berechne_cpk(daten: np.ndarray, usl: float, lsl: float) -> Dict:
    """Berechnet Cp und Cpk."""
    mu = np.mean(daten)
    sigma = np.std(daten, ddof=1)

    if sigma == 0:
        return {"cp": float("inf"), "cpk": float("inf"), "mu": mu, "sigma": 0,
                "bewertung_industrie": "–", "bewertung_katapult": "–"}

    cp = (usl - lsl) / (6 * sigma)
    cpk = min((usl - mu) / (3 * sigma), (mu - lsl) / (3 * sigma))

    # Bewertung Industrie
    if cpk < 0.67:
        bew_ind = "❌ Nicht fähig"
    elif cpk < 1.0:
        bew_ind = "❌ Nicht fähig"
    elif cpk < 1.33:
        bew_ind = "⚠️ Bedingt fähig"
    else:
        bew_ind = "✅ Prozess fähig"

    # Bewertung Katapult
    if cpk < 0.67:
        bew_kat = "❌ Deutlich verbesserungsbedürftig"
    elif cpk < 1.0:
        bew_kat = "⚠️ Deutliche Verbesserung ggü. Baseline"
    elif cpk < 1.33:
        bew_kat = "✅ Gute Prozessfähigkeit für ein Experiment"
    else:
        bew_kat = "✅ Hervorragend"

    return {
        "cp": cp,
        "cpk": cpk,
        "mu": mu,
        "sigma": sigma,
        "usl": usl,
        "lsl": lsl,
        "bewertung_industrie": bew_ind,
        "bewertung_katapult": bew_kat,
    }


def zeige_cpk(ergebnis: Dict):
    """Zeigt die Cpk-Ergebnisse mit Dual-Skala an."""
    html = f"""
    <div style="padding:12px; background:#F9FAFB; border:1px solid #E5E7EB;
                border-radius:6px; margin:10px 0;">
        <h4 style="margin:0 0 8px 0;">Prozessfähigkeit</h4>
        <table style="border-collapse:collapse; width:80%; border:1px solid #E5E7EB;">
            <tr style="background:{LIGHT_BLUE};">
                <th style="padding:8px;">Kennzahl</th>
                <th style="padding:8px;">Wert</th>
            </tr>
            <tr>
                <td style="padding:6px;">Cp</td>
                <td style="padding:6px; text-align:center; font-weight:bold;">{ergebnis['cp']:.2f}</td>
            </tr>
            <tr>
                <td style="padding:6px;">Cpk</td>
                <td style="padding:6px; text-align:center; font-weight:bold; font-size:1.2em;">{ergebnis['cpk']:.2f}</td>
            </tr>
            <tr>
                <td style="padding:6px;">μ</td>
                <td style="padding:6px; text-align:center;">{ergebnis['mu']:.1f} cm</td>
            </tr>
            <tr>
                <td style="padding:6px;">σ</td>
                <td style="padding:6px; text-align:center;">{ergebnis['sigma']:.1f} cm</td>
            </tr>
        </table>
        <h4 style="margin:12px 0 6px 0;">Bewertung (Dual-Skala)</h4>
        <table style="border-collapse:collapse; width:80%; border:1px solid #E5E7EB;">
            <tr style="background:{LIGHT_BLUE};">
                <th style="padding:8px;">Perspektive</th>
                <th style="padding:8px;">Bewertung</th>
            </tr>
            <tr>
                <td style="padding:6px;">Industriestandard</td>
                <td style="padding:6px;">{ergebnis['bewertung_industrie']}</td>
            </tr>
            <tr>
                <td style="padding:6px;">Studentisches Katapult</td>
                <td style="padding:6px;">{ergebnis['bewertung_katapult']}</td>
            </tr>
        </table>
    </div>"""
    display(HTML(html))


def plot_cpk_verteilung(ergebnis: Dict) -> plt.Figure:
    """Normalverteilungskurve mit USL/LSL und μ ± 3σ Bändern."""
    mu, sigma = ergebnis["mu"], ergebnis["sigma"]
    usl, lsl = ergebnis["usl"], ergebnis["lsl"]

    fig, ax = plt.subplots(figsize=(9, 5))
    x = np.linspace(mu - 4.5 * sigma, mu + 4.5 * sigma, 400) if sigma > 0 \
        else np.linspace(lsl - 10, usl + 10, 400)
    y = stats.norm.pdf(x, mu, sigma) if sigma > 0 else np.zeros_like(x)

    ax.plot(x, y, color=BLUE, linewidth=2.5)
    ax.fill_between(x, y, alpha=0.15, color=BLUE)

    ax.axvline(usl, color=RED, linewidth=2, linestyle="--", label=f"USL = {usl:.0f}")
    ax.axvline(lsl, color=RED, linewidth=2, linestyle="--", label=f"LSL = {lsl:.0f}")
    ax.axvline(mu, color=GRAY, linewidth=1.5, linestyle=":", label=f"μ = {mu:.1f}")

    if sigma > 0:
        for s in [-3, 3]:
            val = mu + s * sigma
            ax.axvline(val, color=ORANGE, linewidth=1, linestyle=":")
            ax.text(val, max(y) * 0.05, f"μ{'+'if s>0 else ''}{s}σ",
                    ha="center", fontsize=8, color=ORANGE)

    ax.text(0.02, 0.95, f"Cp = {ergebnis['cp']:.2f}  |  Cpk = {ergebnis['cpk']:.2f}",
            transform=ax.transAxes, fontsize=12, fontweight="bold", color=BLUE, va="top",
            bbox=dict(boxstyle="round,pad=0.4", facecolor=LIGHT_BLUE, edgecolor=BLUE, alpha=0.8))

    ax.set_title("Prozessfähigkeit: Verteilung vs. Spezifikation", fontsize=13, fontweight="bold")
    ax.set_xlabel("Wurfweite (cm)")
    ax.set_yticks([])
    ax.legend(fontsize=9, loc="upper right")
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    fig.tight_layout()
    return fig


def plot_vorher_nachher(baseline: np.ndarray, konfirmation: np.ndarray,
                        zielweite: float, toleranz: float,
                        modus: str = "1D") -> plt.Figure:
    """Vorher/Nachher-Zielscheibe (Baseline vs. Konfirmation)."""
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 5))

    plot_zielscheibe(baseline, zielweite, toleranz, modus=modus,
                     titel="Vorher (Baseline)", farbe=RED, ax=ax1)
    plot_zielscheibe(konfirmation, zielweite, toleranz, modus=modus,
                     titel="Nachher (Konfirmation)", farbe=GREEN, ax=ax2)

    fig.suptitle("Vorher / Nachher – Zielscheiben-Vergleich",
                 fontsize=14, fontweight="bold")
    fig.tight_layout(rect=[0, 0, 1, 0.93])
    return fig


# ═══════════════════════════════════════════════════════════════════
# 8. EXPORT
# ═══════════════════════════════════════════════════════════════════

def exportiere_phase_auf_drive(projekt: Projekt, phase: str = None):
    """Exportiert alle bisherigen Ergebnisse in einen Phase-Unterordner auf Drive."""
    if phase is None:
        phase = _aktuelle_phase(projekt)

    save_dir = _fortschritt_verzeichnis(projekt)
    if not save_dir:
        print("⚠️ Kein Speicherort verfügbar.")
        return

    phase_dir = os.path.join(save_dir, phase)
    os.makedirs(phase_dir, exist_ok=True)

    dateien = []

    # 1. Alle Figuren als PNG
    for name, fig in projekt.figuren.items():
        pfad = os.path.join(phase_dir, f"{name}.png")
        fig.savefig(pfad, format="png", bbox_inches="tight", dpi=150, facecolor="white")
        dateien.append(f"plots/{name}.png")

    # 2. Alle CSVs
    for name, df in projekt.csv_daten.items():
        pfad = os.path.join(phase_dir, f"{name}.csv")
        df.to_csv(pfad, index=False)
        dateien.append(f"daten/{name}.csv")

    # 3. Baseline-Daten
    if len(projekt.baseline_wuerfe) > 0:
        pfad = os.path.join(phase_dir, "baseline_wuerfe.csv")
        pd.DataFrame({"Wurf_Nr": range(1, len(projekt.baseline_wuerfe) + 1),
                       "Weite_cm": projekt.baseline_wuerfe}).to_csv(pfad, index=False)
        dateien.append("baseline_wuerfe.csv")

    # 4. Konfirmationsdaten
    if len(projekt.konfirmation_wuerfe) > 0:
        pfad = os.path.join(phase_dir, "konfirmation_wuerfe.csv")
        pd.DataFrame({"Wurf_Nr": range(1, len(projekt.konfirmation_wuerfe) + 1),
                       "Weite_cm": projekt.konfirmation_wuerfe}).to_csv(pfad, index=False)
        dateien.append("konfirmation_wuerfe.csv")

    # 5. Zusammenfassung
    pfad = os.path.join(phase_dir, "zusammenfassung.txt")
    with open(pfad, "w") as f:
        f.write(_erstelle_zusammenfassung(projekt))
    dateien.append("zusammenfassung.txt")

    # 6. Fortschritt-JSON
    speichere_fortschritt(projekt)

    print(f"💾 {len(dateien)} Dateien in Google Drive gespeichert:")
    print(f"   📂 {phase_dir}")
    for d in dateien[:8]:
        print(f"      • {d}")
    if len(dateien) > 8:
        print(f"      ... und {len(dateien) - 8} weitere")


def exportiere_zip(projekt: Projekt, output_path: str = "DMAIC_Ergebnisse.zip") -> str:
    """Exportiert alle Figuren und Daten als ZIP-Datei."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        # PNGs
        for name, fig in projekt.figuren.items():
            img_buf = io.BytesIO()
            fig.savefig(img_buf, format="png", bbox_inches="tight", dpi=150, facecolor="white")
            img_buf.seek(0)
            zf.writestr(f"plots/{name}.png", img_buf.read())

        # CSVs
        for name, df in projekt.csv_daten.items():
            csv_buf = df.to_csv(index=False)
            zf.writestr(f"daten/{name}.csv", csv_buf)

        # Zusammenfassung
        zusammenfassung = _erstelle_zusammenfassung(projekt)
        zf.writestr("zusammenfassung.txt", zusammenfassung)

    with open(output_path, "wb") as f:
        f.write(buf.getvalue())

    return os.path.abspath(output_path)


def _erstelle_zusammenfassung(projekt: Projekt) -> str:
    """Erstellt eine Text-Zusammenfassung aller Ergebnisse."""
    lines = [
        f"DMAIC Katapult-Versuch – Zusammenfassung",
        f"=========================================",
        f"Gruppe: {projekt.gruppenname} (Nr. {projekt.gruppennummer})",
        f"Datum: {date.today().strftime('%d.%m.%Y')}",
        f"",
        f"DEFINE",
        f"  Zielweite: {projekt.zielweite:.0f} cm",
        f"  Toleranz: ±{projekt.toleranz:.0f} cm",
        f"  Messmodus: {projekt.messmodus}",
    ]

    if len(projekt.testwuerfe) > 0:
        s = berechne_testwurf_statistik(projekt.testwuerfe)
        lines.extend([
            f"  Testwürfe: μ={s['mean']:.1f}, σ={s['std']:.1f}, CV={s['cv']:.1f}%",
        ])

    if projekt.msa_grr:
        lines.extend([
            f"",
            f"MEASURE",
            f"  %GRR: {projekt.msa_grr['pct_grr']:.1f}%",
            f"  Bewertung: {projekt.msa_grr['bewertung']}",
        ])

    if projekt.baseline_stats:
        b = projekt.baseline_stats
        lines.extend([
            f"  Baseline: μ={b['mean']:.1f}, σ={b['std']:.1f}, n={b['n']}",
            f"  Shapiro-Wilk p={b['shapiro_p']:.3f}",
        ])

    if projekt.modell is not None:
        m = projekt.modell
        lines.extend([
            f"",
            f"ANALYZE",
            f"  R² = {m.rsquared:.4f}",
            f"  Adj. R² = {m.rsquared_adj:.4f}",
            f"  RMSE = {np.sqrt(m.mse_resid):.2f}",
        ])

    if projekt.konfirmation_ergebnis:
        k = projekt.konfirmation_ergebnis
        lines.extend([
            f"",
            f"IMPROVE",
            f"  {k['bewertung']}",
            f"  Vorhersage: {k['vorhersage']:.1f} cm",
            f"  Ist: μ={k['mean']:.1f}, σ={k['std']:.1f}",
        ])

    if projekt.cpk_ergebnis:
        c = projekt.cpk_ergebnis
        lines.extend([
            f"",
            f"CONTROL",
            f"  Cpk = {c['cpk']:.2f}",
            f"  Industrie: {c['bewertung_industrie']}",
            f"  Katapult: {c['bewertung_katapult']}",
        ])

    return "\n".join(lines)


def download_colab(filepath: str):
    """Löst einen Download in Google Colab aus."""
    try:
        from google.colab import files as colab_files
        colab_files.download(filepath)
    except ImportError:
        print(f"Datei gespeichert: {filepath}")
        print("(Automatischer Download nur in Google Colab verfügbar)")


# ═══════════════════════════════════════════════════════════════════
# 9. PERSISTENZ (Auto-Save / Restore für Colab)
# ═══════════════════════════════════════════════════════════════════

_DRIVE_BASE = "/content/drive/MyDrive/DMAIC_Katapult"
_LOCAL_BASE = "./DMAIC_Daten"


def _fortschritt_verzeichnis(projekt: Projekt) -> str:
    """Gibt das Speicherverzeichnis für ein Projekt zurück (Drive oder lokal)."""
    ordner = f"{projekt.gruppenname}_{projekt.gruppennummer}"
    if os.path.isdir("/content/drive/MyDrive"):
        return os.path.join(_DRIVE_BASE, ordner)
    return os.path.join(_LOCAL_BASE, ordner)


def _sanitize(obj):
    """Konvertiert numpy-Typen rekursiv in JSON-kompatible Python-Typen."""
    if isinstance(obj, dict):
        return {k: _sanitize(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [_sanitize(v) for v in obj]
    if isinstance(obj, np.ndarray):
        return _sanitize(obj.tolist())
    if isinstance(obj, (np.integer,)):
        return int(obj)
    if isinstance(obj, (np.floating,)):
        v = float(obj)
        if np.isnan(v) or np.isinf(v):
            return None
        return v
    if isinstance(obj, (np.bool_,)):
        return bool(obj)
    if isinstance(obj, float) and (np.isnan(obj) or np.isinf(obj)):
        return None
    return obj


def _dataframe_to_dict(df: pd.DataFrame, with_index: bool = False) -> dict:
    """Serialisiert einen DataFrame als JSON-kompatibles Dict."""
    d = {
        "columns": df.columns.tolist(),
        "data": _sanitize(df.values.tolist()),
    }
    if with_index:
        d["index"] = [str(i) for i in df.index.tolist()]
    return d


def _dict_to_dataframe(d: dict) -> pd.DataFrame:
    """Rekonstruiert einen DataFrame aus einem serialisierten Dict."""
    df = pd.DataFrame(data=d["data"], columns=d["columns"])
    if "index" in d:
        df.index = d["index"]
    return df


def _projekt_to_dict(projekt: Projekt) -> dict:
    """Konvertiert ein Projekt-Objekt in ein JSON-serialisierbares Dict."""
    d = {}

    # Skalare
    for key in ("gruppenname", "gruppennummer", "seed", "zielweite", "toleranz", "messmodus"):
        d[key] = getattr(projekt, key)

    # Numpy-Arrays
    for key in ("testwuerfe", "baseline_wuerfe", "konfirmation_wuerfe",
                "vermessung_min_wuerfe", "vermessung_max_wuerfe"):
        arr = getattr(projekt, key)
        d[key] = arr.tolist() if len(arr) > 0 else []

    # Einfache Dicts / Listen
    d["charter"] = projekt.charter
    d["faktoren"] = projekt.faktoren
    d["faktoren_doe"] = projekt.faktoren_doe
    d["pruning_log"] = projekt.pruning_log

    # Vermessung (DEFINE)
    d["vermessung_min_einstellung"] = projekt.vermessung_min_einstellung
    d["vermessung_max_einstellung"] = projekt.vermessung_max_einstellung
    d["vermessung_beschreibung"] = projekt.vermessung_beschreibung
    # Annäherung & initiale Einstellung
    d["annaeherung_log"] = _sanitize(projekt.annaeherung_log)
    d["initiale_einstellung"] = projekt.initiale_einstellung

    # MSA
    if projekt.msa_type1 is not None:
        d["msa_type1"] = _sanitize(projekt.msa_type1)

    if projekt.msa_grr is not None:
        grr = {}
        for k, v in projekt.msa_grr.items():
            if k == "anova_table" and isinstance(v, pd.DataFrame):
                grr[k] = _dataframe_to_dict(v, with_index=True)
            else:
                grr[k] = _sanitize(v)
        d["msa_grr"] = grr

    # MSA-Rohdaten (Long-Format DataFrame)
    if projekt.msa_rohdaten is not None:
        d["msa_rohdaten"] = _dataframe_to_dict(projekt.msa_rohdaten)

    # DoE-Ergebnisse
    if projekt.doe_ergebnisse is not None:
        d["doe_ergebnisse"] = _dataframe_to_dict(projekt.doe_ergebnisse)

    # Optimale Einstellung
    if projekt.optimale_einstellung is not None:
        d["optimale_einstellung"] = _sanitize(projekt.optimale_einstellung)

    # Config-Felder
    if projekt.versuchsplan_config is not None:
        d["versuchsplan_config"] = projekt.versuchsplan_config
    if projekt.optimierung_config is not None:
        d["optimierung_config"] = projekt.optimierung_config

    # Metadaten
    d["_version"] = 1
    d["_saved_at"] = datetime.now().isoformat(timespec="seconds")
    d["_phase"] = _aktuelle_phase(projekt)

    return _sanitize(d)


def _aktuelle_phase(projekt: Projekt) -> str:
    """Ermittelt die zuletzt abgeschlossene DMAIC-Phase."""
    if len(projekt.konfirmation_wuerfe) > 0:
        return "CONTROL"
    if projekt.optimale_einstellung is not None:
        return "IMPROVE"
    if projekt.doe_ergebnisse is not None:
        return "ANALYZE"
    if len(projekt.baseline_wuerfe) > 0 or projekt.msa_grr is not None:
        return "MEASURE"
    if len(projekt.testwuerfe) > 0 or projekt.charter:
        return "DEFINE"
    return "START"


def _dict_to_projekt(d: dict) -> Projekt:
    """Rekonstruiert ein Projekt-Objekt aus einem gespeicherten Dict."""
    p = Projekt(
        gruppenname=d.get("gruppenname", ""),
        gruppennummer=d.get("gruppennummer", 1),
        seed=d.get("seed", 42),
        zielweite=d.get("zielweite", 0.0),
        toleranz=d.get("toleranz", 15.0),
        messmodus=d.get("messmodus", "1D"),
    )

    # Arrays
    for key in ("testwuerfe", "baseline_wuerfe", "konfirmation_wuerfe",
                "vermessung_min_wuerfe", "vermessung_max_wuerfe"):
        val = d.get(key, [])
        setattr(p, key, np.array(val) if val else np.array([]))

    # Einfache Felder
    p.charter = d.get("charter", {})
    p.faktoren = d.get("faktoren", [])
    p.faktoren_doe = d.get("faktoren_doe", [])
    p.pruning_log = d.get("pruning_log", [])

    # Vermessung (DEFINE)
    p.vermessung_min_einstellung = d.get("vermessung_min_einstellung", {})
    p.vermessung_max_einstellung = d.get("vermessung_max_einstellung", {})
    p.vermessung_beschreibung = d.get("vermessung_beschreibung", "")
    # Annäherung & initiale Einstellung (Legacy-Key "typische_einstellung" akzeptieren)
    p.annaeherung_log = d.get("annaeherung_log", [])
    p.initiale_einstellung = d.get("initiale_einstellung",
                                   d.get("typische_einstellung", {}))

    # MSA
    p.msa_type1 = d.get("msa_type1")

    if "msa_grr" in d and d["msa_grr"] is not None:
        grr = dict(d["msa_grr"])
        if "anova_table" in grr and isinstance(grr["anova_table"], dict):
            grr["anova_table"] = _dict_to_dataframe(grr["anova_table"])
        p.msa_grr = grr

    if "msa_rohdaten" in d and d["msa_rohdaten"] is not None:
        p.msa_rohdaten = _dict_to_dataframe(d["msa_rohdaten"])

    # DoE
    if "doe_ergebnisse" in d and d["doe_ergebnisse"] is not None:
        p.doe_ergebnisse = _dict_to_dataframe(d["doe_ergebnisse"])
        p.csv_daten["doe_ergebnisse"] = p.doe_ergebnisse

    # Optimierung
    p.optimale_einstellung = d.get("optimale_einstellung")

    # Config
    p.versuchsplan_config = d.get("versuchsplan_config")
    p.optimierung_config = d.get("optimierung_config")

    return p


def _recompute_derived(projekt: Projekt):
    """Berechnet abgeleitete Felder (Modell, Statistiken) aus gespeicherten Rohdaten neu."""
    # Baseline-Statistiken
    if len(projekt.baseline_wuerfe) > 0:
        try:
            projekt.baseline_stats = analysiere_baseline(projekt.baseline_wuerfe)
        except Exception as e:
            print(f"⚠️ Baseline-Statistiken konnten nicht berechnet werden: {e}")

    # Versuchsplan
    _fak_doe = _effektive_faktoren(projekt)
    if _fak_doe and projekt.versuchsplan_config:
        try:
            cfg = projekt.versuchsplan_config
            projekt.versuchsplan = generiere_versuchsplan(
                _fak_doe,
                wiederholungen=cfg.get("wiederholungen", 3),
                blocking=cfg.get("blocking", False),
                centerpoints=cfg.get("centerpoints", 3),
                seed=projekt.seed,
                design=cfg.get("design", "voll"),
            )
        except Exception as e:
            print(f"⚠️ Versuchsplan konnte nicht regeneriert werden: {e}")

    # Regressionsmodell
    if projekt.doe_ergebnisse is not None and _fak_doe:
        try:
            projekt.modell = fitte_modell(projekt.doe_ergebnisse, _fak_doe)
            projekt.modell_gepruned, projekt.pruning_log = hierarchisches_pruning(projekt.modell)
            projekt.modell = projekt.modell_gepruned
        except Exception as e:
            print(f"⚠️ Modell konnte nicht neu berechnet werden: {e}")

    # Konfirmation
    if len(projekt.konfirmation_wuerfe) > 0 and projekt.optimale_einstellung:
        try:
            opt = projekt.optimale_einstellung
            projekt.konfirmation_ergebnis = analysiere_konfirmation(
                projekt.konfirmation_wuerfe,
                vorhersage=opt.get("vorhersage", 0),
                pi_low=opt.get("pi_low", 0),
                pi_high=opt.get("pi_high", 1000),
                zielweite=projekt.zielweite,
                toleranz=projekt.toleranz,
            )
        except Exception as e:
            print(f"⚠️ Konfirmationsanalyse konnte nicht berechnet werden: {e}")

    # I-MR
    if len(projekt.konfirmation_wuerfe) > 0:
        try:
            projekt.imr_ergebnis = berechne_imr(projekt.konfirmation_wuerfe)
        except Exception as e:
            print(f"⚠️ I-MR konnte nicht berechnet werden: {e}")

    # Cpk
    if len(projekt.konfirmation_wuerfe) > 0:
        try:
            usl = projekt.zielweite + projekt.toleranz
            lsl = projekt.zielweite - projekt.toleranz
            projekt.cpk_ergebnis = berechne_cpk(projekt.konfirmation_wuerfe, usl, lsl)
        except Exception as e:
            print(f"⚠️ Cpk konnte nicht berechnet werden: {e}")


def speichere_fortschritt(projekt: Projekt):
    """Speichert den Projektfortschritt als JSON (Google Drive oder lokal)."""
    try:
        save_dir = _fortschritt_verzeichnis(projekt)
        os.makedirs(save_dir, exist_ok=True)
        pfad = os.path.join(save_dir, "fortschritt.json")

        d = _projekt_to_dict(projekt)

        # Atomar schreiben (temp → rename)
        fd, tmp = tempfile.mkstemp(dir=save_dir, suffix=".json")
        try:
            with os.fdopen(fd, "w") as f:
                json.dump(d, f, ensure_ascii=False, indent=2)
            os.replace(tmp, pfad)
        except Exception:
            if os.path.exists(tmp):
                os.remove(tmp)
            raise

        print(f"💾 Fortschritt gespeichert ({d['_phase']})")
    except Exception as e:
        print(f"⚠️ Speichern fehlgeschlagen: {e}")


def lade_fortschritt(gruppenname: str, gruppennummer: int) -> Optional[Projekt]:
    """Lädt gespeicherten Fortschritt und berechnet abgeleitete Felder neu."""
    ordner = f"{gruppenname}_{gruppennummer}"

    # Drive oder lokal suchen
    for base in (_DRIVE_BASE, _LOCAL_BASE):
        pfad = os.path.join(base, ordner, "fortschritt.json")
        if os.path.exists(pfad):
            break
    else:
        return None

    try:
        with open(pfad) as f:
            d = json.load(f)

        projekt = _dict_to_projekt(d)

        # Abgeleitete Felder neu berechnen
        _recompute_derived(projekt)

        # Zusammenfassung
        saved_at = d.get("_saved_at", "unbekannt")
        phase = d.get("_phase", "unbekannt")
        teile = []
        if len(projekt.testwuerfe) > 0:
            teile.append(f"Testwürfe: {len(projekt.testwuerfe)}")
        if len(projekt.baseline_wuerfe) > 0:
            teile.append(f"Baseline: {len(projekt.baseline_wuerfe)}")
        if projekt.doe_ergebnisse is not None:
            teile.append(f"DoE: {len(projekt.doe_ergebnisse)} Runs")
        if len(projekt.konfirmation_wuerfe) > 0:
            teile.append(f"Konfirmation: {len(projekt.konfirmation_wuerfe)}")
        if projekt.modell is not None:
            teile.append(f"Modell R²={projekt.modell.rsquared:.3f}")

        try:
            ts = datetime.fromisoformat(saved_at)
            zeit_str = ts.strftime("%d.%m.%Y um %H:%M")
        except (ValueError, TypeError):
            zeit_str = saved_at

        return projekt

    except Exception as e:
        print(f"⚠️ Fortschritt konnte nicht geladen werden: {e}")
        return None


def finde_speicherstaende() -> List[Dict]:
    """Sucht im Drive und lokal nach vorhandenen DMAIC-Speicherständen."""
    ergebnisse = []
    for base in (_DRIVE_BASE, _LOCAL_BASE):
        if not os.path.isdir(base):
            continue
        for ordner in sorted(os.listdir(base)):
            pfad = os.path.join(base, ordner, "fortschritt.json")
            if not os.path.exists(pfad):
                continue
            try:
                with open(pfad) as f:
                    d = json.load(f)
                ergebnisse.append({
                    "pfad": pfad,
                    "gruppenname": d.get("gruppenname", "?"),
                    "gruppennummer": d.get("gruppennummer", 0),
                    "phase": d.get("_phase", "?"),
                    "saved_at": d.get("_saved_at", "?"),
                    "zielweite": d.get("zielweite", 0),
                })
            except Exception:
                continue
    return ergebnisse


def zeige_speicherstand_auswahl(staende: List[Dict]):
    """Zeigt eine nummerierte HTML-Tabelle aller Speicherstände."""
    if not staende:
        print("Keine Speicherstände gefunden.")
        return

    rows = ""
    for i, s in enumerate(staende, 1):
        try:
            ts = datetime.fromisoformat(s["saved_at"])
            zeit = ts.strftime("%d.%m.%Y %H:%M")
        except (ValueError, TypeError):
            zeit = s["saved_at"]
        rows += f"""<tr>
            <td style="padding:6px; text-align:center; font-weight:bold;">{i}</td>
            <td style="padding:6px;">{s['gruppenname']}</td>
            <td style="padding:6px; text-align:center;">{s['gruppennummer']}</td>
            <td style="padding:6px; text-align:center;">{s['phase']}</td>
            <td style="padding:6px; text-align:center;">{s['zielweite']:.0f} cm</td>
            <td style="padding:6px;">{zeit}</td>
        </tr>"""

    html = f"""
    <div style="padding:12px; border:2px solid {BLUE}; border-radius:8px;
                background:{LIGHT_BLUE}; margin:10px 0;">
        <h3 style="margin:0 0 8px 0;">📂 Vorhandene Speicherstände</h3>
        <table style="border-collapse:collapse; width:100%; background:white; border-radius:4px;">
            <tr style="background:{BLUE}; color:white;">
                <th style="padding:8px;">Nr.</th>
                <th style="padding:8px;">Gruppe</th>
                <th style="padding:8px;">Nr.</th>
                <th style="padding:8px;">Phase</th>
                <th style="padding:8px;">Zielweite</th>
                <th style="padding:8px;">Gespeichert</th>
            </tr>
            {rows}
        </table>
    </div>"""
    display(HTML(html))


def zeige_restore_zusammenfassung(projekt: Projekt):
    """Zeigt eine übersichtliche HTML-Zusammenfassung des geladenen Fortschritts."""
    phase = _aktuelle_phase(projekt)

    # Phasen-Status-Leiste
    phasen = ["DEFINE", "MEASURE", "ANALYZE", "IMPROVE", "CONTROL"]
    phasen_map = {p: i for i, p in enumerate(phasen)}
    aktiv_idx = phasen_map.get(phase, -1)

    status_dots = ""
    for i, p in enumerate(phasen):
        if i < aktiv_idx:
            farbe = GREEN
            symbol = "✅"
        elif i == aktiv_idx:
            farbe = BLUE
            symbol = "🔵"
        else:
            farbe = GRAY
            symbol = "⚪"
        status_dots += f'<span style="margin:0 6px;">{symbol} {p}</span>'

    # Detail-Zeilen sammeln
    details = []
    details.append(f"Zielweite: {projekt.zielweite:.0f} cm "
                   f"(Toleranz ±{projekt.toleranz:.0f} cm)")
    if len(projekt.vermessung_min_wuerfe) > 0 and len(projekt.vermessung_max_wuerfe) > 0:
        _min = float(np.mean(projekt.vermessung_min_wuerfe))
        _max = float(np.mean(projekt.vermessung_max_wuerfe))
        details.append(f"Katapult vermessen: {_min:.0f}–{_max:.0f} cm")
    if projekt.annaeherung_log:
        details.append(f"Annäherung: {len(projekt.annaeherung_log)} Iteration(en) protokolliert")
    if projekt.initiale_einstellung:
        details.append(f"Initiale Einstellung: {len(projekt.initiale_einstellung)} Faktor(en) festgelegt")
    if len(projekt.testwuerfe) > 0:
        details.append(f"Testwürfe: {len(projekt.testwuerfe)} Stück "
                       f"(μ={np.mean(projekt.testwuerfe):.1f} cm)")
    if projekt.charter and projekt.charter.get("Problemstellung"):
        details.append(f"Charter: ausgefüllt")
    if projekt.msa_type1:
        details.append(f"MSA Type-1: {len(projekt.msa_type1)} Person(en)")
    if projekt.msa_grr:
        details.append(f"Gage R&R: %GRR = {projekt.msa_grr.get('pct_grr', 0):.1f}%")
    if len(projekt.baseline_wuerfe) > 0:
        details.append(f"Baseline: {len(projekt.baseline_wuerfe)} Würfe "
                       f"(μ={np.mean(projekt.baseline_wuerfe):.1f}, "
                       f"σ={np.std(projekt.baseline_wuerfe, ddof=1):.1f})")
    if projekt.faktoren:
        namen = ", ".join(f["name"] for f in projekt.faktoren)
        details.append(f"Faktoren: {len(projekt.faktoren)} ({namen})")
    if projekt.doe_ergebnisse is not None:
        details.append(f"DoE: {len(projekt.doe_ergebnisse)} Runs")
    if projekt.modell is not None:
        details.append(f"Modell: R² = {projekt.modell.rsquared:.4f}")
    if projekt.optimale_einstellung:
        s = projekt.optimale_einstellung.get("strategie_label", "")
        details.append(f"Optimierung: {s}")
    if len(projekt.konfirmation_wuerfe) > 0:
        details.append(f"Konfirmation: {len(projekt.konfirmation_wuerfe)} Würfe")
    if projekt.cpk_ergebnis:
        details.append(f"Cpk = {projekt.cpk_ergebnis['cpk']:.2f}")

    details_html = "".join(f"<li>{d}</li>" for d in details)

    # Nächste Aktion
    if phase == "START":
        naechstes = "Führt die Zellen der DEFINE-Phase aus (Testwürfe, Charter)."
    elif phase == "DEFINE":
        naechstes = "Weiter mit <strong>MEASURE</strong>: MSA durchführen und Baseline-Würfe eingeben."
    elif phase == "MEASURE":
        naechstes = "Weiter mit <strong>ANALYZE</strong>: Faktoren für das DoE verfeinern und Versuchsplan erstellen."
    elif phase == "ANALYZE":
        naechstes = "Weiter mit <strong>IMPROVE</strong>: Konturplot analysieren und Einstellungen optimieren."
    elif phase == "IMPROVE":
        naechstes = "Weiter mit <strong>CONTROL</strong>: Konfirmation auswerten und Cpk berechnen."
    else:
        naechstes = "Alle Phasen abgeschlossen! Exportiert eure Ergebnisse als ZIP."

    html = f"""
    <div style="padding:16px; border:2px solid {BLUE}; border-radius:8px;
                background:{LIGHT_BLUE}; margin:10px 0;">
        <h3 style="margin:0 0 8px 0; color:{BLUE};">🔄 Fortschritt wiederhergestellt</h3>
        <div style="margin:8px 0; font-size:1.1em;">{status_dots}</div>
        <hr style="border:none; border-top:1px solid {BLUE}40; margin:10px 0;">
        <ul style="margin:4px 0; padding-left:20px; line-height:1.6;">{details_html}</ul>
        <hr style="border:none; border-top:1px solid {BLUE}40; margin:10px 0;">
        <div style="color:{BLUE};">
            <strong>Nächster Schritt:</strong> {naechstes}
        </div>
    </div>"""
    display(HTML(html))
