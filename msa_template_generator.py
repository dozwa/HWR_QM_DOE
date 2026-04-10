#!/usr/bin/env python3
"""
MSA-Template-Generator für den DMAIC Katapult-Versuch.
Erzeugt MSA_Messung_Template.xlsx mit zwei Blättern:
  - Type-1: Messmittelfähigkeit
  - Reproduzierbarkeit: Gekreuzte Gage R&R

Nutzung:
  - Standalone:  python msa_template_generator.py
  - Import:      from msa_template_generator import generate_msa_template
  - Google Colab: from msa_template_generator import generate_for_colab
"""

import os
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Styling-Konstanten
# ---------------------------------------------------------------------------

_HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_META_KEY_FONT = Font(name="Calibri", bold=True, size=11)
_META_VAL_FONT = Font(name="Calibri", size=11, italic=True, color="666666")
_NORMAL_FONT = Font(name="Calibri", size=11)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")


# ---------------------------------------------------------------------------
# Hilfsfunktionen
# ---------------------------------------------------------------------------


def _apply_header_style(cell):
    """Formatiert eine Zelle als Tabellen-Header."""
    cell.font = _HEADER_FONT
    cell.fill = _HEADER_FILL
    cell.alignment = _CENTER
    cell.border = _THIN_BORDER


def _apply_data_style(cell, alignment=None):
    """Formatiert eine normale Datenzelle."""
    cell.font = _NORMAL_FONT
    cell.border = _THIN_BORDER
    cell.alignment = alignment or _CENTER


def _apply_meta_key_style(cell):
    """Formatiert eine Metadaten-Beschriftung."""
    cell.font = _META_KEY_FONT
    cell.alignment = _LEFT


def _apply_meta_val_style(cell):
    """Formatiert ein Metadaten-Eingabefeld."""
    cell.font = _META_VAL_FONT
    cell.alignment = _LEFT
    cell.border = Border(bottom=Side(style="thin", color="AAAAAA"))


def _write_metadata_block(ws, gruppenname, messmodus, start_row=1):
    """
    Schreibt den Metadaten-Kopfblock (Gruppe, Datum, Messmittel, Messmodus)
    und gibt die nächste verfügbare Zeile zurück.
    """
    meta = [
        ("Gruppe:", gruppenname if gruppenname else "<Gruppenname eintragen>"),
        ("Datum:", date.today().strftime("%d.%m.%Y")),
        ("Messmittel:", "<z.\u202fB. Maßband, Laser-Entfernungsmesser>"),
        ("Messmodus:", messmodus),
    ]

    row = start_row
    for key, value in meta:
        key_cell = ws.cell(row=row, column=1, value=key)
        _apply_meta_key_style(key_cell)
        val_cell = ws.cell(row=row, column=2, value=value)
        _apply_meta_val_style(val_cell)
        # Metadaten-Wert über mehrere Spalten zusammenführen
        ws.merge_cells(
            start_row=row, start_column=2,
            end_row=row, end_column=4,
        )
        row += 1

    # Leerzeile nach Metadaten
    row += 1
    return row


# ---------------------------------------------------------------------------
# Sheet-Builder
# ---------------------------------------------------------------------------


def _build_type1_sheet(wb, gruppenname, num_personen, num_messungen, messmodus):
    """
    Erstellt das Blatt 'Type-1' für die Messmittelfähigkeit.

    Spalten:
      Messung-Nr. | Referenzwert | Person 1 | Person 2 | … | Person n
    """
    ws = wb.active
    ws.title = "Type-1"

    # -- Metadaten --
    data_start = _write_metadata_block(ws, gruppenname, messmodus)

    # -- Hinweiszeile --
    hint_cell = ws.cell(
        row=data_start,
        column=1,
        value=(
            "Anleitung: Alle Personen messen unabhängig denselben "
            "Referenzpunkt (z.\u202fB. Klebestreifen auf dem Boden). "
            "Referenzwert = bekannte Distanz in cm."
        ),
    )
    hint_cell.font = Font(name="Calibri", size=10, italic=True, color="888888")
    ws.merge_cells(
        start_row=data_start, start_column=1,
        end_row=data_start, end_column=2 + num_personen,
    )
    data_start += 2  # Leerzeile nach Hinweis

    # -- Header-Zeile --
    header_row = data_start
    headers = ["Messung-Nr.", "Referenzwert (cm)"]
    for p in range(1, num_personen + 1):
        headers.append(f"Person {p} (cm)")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        _apply_header_style(cell)

    # -- Datenzeilen --
    for i in range(1, num_messungen + 1):
        row = header_row + i
        # Messung-Nr.
        nr_cell = ws.cell(row=row, column=1, value=i)
        _apply_data_style(nr_cell)

        # Referenzwert – Platzhalter (Studierende füllen den tatsächlichen Wert ein)
        ref_cell = ws.cell(row=row, column=2, value="<Ref>")
        _apply_data_style(ref_cell)
        ref_cell.font = Font(name="Calibri", size=11, italic=True, color="999999")

        # Messwerte pro Person – leer lassen
        for p in range(num_personen):
            empty_cell = ws.cell(row=row, column=3 + p, value=None)
            _apply_data_style(empty_cell)

    # -- Spaltenbreiten --
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 20
    for p in range(num_personen):
        col_letter = get_column_letter(3 + p)
        ws.column_dimensions[col_letter].width = 18

    # Druckbereich-Info
    ws.sheet_properties.pageSetUpPr = None  # openpyxl default


def _build_reproduzierbarkeit_sheet(
    wb, gruppenname, num_personen, num_wuerfe, messmodus
):
    """
    Erstellt das Blatt 'Reproduzierbarkeit' für die gekreuzte Gage R&R.

    Spalten:
      Wurf-ID | Messung Person 1 | Messung Person 2 | … | Messung Person n
    """
    ws = wb.create_sheet(title="Reproduzierbarkeit")

    # -- Metadaten --
    data_start = _write_metadata_block(ws, gruppenname, messmodus)

    # -- Hinweiszeile --
    hint_cell = ws.cell(
        row=data_start,
        column=1,
        value=(
            "Anleitung: Nach jedem Wurf den Auftrittspunkt markieren "
            "(Kreppband). Jede Person misst unabhängig die Distanz – "
            "ohne Absprache. Erst danach Markierung entfernen."
        ),
    )
    hint_cell.font = Font(name="Calibri", size=10, italic=True, color="888888")
    ws.merge_cells(
        start_row=data_start, start_column=1,
        end_row=data_start, end_column=1 + num_personen,
    )
    data_start += 2  # Leerzeile nach Hinweis

    # -- Header-Zeile --
    header_row = data_start
    headers = ["Wurf-ID"]
    for p in range(1, num_personen + 1):
        headers.append(f"Messung Person {p} (cm)")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        _apply_header_style(cell)

    # -- Datenzeilen --
    for i in range(1, num_wuerfe + 1):
        row = header_row + i
        # Wurf-ID
        id_cell = ws.cell(row=row, column=1, value=i)
        _apply_data_style(id_cell)

        # Messwerte pro Person – leer lassen
        for p in range(num_personen):
            empty_cell = ws.cell(row=row, column=2 + p, value=None)
            _apply_data_style(empty_cell)

    # -- Spaltenbreiten --
    ws.column_dimensions["A"].width = 14
    for p in range(num_personen):
        col_letter = get_column_letter(2 + p)
        ws.column_dimensions[col_letter].width = 24


# ---------------------------------------------------------------------------
# Öffentliche API
# ---------------------------------------------------------------------------


def generate_msa_template(
    gruppenname="",
    num_personen=3,
    num_messungen=10,
    num_wuerfe=10,
    messmodus="1D",
    output_path=None,
):
    """
    Erzeugt die Datei MSA_Messung_Template.xlsx.

    Parameter
    ---------
    gruppenname : str
        Name der Studierendengruppe (wird in Metadaten eingetragen).
    num_personen : int
        Anzahl der messenden Personen (mindestens 2, Standard 3).
    num_messungen : int
        Anzahl der Wiederholungsmessungen im Type-1-Blatt (Standard 10).
    num_wuerfe : int
        Anzahl der Würfe im Reproduzierbarkeits-Blatt (Standard 10).
    messmodus : str
        '1D' (nur Weite) oder '2D' (Weite + Querversatz).
    output_path : str oder None
        Vollständiger Dateipfad für die Ausgabe.  Wenn None, wird
        'MSA_Messung_Template.xlsx' im aktuellen Verzeichnis erzeugt.

    Rückgabe
    --------
    str – Absoluter Pfad der erzeugten Datei.
    """
    # --- Validierung ---
    if num_personen < 2:
        raise ValueError("Es werden mindestens 2 Personen benötigt.")
    if num_messungen < 1:
        raise ValueError("Mindestens 1 Messung erforderlich.")
    if num_wuerfe < 1:
        raise ValueError("Mindestens 1 Wurf erforderlich.")
    messmodus = messmodus.upper()
    if messmodus not in ("1D", "2D"):
        raise ValueError("Messmodus muss '1D' oder '2D' sein.")

    # --- Workbook erstellen ---
    wb = Workbook()

    _build_type1_sheet(wb, gruppenname, num_personen, num_messungen, messmodus)
    _build_reproduzierbarkeit_sheet(
        wb, gruppenname, num_personen, num_wuerfe, messmodus
    )

    # --- Speichern ---
    if output_path is None:
        output_path = os.path.join(os.getcwd(), "MSA_Messung_Template.xlsx")
    else:
        output_path = os.path.abspath(output_path)

    # Sicherstellen, dass das Zielverzeichnis existiert
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    wb.save(output_path)
    print(f"MSA-Template gespeichert: {output_path}")
    return output_path


def generate_for_colab(gruppenname, num_personen=3, messmodus="1D"):
    """
    Erzeugt das MSA-Template und löst einen Download in Google Colab aus.

    Nutzung in einer Colab-Zelle:
        from msa_template_generator import generate_for_colab
        generate_for_colab("Gruppe A", num_personen=4, messmodus="1D")

    Parameter
    ---------
    gruppenname : str
        Name der Studierendengruppe.
    num_personen : int
        Anzahl der messenden Personen (Standard 3).
    messmodus : str
        '1D' oder '2D'.

    Rückgabe
    --------
    str – Pfad der erzeugten Datei (im Colab-Dateisystem).
    """
    filename = "MSA_Messung_Template.xlsx"
    filepath = generate_msa_template(
        gruppenname=gruppenname,
        num_personen=num_personen,
        messmodus=messmodus,
        output_path=filename,
    )

    # Colab-Download auslösen (funktioniert nur innerhalb von Google Colab)
    try:
        from google.colab import files as colab_files  # type: ignore
        colab_files.download(filepath)
    except ImportError:
        print(
            "Hinweis: google.colab nicht verfügbar – "
            "Datei wurde lokal gespeichert, aber kein automatischer "
            "Download ausgelöst."
        )

    return filepath


# ---------------------------------------------------------------------------
# Standalone-Ausführung
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    generate_msa_template()
