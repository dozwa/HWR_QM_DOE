"""DMAIC phase MEASURE: cells 17..31 of the notebook."""
from __future__ import annotations

from .cells import code, colab_code, md, phase_export_cell

_MEASURE_17_X = r"""---
# Phase 2: MEASURE

## Können wir unserer Messung vertrauen?

Bevor ihr optimiert, müsst ihr sicherstellen, dass euer **Messsystem zuverlässig** ist. Sonst könnten Unterschiede in euren Daten vom Messen kommen – nicht vom Katapult.

**Zwei Übungen:**
1. **Type-1 Studie:** Wie gut ist euer Maßband? (Wiederholbarkeit)
2. **Reproduzierbarkeitsstudie:** Messen alle gleich?"""

_MEASURE_18_MESSSYSTEMANALYSE_MSA_ERINNERU = r"""### Messsystemanalyse (MSA) – Erinnerung

Die MSA prüft zwei Dinge:
- **Wiederholbarkeit (Repeatability):** Selbe Person, selbe Messung → gleiches Ergebnis?
- **Reproduzierbarkeit (Reproducibility):** Verschiedene Personen → gleiches Ergebnis?

> Euer Messergebnis ist nicht die Wahrheit, sondern eine **Schätzung**. Die MSA prüft, wie gut diese Schätzung ist.

**Varianzzerlegung:**
$$\sigma^2_{total} = \sigma^2_{Teil} + \sigma^2_{Mess}$$
$$\sigma^2_{Mess} = \sigma^2_{Wiederhol} + \sigma^2_{Reprod}$$"""

_MEASURE_19_TITLE_REFERENZ_ACCURACY_VS_PRE = r"""fig = helper.plot_4_zielscheiben_referenz()
helper._save_fig(projekt, fig, "measure_zielscheiben_referenz")
plt.show()"""

_MEASURE_20_MSA_DURCHF_HRUNG = r"""### MSA-Durchführung

#### Übung A: Type-1 (Messmittelfähigkeit)
1. Markiert einen **festen Punkt** auf dem Boden (Klebeband bei z.B. 400 cm)
2. **Jede Person** misst diesen Punkt **10× unabhängig** mit dem Maßband
3. **Verdeckt notieren** – ohne die Werte der anderen zu sehen!

#### Übung B: Reproduzierbarkeit
1. Macht **10 Würfe** mit identischen Katapult-Einstellungen
2. Nach jedem Wurf: **Auftreffpunkt sofort markieren** (Kreppband)
3. **Jede Person** misst unabhängig die Distanz zum Auftreffpunkt
4. **Keine Kommunikation** während der Messung!
5. Erst alle Messungen notiert → dann Markierung entfernen → nächster Wurf"""

_MEASURE_21_TITLE_MSA_TEMPLATE_HERUNTERLAD = r"""anzahl_personen = 3 #@param {type:"integer"}

filepath = helper.generate_msa_template(
    gruppenname=projekt.gruppenname,
    num_personen=anzahl_personen,
    messmodus=projekt.messmodus,
)
print(f"✅ Template erstellt: {filepath}")

# Kopie im Drive-Ordner ablegen
import shutil
_save_dir = helper._fortschritt_verzeichnis(projekt)
if _save_dir:
    os.makedirs(_save_dir, exist_ok=True)
    shutil.copy2(filepath, os.path.join(_save_dir, os.path.basename(filepath)))

try:
    from google.colab import files
    files.download(filepath)
    print("📥 Download gestartet!")
except ImportError:
    print(f"📁 Datei gespeichert unter: {filepath}")
"""

_MEASURE_22_MSA_DATEN_HOCHLADEN = r"""### MSA-Daten hochladen

Füllt das Excel-Template aus und ladet es hier hoch. Das Notebook wertet eure MSA automatisch aus."""

_MEASURE_23_TITLE_MSA_DATEN_HOCHLADEN_UND = r"""try:
    from google.colab import files
    print("⬆️ Bitte MSA_Messung_Template.xlsx (ausgefüllt) hochladen:")
    uploaded = files.upload()
    msa_file = list(uploaded.keys())[0]
except ImportError:
    msa_file = "MSA_Messung_Template.xlsx"

# Type-1 auslesen
import openpyxl
wb = openpyxl.load_workbook(msa_file, data_only=True)
ws_type1 = wb["Type-1"]

# Header in Zeile 8, Daten ab Zeile 9
type1_data = {}
for row in ws_type1.iter_rows(min_row=9, max_row=18, values_only=True):
    if row[0] is not None:
        nr = row[0]
        ref = row[1] if isinstance(row[1], (int, float)) else 400
        for i, val in enumerate(row[2:], start=1):
            if val is not None and isinstance(val, (int, float)):
                if f"Person {i}" not in type1_data:
                    type1_data[f"Person {i}"] = []
                type1_data[f"Person {i}"].append(float(val))

if type1_data:
    type1_df = pd.DataFrame(type1_data)
    referenzwert = ref if isinstance(ref, (int, float)) else 400.0
    type1_ergebnis = helper.analysiere_type1(type1_df, referenzwert)
    projekt.msa_type1 = type1_ergebnis
    helper.zeige_type1(type1_ergebnis)
    print(f"\n✅ Type-1 Analyse abgeschlossen (Referenzwert: {referenzwert} cm)")
else:
    print("⚠️ Keine Type-1 Daten gefunden. Bitte Excel prüfen.")

# Excel-Backup auf Drive speichern
import shutil
_save_dir = helper._fortschritt_verzeichnis(projekt)
if _save_dir:
    os.makedirs(_save_dir, exist_ok=True)
    shutil.copy2(msa_file, os.path.join(_save_dir, "MSA_Messung_Template.xlsx"))

helper.speichere_fortschritt(projekt)
"""

_MEASURE_24_TITLE_GAGE_R_R_ANALYSE_ANOVA_M = r"""try:
    ws_repr = wb["Reproduzierbarkeit"]
    
    # Header in Zeile 8, Daten ab Zeile 9
    grr_records = []
    for row in ws_repr.iter_rows(min_row=9, max_row=18, values_only=True):
        if row[0] is not None and isinstance(row[0], (int, float)):
            wurf_id = row[0]
            for i, val in enumerate(row[1:], start=1):
                if val is not None and isinstance(val, (int, float)):
                    grr_records.append({
                        "Wurf_ID": int(wurf_id),
                        "Person": f"Person {i}",
                        "Messwert": float(val),
                    })
    
    if grr_records:
        grr_df = pd.DataFrame(grr_records)
        projekt.msa_rohdaten = grr_df
        grr_ergebnis = helper.analysiere_gage_rr(grr_df)
        projekt.msa_grr = grr_ergebnis
        helper.zeige_gage_rr(grr_ergebnis)
        helper.speichere_fortschritt(projekt)
    else:
        print("⚠️ Keine Reproduzierbarkeitsdaten gefunden.")
except Exception as e:
    print(f"❌ Fehler bei der Gage R&R Analyse: {e}")
    print("Prüft eure MSA-Excel: Sind Daten im Blatt 'Reproduzierbarkeit' vorhanden?")
"""

_MEASURE_25_TITLE_MSA_VISUALISIERUNGEN = r"""if projekt.msa_rohdaten is not None:
    grr_df = projekt.msa_rohdaten

    # Boxplot
    fig1 = helper.plot_msa_boxplot(grr_df)
    helper._save_fig(projekt, fig1, "measure_msa_boxplot")
    plt.show()

    # Interaktionsplot
    fig2 = helper.plot_msa_interaktion(grr_df)
    helper._save_fig(projekt, fig2, "measure_msa_interaktion")
    plt.show()

    # Zielscheiben pro Messer
    fig3 = helper.plot_msa_zielscheiben_pro_messer(grr_df)
    helper._save_fig(projekt, fig3, "measure_msa_zielscheiben_pro_messer")
    plt.show()
else:
    print("ℹ️ Keine MSA-Rohdaten vorhanden. Bitte zuerst Gage R&R durchführen.")
"""

_MEASURE_26_PROZESS_BASELINE = r"""### Prozess-Baseline

Jetzt erhebt ihr euren **Ist-Zustand**: Wie gut trifft euer Katapult aktuell?

1. Stellt euer Katapult auf die **initiale Einstellung** (aus DEFINE)
2. Macht **mindestens 10, gern bis zu 20 Würfe** (mehr Würfe = aussagekräftigere Baseline)
3. Messt jede Wurfweite und tragt sie unten ein

> Die Baseline ist euer Referenzpunkt. Am Ende des Tages vergleichen wir: **Baseline vs. Konfirmation**."""

_MEASURE_27_TITLE_BASELINE_W_RFE_EINGEBEN = r"""wurf_01 = 0.0 #@param {type:"number"}
wurf_02 = 0.0 #@param {type:"number"}
wurf_03 = 0.0 #@param {type:"number"}
wurf_04 = 0.0 #@param {type:"number"}
wurf_05 = 0.0 #@param {type:"number"}
wurf_06 = 0.0 #@param {type:"number"}
wurf_07 = 0.0 #@param {type:"number"}
wurf_08 = 0.0 #@param {type:"number"}
wurf_09 = 0.0 #@param {type:"number"}
wurf_10 = 0.0 #@param {type:"number"}
wurf_11 = 0.0 #@param {type:"number"}
wurf_12 = 0.0 #@param {type:"number"}
wurf_13 = 0.0 #@param {type:"number"}
wurf_14 = 0.0 #@param {type:"number"}
wurf_15 = 0.0 #@param {type:"number"}
wurf_16 = 0.0 #@param {type:"number"}
wurf_17 = 0.0 #@param {type:"number"}
wurf_18 = 0.0 #@param {type:"number"}
wurf_19 = 0.0 #@param {type:"number"}
wurf_20 = 0.0 #@param {type:"number"}

_alle = [wurf_01, wurf_02, wurf_03, wurf_04, wurf_05,
         wurf_06, wurf_07, wurf_08, wurf_09, wurf_10,
         wurf_11, wurf_12, wurf_13, wurf_14, wurf_15,
         wurf_16, wurf_17, wurf_18, wurf_19, wurf_20]
werte = [round(w, 1) for w in _alle if w > 0]

if werte:
    projekt.baseline_wuerfe = np.array(werte)
    print(f"✅ {len(werte)} Baseline-Würfe eingetragen")
    print(f"   Werte: {werte}")
    helper.speichere_fortschritt(projekt)
elif len(projekt.baseline_wuerfe) > 0:
    print(f"ℹ️ {len(projekt.baseline_wuerfe)} Baseline-Würfe aus gespeichertem Fortschritt geladen.")
else:
    print("⚠️ Bitte mindestens 10 Baseline-Würfe eingeben (Weite in cm mit 1 Nachkommastelle, z.B. 345.2).")
"""

_MEASURE_28_TITLE_BASELINE_AUSWERTUNG = r"""if len(projekt.baseline_wuerfe) > 0:
    projekt.baseline_stats = helper.analysiere_baseline(projekt.baseline_wuerfe)
    b = projekt.baseline_stats

    # Histogramm
    fig = helper.plot_baseline_histogramm(
        projekt.baseline_wuerfe, projekt.zielweite, projekt.toleranz
    )
    helper._save_fig(projekt, fig, "measure_baseline_histogramm")
    plt.show()

    # Shapiro-Wilk Bewertung
    if not np.isnan(b['shapiro_p']):
        shapiro_schwellen = [
            (0.05, "⚠️", "Abweichung von Normalverteilung – Daten und Histogramm prüfen"),
            (float('inf'), "✅", "Normalverteilung kann angenommen werden"),
        ]
        helper.zeige_ampel(b['shapiro_p'], shapiro_schwellen,
                          titel="Shapiro-Wilk p-Wert:")

    # Zielscheibe
    fig2 = helper.plot_zielscheibe(
        projekt.baseline_wuerfe, projekt.zielweite, projekt.toleranz,
        modus=projekt.messmodus, titel="Baseline – Ist-Zustand"
    )
    helper._save_fig(projekt, fig2, "measure_baseline_zielscheibe")
    plt.show()

    helper.hinweis_bericht("Baseline-Histogramm und Zielscheibe sind die Ausgangsbasis für die Bewertung des Verbesserungserfolgs.")
else:
    print("⚠️ Bitte zuerst Baseline-Würfe eingeben!")"""

_MEASURE_29_DIV_STYLE_PADDING_10PX_BORDER = r"""<div style="padding:10px; border-left:4px solid #2563EB; background:#DBEAFE; border-radius:4px;">
📋 <strong>Für den Bericht:</strong> MSA-Ergebnisse (Type-1, Gage R&R) und Baseline-Histogramm sind die zentralen Outputs der Measure-Phase.
</div>"""

_MEASURE_30_DETAILS_STYLE_MARGIN_10PX_0_PA = r"""<details style="margin:10px 0; padding:8px; background:#F9FAFB; border:1px solid #E5E7EB; border-radius:6px;">
<summary style="cursor:pointer; font-weight:bold; color:#2563EB;">
🔍 Für Neugierige: Warum brauchen wir eine Messsystemanalyse?
</summary>
<div style="margin-top:8px; padding:8px; font-size:0.95em;">

Stellt euch vor, euer Katapult wirft tatsächlich immer auf 450 cm – aber euer Messsystem schwankt um ±20 cm. Dann sieht es so aus, als wäre das Katapult schlecht, obwohl nur die Messung schlecht ist.

Die MSA trennt die **Messunsicherheit** von der **echten Prozessvariation**. Nur so könnt ihr sicher sein, dass Verbesserungen am Katapult auch wirklich Verbesserungen sind – und nicht nur Messartefakte.

**Faustregel (AIAG):**
- %GRR < 10%: Messsystem geeignet
- %GRR 10–30%: Bedingt geeignet
- %GRR > 30%: Nicht geeignet – zuerst Messung verbessern!
</div>
</details>"""


def cells():
    return [
        md(_MEASURE_17_X),
        md(_MEASURE_18_MESSSYSTEMANALYSE_MSA_ERINNERU),
        colab_code("📊 Referenz: Accuracy vs. Precision (4 Quadranten)", _MEASURE_19_TITLE_REFERENZ_ACCURACY_VS_PRE),
        md(_MEASURE_20_MSA_DURCHF_HRUNG),
        colab_code("📥 MSA-Template herunterladen", _MEASURE_21_TITLE_MSA_TEMPLATE_HERUNTERLAD),
        md(_MEASURE_22_MSA_DATEN_HOCHLADEN),
        colab_code("⬆️ MSA-Daten hochladen und Type-1 auswerten", _MEASURE_23_TITLE_MSA_DATEN_HOCHLADEN_UND),
        colab_code("📊 Gage R&R Analyse (ANOVA-Methode, AIAG-konform)", _MEASURE_24_TITLE_GAGE_R_R_ANALYSE_ANOVA_M),
        colab_code("📊 MSA-Visualisierungen", _MEASURE_25_TITLE_MSA_VISUALISIERUNGEN),
        md(_MEASURE_26_PROZESS_BASELINE),
        colab_code("📝 Baseline-Würfe eingeben (Weite in cm, 1 Nachkommastelle)", _MEASURE_27_TITLE_BASELINE_W_RFE_EINGEBEN),
        colab_code("📊 Baseline-Auswertung", _MEASURE_28_TITLE_BASELINE_AUSWERTUNG),
        md(_MEASURE_29_DIV_STYLE_PADDING_10PX_BORDER),
        md(_MEASURE_30_DETAILS_STYLE_MARGIN_10PX_0_PA),
        phase_export_cell("MEASURE"),
    ]
